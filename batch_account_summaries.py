"""Resumable overnight account-summary processor.

Each Excel file is loaded through app_account.py and its existing account
summary download route. The resulting three sheets are stored in SQLite.
"""

from __future__ import annotations

import argparse
import concurrent.futures
import gc
import hashlib
import json
import logging
import os
import posixpath
import re
import sqlite3
import sys
import time
import xml.etree.ElementTree as ET
import zipfile
from contextlib import contextmanager, redirect_stderr, redirect_stdout
from io import BytesIO, StringIO
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

import app_account
from summary_database import (
    DEFAULT_DATABASE_PATH,
    connect_database,
    initialize_database,
    mark_file_failed,
    save_file_summaries,
    set_worker_state,
    utc_now,
)


LOGGER = logging.getLogger("account_summary_batch")
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
FAST_DUPLICATE_AUDIT_VERSION = (
    "ack-credit-tid-no-leading-zero-credit-account-last4-no-other-bank-v3"
)
DUPLICATE_SUMMARY_VERSION = (
    "ack-credit-tid-no-leading-zero-credit-account-last4-no-other-bank-v6"
)
CELL_COLUMN_PATTERN = re.compile(r"([A-Za-z]+)")


def configure_logging(database_path: Path, verbose: bool = False) -> None:
    database_path.parent.mkdir(parents=True, exist_ok=True)
    log_path = database_path.parent / "account_summary_worker.log"
    LOGGER.setLevel(logging.DEBUG if verbose else logging.INFO)
    LOGGER.handlers.clear()
    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    console = logging.StreamHandler()
    console.setFormatter(formatter)
    console.setLevel(logging.DEBUG if verbose else logging.INFO)
    LOGGER.addHandler(console)
    file_handler = RotatingFileHandler(
        log_path,
        maxBytes=10 * 1024 * 1024,
        backupCount=3,
        encoding="utf-8",
    )
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.DEBUG)
    LOGGER.addHandler(file_handler)


class SingleWorkerLock:
    """Hold a non-blocking process lock next to the SQLite database."""

    def __init__(self, database_path: Path) -> None:
        self.path = database_path.with_suffix(database_path.suffix + ".worker.lock")
        self.handle: Any = None

    def __enter__(self) -> "SingleWorkerLock":
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.handle = self.path.open("a+b")
        self.handle.seek(0, os.SEEK_END)
        if self.handle.tell() == 0:
            self.handle.write(b"0")
            self.handle.flush()
        self.handle.seek(0)
        try:
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(self.handle.fileno(), msvcrt.LK_NBLCK, 1)
            else:
                import fcntl

                fcntl.flock(self.handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        except OSError as exc:
            self.handle.close()
            raise RuntimeError(
                "Another account-summary worker is already using this database."
            ) from exc
        return self

    def __exit__(self, exc_type: Any, exc: Any, traceback: Any) -> None:
        if not self.handle:
            return
        try:
            self.handle.seek(0)
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(self.handle.fileno(), msvcrt.LK_UNLCK, 1)
            else:
                import fcntl

                fcntl.flock(self.handle.fileno(), fcntl.LOCK_UN)
        finally:
            self.handle.close()


@contextmanager
def keep_windows_awake(enabled: bool) -> Iterator[None]:
    if not enabled or os.name != "nt":
        yield
        return
    import ctypes

    es_continuous = 0x80000000
    es_system_required = 0x00000001
    ctypes.windll.kernel32.SetThreadExecutionState(
        es_continuous | es_system_required
    )
    try:
        yield
    finally:
        ctypes.windll.kernel32.SetThreadExecutionState(es_continuous)


def iter_excel_files(input_directory: Path) -> Iterator[Path]:
    for root, directory_names, file_names in os.walk(input_directory):
        directory_names[:] = sorted(
            name
            for name in directory_names
            if not name.startswith(".")
        )
        for file_name in sorted(file_names):
            if file_name.startswith("~$"):
                continue
            path = Path(root) / file_name
            if path.suffix.lower() in SUPPORTED_EXTENSIONS:
                yield path.resolve()


def calculate_file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        shared_strings = ET.parse(archive.open("xl/sharedStrings.xml")).getroot()
    except KeyError:
        return []
    values: list[str] = []
    for item in shared_strings:
        if _xml_local_name(item.tag) != "si":
            continue
        values.append(
            "".join(
                node.text or ""
                for node in item.iter()
                if _xml_local_name(node.tag) == "t"
            )
        )
    return values


def _money_transfer_sheet_path(archive: zipfile.ZipFile) -> str:
    workbook = ET.parse(archive.open("xl/workbook.xml")).getroot()
    relationship_id = None
    for node in workbook.iter():
        if (
            _xml_local_name(node.tag) == "sheet"
            and "money transfer" in node.attrib.get("name", "").casefold()
        ):
            relationship_id = next(
                (
                    value
                    for key, value in node.attrib.items()
                    if _xml_local_name(key) == "id"
                ),
                None,
            )
            break
    if not relationship_id:
        raise ValueError("Money Transfer sheet not found")

    relationships = ET.parse(
        archive.open("xl/_rels/workbook.xml.rels")
    ).getroot()
    target = None
    for node in relationships:
        if (
            _xml_local_name(node.tag) == "Relationship"
            and node.attrib.get("Id") == relationship_id
        ):
            target = node.attrib.get("Target")
            break
    if not target:
        raise ValueError("Money Transfer worksheet relationship is missing")
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def _xlsx_cell_value(
    cell: ET.Element,
    shared_strings: list[str],
) -> str:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        return "".join(
            node.text or ""
            for node in cell.iter()
            if _xml_local_name(node.tag) == "t"
        )
    raw_value = next(
        (
            node.text or ""
            for node in cell
            if _xml_local_name(node.tag) == "v"
        ),
        "",
    )
    if cell_type == "s" and raw_value:
        try:
            return shared_strings[int(raw_value)]
        except (IndexError, ValueError):
            return ""
    return raw_value


def audit_workbook_duplicate_keys(source_path: Path) -> tuple[int, int]:
    """Count repeated B/G/J credit keys, excluding E=Other/Others."""
    with zipfile.ZipFile(source_path) as archive:
        shared_strings = _read_shared_strings(archive)
        worksheet_path = _money_transfer_sheet_path(archive)
        seen: set[tuple[str, str, str]] = set()
        duplicates = 0
        data_rows = 0
        with archive.open(worksheet_path) as worksheet:
            for _event, row in ET.iterparse(worksheet, events=("end",)):
                if _xml_local_name(row.tag) != "row":
                    continue
                try:
                    row_number = int(row.attrib.get("r", "0") or 0)
                except ValueError:
                    row_number = 0
                if row_number <= 1:
                    row.clear()
                    continue
                selected: dict[str, str] = {}
                for cell in row:
                    if _xml_local_name(cell.tag) != "c":
                        continue
                    reference = cell.attrib.get("r", "")
                    match = CELL_COLUMN_PATTERN.match(reference)
                    if not match:
                        continue
                    column = match.group(1).upper()
                    if column in {"B", "E", "G", "J"}:
                        selected[column] = _xlsx_cell_value(
                            cell,
                            shared_strings,
                        )
                acknowledgement = app_account._identity_text(selected.get("B", ""))
                credited_account_last_four = app_account._account_last_four(
                    selected.get("G", "")
                )
                credited_transaction_id = app_account._credited_transaction_identity(
                    selected.get("J", "")
                )
                is_other_bank = app_account._is_money_transfer_to_others_bank(
                    selected.get("E", "")
                )
                if (
                    acknowledgement
                    or credited_account_last_four
                    or credited_transaction_id
                ):
                    data_rows += 1
                if (
                    not is_other_bank
                    and acknowledgement
                    and credited_account_last_four
                    and credited_transaction_id
                ):
                    key = (
                        acknowledgement,
                        credited_transaction_id,
                        credited_account_last_four,
                    )
                    if key in seen:
                        duplicates += 1
                    else:
                        seen.add(key)
                row.clear()
        return duplicates, data_rows


def fast_audit_and_queue_duplicate_files(
    database_path: Path,
    *,
    workers: int,
) -> dict[str, int]:
    """Audit ACK, bank, credited account, and TID columns efficiently."""
    connection = connect_database(database_path)
    try:
        candidates = [
            dict(row)
            for row in connection.execute(
                """
                SELECT id, source_path, fast_duplicate_audit_version,
                       fast_duplicate_rows_found, fast_duplicate_audit_error,
                       fast_duplicate_reprocessed_version
                FROM source_files
                WHERE duplicate_of_source_file_id IS NULL
                ORDER BY id
                """
            )
        ]
    finally:
        connection.close()

    counts = {
        "files": len(candidates),
        "scanned": 0,
        "cached": 0,
        "duplicate_files": 0,
        "duplicate_rows": 0,
        "errors": 0,
        "queued": 0,
    }
    files_to_scan: list[dict[str, Any]] = []
    queue_ids: set[int] = set()
    for candidate in candidates:
        if (
            candidate["fast_duplicate_audit_version"]
            == FAST_DUPLICATE_AUDIT_VERSION
            and not candidate["fast_duplicate_audit_error"]
        ):
            duplicate_rows = int(
                candidate["fast_duplicate_rows_found"] or 0
            )
            counts["cached"] += 1
            counts["duplicate_rows"] += duplicate_rows
            if (
                duplicate_rows
                and candidate["fast_duplicate_reprocessed_version"]
                != DUPLICATE_SUMMARY_VERSION
            ):
                counts["duplicate_files"] += 1
                queue_ids.add(int(candidate["id"]))
        else:
            files_to_scan.append(candidate)

    pending_updates: list[tuple[str, int, str | None, int]] = []

    def flush_updates() -> None:
        if not pending_updates:
            return
        update_connection = connect_database(database_path)
        try:
            with update_connection:
                update_connection.executemany(
                    """
                    UPDATE source_files
                    SET fast_duplicate_audit_version = ?,
                        fast_duplicate_rows_found = ?,
                        fast_duplicate_audit_error = ?
                    WHERE id = ?
                    """,
                    pending_updates,
                )
        finally:
            update_connection.close()
        pending_updates.clear()

    started = time.perf_counter()
    with concurrent.futures.ThreadPoolExecutor(
        max_workers=max(1, workers)
    ) as executor:
        futures = {
            executor.submit(
                audit_workbook_duplicate_keys,
                Path(candidate["source_path"]),
            ): candidate
            for candidate in files_to_scan
        }
        for future in concurrent.futures.as_completed(futures):
            candidate = futures[future]
            error_message = None
            try:
                duplicate_rows, _data_rows = future.result()
            except Exception as exc:
                duplicate_rows = 0
                error_message = f"{type(exc).__name__}: {exc}"[:1000]
                counts["errors"] += 1
                queue_ids.add(int(candidate["id"]))
            else:
                counts["duplicate_rows"] += duplicate_rows
                if (
                    duplicate_rows
                    and candidate["fast_duplicate_reprocessed_version"]
                    != DUPLICATE_SUMMARY_VERSION
                ):
                    counts["duplicate_files"] += 1
                    queue_ids.add(int(candidate["id"]))
            counts["scanned"] += 1
            pending_updates.append(
                (
                    FAST_DUPLICATE_AUDIT_VERSION,
                    duplicate_rows,
                    error_message,
                    int(candidate["id"]),
                )
            )
            if len(pending_updates) >= 250:
                flush_updates()
            if counts["scanned"] % 1000 == 0:
                elapsed = max(time.perf_counter() - started, 0.001)
                LOGGER.info(
                    "Fast duplicate audit: %s/%s scanned (%.0f files/sec), "
                    "%s affected, %s unreadable.",
                    counts["scanned"],
                    len(files_to_scan),
                    counts["scanned"] / elapsed,
                    counts["duplicate_files"],
                    counts["errors"],
                )
                set_worker_state(
                    database_path,
                    is_running=True,
                    process_id=os.getpid(),
                    current_file=None,
                    message=(
                        f"Fast duplicate audit: {counts['scanned']}/"
                        f"{len(files_to_scan)} files"
                    ),
                )
    flush_updates()

    queue_connection = connect_database(database_path)
    try:
        with queue_connection:
            for start in range(0, len(queue_ids), 500):
                chunk = list(queue_ids)[start : start + 500]
                placeholders = ",".join("?" for _ in chunk)
                cursor = queue_connection.execute(
                    f"""
                    UPDATE source_files
                    SET status = 'pending', attempts = 0, error_message = NULL
                    WHERE id IN ({placeholders})
                    """,
                    chunk,
                )
                counts["queued"] += cursor.rowcount
    finally:
        queue_connection.close()
    return counts


def _mark_duplicate_in_connection(
    connection: sqlite3.Connection,
    *,
    source_file_id: int,
    canonical_source_file_id: int,
    content_sha256: str,
    completed_at: str,
) -> None:
    for table_name in (
        "account_summaries",
        "bank_summaries",
        "partial_bank_summaries",
        "money_transfer_to_others_rows",
    ):
        connection.execute(
            f"DELETE FROM {table_name} WHERE source_file_id = ?",
            (source_file_id,),
        )
    connection.execute(
        """
        UPDATE source_files
        SET status = 'completed',
            completed_at = ?,
            duration_seconds = 0,
            acknowledgement_count = 0,
            account_row_count = 0,
            bank_row_count = 0,
            partial_bank_row_count = 0,
            money_transfer_other_row_count = 0,
            content_sha256 = ?,
            duplicate_of_source_file_id = ?,
            main_rows_read = 0,
            duplicate_transaction_rows_removed = 0,
            other_rows_read = 0,
            duplicate_other_rows_removed = 0,
            duplicate_summary_rows_removed = 0,
            error_message = NULL
        WHERE id = ?
        """,
        (
            completed_at,
            content_sha256,
            canonical_source_file_id,
            source_file_id,
        ),
    )


def discover_files(database_path: Path, input_directory: Path) -> dict[str, int]:
    """Register files and identify exact workbook copies by SHA-256."""
    connection = connect_database(database_path)
    now = utc_now()
    counts = {
        "seen": 0,
        "new": 0,
        "changed": 0,
        "unchanged": 0,
        "hashed": 0,
        "duplicate_files": 0,
    }
    try:
        existing = {
            row["source_path"].casefold(): dict(row)
            for row in connection.execute(
                """
                SELECT id, source_path, fingerprint, content_sha256, status,
                       duplicate_of_source_file_id
                FROM source_files
                """
            )
        }
        content_owners: dict[str, int] = {}
        with connection:
            for path in iter_excel_files(input_directory):
                try:
                    stat = path.stat()
                except OSError as exc:
                    LOGGER.warning("Could not inspect %s: %s", path, exc)
                    continue
                counts["seen"] += 1
                fingerprint = f"{stat.st_size}:{stat.st_mtime_ns}"
                existing_row = existing.get(str(path).casefold())
                unchanged = (
                    existing_row is not None
                    and existing_row["fingerprint"] == fingerprint
                )
                content_sha256 = (
                    existing_row.get("content_sha256") if unchanged else None
                )
                if not content_sha256:
                    content_sha256 = calculate_file_sha256(path)
                    counts["hashed"] += 1
                canonical_id = content_owners.get(content_sha256)

                if existing_row is None:
                    cursor = connection.execute(
                        """
                        INSERT INTO source_files (
                            source_path,
                            file_name,
                            file_size,
                            mtime_ns,
                            fingerprint,
                            content_sha256,
                            status,
                            discovered_at,
                            last_seen_at,
                            duplicate_of_source_file_id,
                            completed_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            str(path),
                            path.name,
                            stat.st_size,
                            stat.st_mtime_ns,
                            fingerprint,
                            content_sha256,
                            "completed" if canonical_id is not None else "pending",
                            now,
                            now,
                            canonical_id,
                            now if canonical_id is not None else None,
                        ),
                    )
                    new_id = cursor.lastrowid
                    if canonical_id is not None:
                        counts["duplicate_files"] += 1
                    counts["new"] += 1
                    # Pending files are checked again at claim time, after a
                    # prior identical file may have completed successfully.
                    if canonical_id is None:
                        existing[str(path).casefold()] = {
                            "id": new_id,
                            "fingerprint": fingerprint,
                            "content_sha256": content_sha256,
                            "status": "pending",
                            "duplicate_of_source_file_id": None,
                        }
                elif not unchanged:
                    if canonical_id is not None and canonical_id != existing_row["id"]:
                        _mark_duplicate_in_connection(
                            connection,
                            source_file_id=existing_row["id"],
                            canonical_source_file_id=canonical_id,
                            content_sha256=content_sha256,
                            completed_at=now,
                        )
                        connection.execute(
                            """
                            UPDATE source_files
                            SET file_name = ?, file_size = ?, mtime_ns = ?,
                                fingerprint = ?, last_seen_at = ?
                            WHERE id = ?
                            """,
                            (
                                path.name,
                                stat.st_size,
                                stat.st_mtime_ns,
                                fingerprint,
                                now,
                                existing_row["id"],
                            ),
                        )
                        counts["duplicate_files"] += 1
                    else:
                        connection.execute(
                            """
                            UPDATE source_files
                            SET file_name = ?,
                                file_size = ?,
                                mtime_ns = ?,
                                fingerprint = ?,
                                content_sha256 = ?,
                                duplicate_of_source_file_id = NULL,
                                status = 'pending',
                                attempts = 0,
                                fast_duplicate_audit_version = NULL,
                                fast_duplicate_rows_found = 0,
                                fast_duplicate_audit_error = NULL,
                                fast_duplicate_reprocessed_version = NULL,
                                last_seen_at = ?,
                                error_message = NULL
                            WHERE id = ?
                            """,
                            (
                                path.name,
                                stat.st_size,
                                stat.st_mtime_ns,
                                fingerprint,
                                content_sha256,
                                now,
                                existing_row["id"],
                            ),
                        )
                    counts["changed"] += 1
                else:
                    if canonical_id is not None and canonical_id != existing_row["id"]:
                        if existing_row.get("duplicate_of_source_file_id") != canonical_id:
                            _mark_duplicate_in_connection(
                                connection,
                                source_file_id=existing_row["id"],
                                canonical_source_file_id=canonical_id,
                                content_sha256=content_sha256,
                                completed_at=now,
                            )
                            counts["duplicate_files"] += 1
                    elif not existing_row.get("content_sha256"):
                        connection.execute(
                            """
                            UPDATE source_files
                            SET content_sha256 = ?, last_seen_at = ?
                            WHERE id = ?
                            """,
                            (content_sha256, now, existing_row["id"]),
                        )
                    counts["unchanged"] += 1

                current_id = (
                    existing_row["id"] if existing_row is not None else new_id
                )
                current_state = connection.execute(
                    """
                    SELECT status, duplicate_of_source_file_id
                    FROM source_files
                    WHERE id = ?
                    """,
                    (current_id,),
                ).fetchone()
                if (
                    current_state["status"] == "completed"
                    and current_state["duplicate_of_source_file_id"] is None
                ):
                    content_owners.setdefault(content_sha256, current_id)

        return counts
    finally:
        connection.close()


def reset_interrupted_files(database_path: Path) -> int:
    connection = connect_database(database_path)
    try:
        with connection:
            cursor = connection.execute(
                """
                UPDATE source_files
                SET status = 'pending',
                    error_message = COALESCE(
                        error_message,
                        'Previous worker stopped during processing; queued again.'
                    )
                WHERE status = 'processing'
                """
            )
        return cursor.rowcount
    finally:
        connection.close()


def reset_failed_files(database_path: Path) -> int:
    connection = connect_database(database_path)
    try:
        with connection:
            cursor = connection.execute(
                """
                UPDATE source_files
                SET status = 'pending', attempts = 0, error_message = NULL
                WHERE status = 'failed'
                """
            )
        return cursor.rowcount
    finally:
        connection.close()


def reset_all_files(database_path: Path) -> int:
    connection = connect_database(database_path)
    try:
        with connection:
            cursor = connection.execute(
                """
                UPDATE source_files
                SET status = 'pending', attempts = 0, error_message = NULL
                WHERE duplicate_of_source_file_id IS NULL
                """
            )
        return cursor.rowcount
    finally:
        connection.close()


def reset_files_with_detected_duplicates(database_path: Path) -> int:
    """Queue files where the previous report recorded duplicate transactions."""
    connection = connect_database(database_path)
    try:
        with connection:
            cursor = connection.execute(
                """
                UPDATE source_files
                SET status = 'pending',
                    attempts = 0,
                    error_message = NULL
                WHERE id IN (
                    SELECT DISTINCT source_file_id
                    FROM account_summaries
                    WHERE TRIM(COALESCE(duplicate_entry_info, ''))
                          NOT IN ('', 'None')
                    UNION
                    SELECT id
                    FROM source_files
                    WHERE duplicate_transaction_rows_removed > 0
                       OR duplicate_other_rows_removed > 0
                       OR duplicate_summary_rows_removed > 0
                )
                """
            )
        return cursor.rowcount
    finally:
        connection.close()


def claim_next_file(
    database_path: Path,
    *,
    maximum_attempts: int,
) -> sqlite3.Row | None:
    connection = connect_database(database_path)
    try:
        while True:
            connection.execute("BEGIN IMMEDIATE")
            row = connection.execute(
                """
                SELECT sf.id, sf.source_path, sf.file_name, sf.attempts,
                       sf.content_sha256
                FROM source_files sf
                WHERE (
                        sf.status = 'pending'
                        OR (sf.status = 'failed' AND sf.attempts < ?)
                    )
                  AND sf.duplicate_of_source_file_id IS NULL
                ORDER BY
                    CASE WHEN EXISTS (
                        SELECT 1
                        FROM account_summaries legacy_duplicate
                        WHERE legacy_duplicate.source_file_id = sf.id
                          AND TRIM(COALESCE(
                              legacy_duplicate.duplicate_entry_info, ''
                          )) NOT IN ('', 'None')
                    ) THEN 0 ELSE 1 END,
                    CASE sf.status WHEN 'pending' THEN 0 ELSE 1 END,
                    sf.id
                LIMIT 1
                """,
                (maximum_attempts,),
            ).fetchone()
            if row is None:
                connection.commit()
                return None

            if row["content_sha256"]:
                canonical = connection.execute(
                    """
                    SELECT id
                    FROM source_files
                    WHERE content_sha256 = ?
                      AND id <> ?
                      AND status = 'completed'
                      AND duplicate_of_source_file_id IS NULL
                    ORDER BY id
                    LIMIT 1
                    """,
                    (row["content_sha256"], row["id"]),
                ).fetchone()
                if canonical is not None:
                    _mark_duplicate_in_connection(
                        connection,
                        source_file_id=row["id"],
                        canonical_source_file_id=canonical["id"],
                        content_sha256=row["content_sha256"],
                        completed_at=utc_now(),
                    )
                    connection.commit()
                    LOGGER.info(
                        "Skipped exact duplicate workbook %s (source ID %s).",
                        row["file_name"],
                        canonical["id"],
                    )
                    continue

            connection.execute(
                """
                UPDATE source_files
                SET status = 'processing',
                    attempts = attempts + 1,
                    started_at = ?,
                    completed_at = NULL,
                    duration_seconds = NULL,
                    error_message = NULL
                WHERE id = ?
                """,
                (utc_now(), row["id"]),
            )
            connection.commit()
            return row
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def parse_summary_workbook(payload: bytes) -> dict[str, list[dict[str, Any]]]:
    expected_sheets = (
        "Account Wise Summary",
        "Bank Wise Summary",
        "Partial Bank Wise Summary",
        "Money Transfer to Others",
    )
    workbook = load_workbook(
        BytesIO(payload),
        read_only=True,
        data_only=True,
    )
    try:
        summaries: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in expected_sheets:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Generated summary is missing sheet: {sheet_name}"
                )
            worksheet = workbook[sheet_name]
            row_iterator = worksheet.iter_rows(values_only=True)
            headers = next(row_iterator, None)
            if not headers:
                summaries[sheet_name] = []
                continue
            normalized_headers = [str(value).strip() for value in headers]
            summaries[sheet_name] = [
                dict(zip(normalized_headers, values))
                for values in row_iterator
                if any(value is not None for value in values)
            ]
        return summaries
    finally:
        workbook.close()


def analyse_with_existing_app(
    source_path: Path,
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, Any]]:
    """Call app_account.py's existing processing and report-generation code."""
    captured_output = StringIO()
    with redirect_stdout(captured_output), redirect_stderr(captured_output):
        app_account.uploaded_files_count = 0
        success, message = app_account.process_excel_file(
            str(source_path),
            is_first_file=True,
        )
        if not success:
            raise ValueError(message)
        with app_account.app.test_request_context(
            "/download-account-summary"
        ):
            response = app_account.download_account_summary()
            if isinstance(response, tuple):
                response = response[0]
            mimetype = getattr(response, "mimetype", "")
            response.direct_passthrough = False
            payload = response.get_data()
            if mimetype == "application/json":
                try:
                    error_data = json.loads(payload.decode("utf-8"))
                except (UnicodeDecodeError, json.JSONDecodeError):
                    error_data = {"message": payload[:500].decode(errors="replace")}
                raise ValueError(
                    error_data.get("message")
                    or error_data.get("error")
                    or "Account summary generation failed."
                )
    LOGGER.debug("app_account.py output for %s:\n%s", source_path, captured_output.getvalue())
    audit = {
        key: int(value or 0)
        for key, value in app_account.last_processing_audit.items()
    }
    audit["duplicate_processing_version"] = DUPLICATE_SUMMARY_VERSION
    return parse_summary_workbook(payload), audit


def analyse_source_in_worker(
    source_path: str,
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, Any]]:
    """Process one source in an isolated child process."""
    try:
        return analyse_with_existing_app(Path(source_path))
    finally:
        release_app_memory()


def release_app_memory() -> None:
    app_account.df_main = None
    app_account.df_other_sheets = {}
    app_account.uploaded_files_count = 0
    app_account.debited_acc_map = {}
    app_account.credited_acc_map = {}
    app_account.debited_trans_id_map = {}
    app_account.credited_trans_id_map = {}
    app_account.breakdown_map = {}
    app_account.last_duplicate_transaction_details = []
    app_account.last_processing_audit = {
        'main_rows_read': 0,
        'main_rows_kept': 0,
        'duplicate_transaction_rows_removed': 0,
        'other_rows_read': 0,
        'other_rows_kept': 0,
        'duplicate_other_rows_removed': 0,
    }
    gc.collect()


def queue_counts(database_path: Path, maximum_attempts: int) -> dict[str, int]:
    connection = connect_database(database_path)
    try:
        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END)
                    AS completed,
                SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END)
                    AS failed,
                SUM(CASE
                    WHEN status = 'pending'
                      OR (status = 'failed' AND attempts < ?)
                    THEN 1 ELSE 0 END) AS actionable
            FROM source_files
            """,
            (maximum_attempts,),
        ).fetchone()
        return {key: int(row[key] or 0) for key in row.keys()}
    finally:
        connection.close()


def process_queue(
    database_path: Path,
    *,
    maximum_attempts: int,
    maximum_files: int | None,
    process_workers: int = 1,
) -> int:
    if process_workers > 1:
        return process_queue_parallel(
            database_path,
            maximum_attempts=maximum_attempts,
            maximum_files=maximum_files,
            process_workers=process_workers,
        )
    processed_this_run = 0
    while maximum_files is None or processed_this_run < maximum_files:
        source = claim_next_file(
            database_path,
            maximum_attempts=maximum_attempts,
        )
        if source is None:
            break
        source_path = Path(source["source_path"])
        started = time.perf_counter()
        set_worker_state(
            database_path,
            is_running=True,
            process_id=os.getpid(),
            current_file=source["file_name"],
            message="Analysing workbook",
        )
        try:
            if not source_path.is_file():
                raise FileNotFoundError(f"Source file no longer exists: {source_path}")
            summaries, processing_audit = analyse_with_existing_app(source_path)
            duration = time.perf_counter() - started
            counts = save_file_summaries(
                database_path,
                source["id"],
                summaries,
                duration_seconds=duration,
                processing_audit=processing_audit,
            )
            processed_this_run += 1
            progress = queue_counts(database_path, maximum_attempts)
            LOGGER.info(
                "[%s/%s] %s | %.2fs | %s ACK | %s account rows | "
                "%s duplicate credited amounts excluded",
                progress["completed"] + progress["failed"],
                progress["total"],
                source["file_name"],
                duration,
                counts["acknowledgements"],
                counts["account"],
                processing_audit["duplicate_transaction_rows_removed"],
            )
        except KeyboardInterrupt:
            duration = time.perf_counter() - started
            mark_file_failed(
                database_path,
                source["id"],
                "Worker stopped by user during this file; it can be retried.",
                duration_seconds=duration,
            )
            raise
        except Exception as exc:
            duration = time.perf_counter() - started
            mark_file_failed(
                database_path,
                source["id"],
                f"{type(exc).__name__}: {exc}",
                duration_seconds=duration,
            )
            processed_this_run += 1
            LOGGER.exception(
                "Failed %s after %.2fs; continuing with the next file.",
                source_path,
                duration,
            )
        finally:
            release_app_memory()
    return processed_this_run


def process_queue_parallel(
    database_path: Path,
    *,
    maximum_attempts: int,
    maximum_files: int | None,
    process_workers: int,
) -> int:
    """Analyse independent workbooks concurrently; write SQLite in the parent."""
    processed_this_run = 0
    claimed_this_run = 0
    pending: dict[concurrent.futures.Future, tuple[dict[str, Any], float]] = {}

    with concurrent.futures.ProcessPoolExecutor(
        max_workers=process_workers
    ) as executor:
        while True:
            while len(pending) < process_workers and (
                maximum_files is None or claimed_this_run < maximum_files
            ):
                source_row = claim_next_file(
                    database_path,
                    maximum_attempts=maximum_attempts,
                )
                if source_row is None:
                    break
                source = dict(source_row)
                future = executor.submit(
                    analyse_source_in_worker,
                    source["source_path"],
                )
                pending[future] = (source, time.perf_counter())
                claimed_this_run += 1

            if not pending:
                break

            set_worker_state(
                database_path,
                is_running=True,
                process_id=os.getpid(),
                current_file=None,
                message=f"Analysing {len(pending)} workbooks in parallel",
            )
            completed, _not_completed = concurrent.futures.wait(
                pending,
                return_when=concurrent.futures.FIRST_COMPLETED,
            )
            for future in completed:
                source, started = pending.pop(future)
                source_path = Path(source["source_path"])
                duration = time.perf_counter() - started
                try:
                    summaries, processing_audit = future.result()
                    counts = save_file_summaries(
                        database_path,
                        source["id"],
                        summaries,
                        duration_seconds=duration,
                        processing_audit=processing_audit,
                    )
                    progress = queue_counts(database_path, maximum_attempts)
                    LOGGER.info(
                        "[%s/%s] %s | %.2fs parallel | %s ACK | "
                        "%s account rows | %s duplicate credited amounts excluded",
                        progress["completed"] + progress["failed"],
                        progress["total"],
                        source["file_name"],
                        duration,
                        counts["acknowledgements"],
                        counts["account"],
                        processing_audit[
                            "duplicate_transaction_rows_removed"
                        ],
                    )
                except Exception as exc:
                    mark_file_failed(
                        database_path,
                        source["id"],
                        f"{type(exc).__name__}: {exc}",
                        duration_seconds=duration,
                    )
                    LOGGER.exception(
                        "Failed %s after %.2fs in parallel worker; continuing.",
                        source_path,
                        duration,
                    )
                processed_this_run += 1

    return processed_this_run


def build_argument_parser() -> argparse.ArgumentParser:
    default_input = Path(r"C:\Users\admin\Desktop\bank_trails")
    if not default_input.exists():
        default_input = Path(__file__).resolve().parent / "uploads_account"
    parser = argparse.ArgumentParser(
        description=(
            "Process Excel files one at a time through app_account.py and save "
            "the account summaries to a resumable SQLite database."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=default_input,
        help=f"Folder containing source workbooks (default: {default_input})",
    )
    parser.add_argument(
        "--database",
        type=Path,
        default=DEFAULT_DATABASE_PATH,
        help=f"SQLite database path (default: {DEFAULT_DATABASE_PATH})",
    )
    parser.add_argument(
        "--watch",
        action="store_true",
        help="Keep scanning for new files after the current queue is finished.",
    )
    parser.add_argument(
        "--scan-interval",
        type=int,
        default=60,
        help="Seconds between scans in watch mode (default: 60).",
    )
    parser.add_argument(
        "--max-attempts",
        type=int,
        default=2,
        help="Maximum attempts for a failing file (default: 2).",
    )
    parser.add_argument(
        "--max-files",
        type=int,
        help="Stop after this many files; useful for a test run.",
    )
    parser.add_argument(
        "--retry-failed",
        action="store_true",
        help="Queue all failed files again and reset their attempt count.",
    )
    parser.add_argument(
        "--reprocess-all",
        action="store_true",
        help="Queue every discovered file again without deleting old data first.",
    )
    parser.add_argument(
        "--reprocess-detected-duplicates",
        action="store_true",
        help=(
            "Queue files whose previous account summary reported duplicate "
            "transactions."
        ),
    )
    parser.add_argument(
        "--fast-reprocess-duplicates",
        action="store_true",
        help=(
            "Read only ACK, bank, credited transaction ID, and credited-account "
            "columns, then queue only files containing duplicate keys."
        ),
    )
    parser.add_argument(
        "--audit-workers",
        type=int,
        default=min(16, max(4, (os.cpu_count() or 4) * 2)),
        help="Parallel workers for the fast duplicate audit (default: %(default)s).",
    )
    parser.add_argument(
        "--process-workers",
        type=int,
        default=min(4, max(1, (os.cpu_count() or 2) // 2)),
        help=(
            "Independent workbook analysis processes (default: %(default)s; "
            "use 1 for sequential processing)."
        ),
    )
    parser.add_argument(
        "--keep-awake",
        action="store_true",
        help="Prevent Windows system sleep while the worker is running.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Write the existing app's debug output to the worker log.",
    )
    return parser


def main() -> int:
    args = build_argument_parser().parse_args()
    input_directory = args.input.expanduser().resolve()
    database_path = args.database.expanduser().resolve()
    if not input_directory.is_dir():
        print(f"Input folder does not exist: {input_directory}", file=sys.stderr)
        return 2
    if args.max_attempts < 1:
        print("--max-attempts must be at least 1.", file=sys.stderr)
        return 2
    if args.audit_workers < 1:
        print("--audit-workers must be at least 1.", file=sys.stderr)
        return 2
    if args.process_workers < 1:
        print("--process-workers must be at least 1.", file=sys.stderr)
        return 2
    reprocess_modes = sum(
        bool(value)
        for value in (
            args.reprocess_all,
            args.reprocess_detected_duplicates,
            args.fast_reprocess_duplicates,
        )
    )
    if reprocess_modes > 1:
        print(
            "Choose only one reprocessing mode.",
            file=sys.stderr,
        )
        return 2

    initialize_database(database_path)
    configure_logging(database_path, args.verbose)

    try:
        with SingleWorkerLock(database_path), keep_windows_awake(args.keep_awake):
            interrupted = reset_interrupted_files(database_path)
            if interrupted:
                LOGGER.warning("Re-queued %s interrupted file(s).", interrupted)
            if args.retry_failed:
                LOGGER.info(
                    "Re-queued %s failed file(s).",
                    reset_failed_files(database_path),
                )

            discovery = discover_files(database_path, input_directory)
            LOGGER.info(
                "Scan complete: %s Excel files, %s new, %s changed, "
                "%s hashed, %s exact copies skipped.",
                discovery["seen"],
                discovery["new"],
                discovery["changed"],
                discovery["hashed"],
                discovery["duplicate_files"],
            )
            if args.reprocess_all:
                LOGGER.info(
                    "Queued %s file(s) for full reprocessing.",
                    reset_all_files(database_path),
                )
            elif args.reprocess_detected_duplicates:
                LOGGER.info(
                    "Queued %s file(s) with previously detected duplicates.",
                    reset_files_with_detected_duplicates(database_path),
                )
            elif args.fast_reprocess_duplicates:
                set_worker_state(
                    database_path,
                    is_running=True,
                    process_id=os.getpid(),
                    current_file=None,
                    message="Starting fast three-column duplicate audit",
                    started_at=utc_now(),
                )
                audit = fast_audit_and_queue_duplicate_files(
                    database_path,
                    workers=args.audit_workers,
                )
                LOGGER.info(
                    "Fast audit complete: %s scanned, %s cached, "
                    "%s affected files (%s duplicate rows), %s unreadable, "
                    "%s queued for full summary recalculation.",
                    audit["scanned"],
                    audit["cached"],
                    audit["duplicate_files"],
                    audit["duplicate_rows"],
                    audit["errors"],
                    audit["queued"],
                )

            set_worker_state(
                database_path,
                is_running=True,
                process_id=os.getpid(),
                current_file=None,
                message="Worker started",
                started_at=utc_now(),
            )
            total_processed = 0
            while True:
                remaining_limit = (
                    None
                    if args.max_files is None
                    else max(0, args.max_files - total_processed)
                )
                if remaining_limit == 0:
                    break
                total_processed += process_queue(
                    database_path,
                    maximum_attempts=args.max_attempts,
                    maximum_files=remaining_limit,
                    process_workers=args.process_workers,
                )
                if args.max_files is not None and total_processed >= args.max_files:
                    break
                if not args.watch:
                    break
                set_worker_state(
                    database_path,
                    is_running=True,
                    process_id=os.getpid(),
                    current_file=None,
                    message=f"Queue complete; next scan in {args.scan_interval}s",
                )
                LOGGER.info(
                    "Queue complete. Watching for new files every %s seconds.",
                    args.scan_interval,
                )
                time.sleep(max(5, args.scan_interval))
                discovery = discover_files(database_path, input_directory)
                if (
                    discovery["new"]
                    or discovery["changed"]
                    or discovery["duplicate_files"]
                ):
                    LOGGER.info(
                        "New scan: %s new, %s changed, %s exact copies skipped.",
                        discovery["new"],
                        discovery["changed"],
                        discovery["duplicate_files"],
                    )

            final_counts = queue_counts(database_path, args.max_attempts)
            LOGGER.info(
                "Worker stopped cleanly. Completed: %s, failed: %s, total: %s.",
                final_counts["completed"],
                final_counts["failed"],
                final_counts["total"],
            )
            set_worker_state(
                database_path,
                is_running=False,
                process_id=None,
                current_file=None,
                message="Worker stopped cleanly",
            )
            return 0
    except KeyboardInterrupt:
        LOGGER.warning("Worker stopped by user. Progress is saved.")
        set_worker_state(
            database_path,
            is_running=False,
            process_id=None,
            current_file=None,
            message="Stopped by user; progress is saved",
        )
        return 130
    except RuntimeError as exc:
        LOGGER.error("%s", exc)
        return 1
    except Exception:
        LOGGER.exception("Worker stopped because of an unexpected error.")
        set_worker_state(
            database_path,
            is_running=False,
            process_id=None,
            current_file=None,
            message="Worker stopped after an unexpected error",
        )
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
