"""SQLite storage and Excel export helpers for account summaries."""

from __future__ import annotations

import json
import math
import sqlite3
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATABASE_PATH = BASE_DIR / "data" / "account_summaries.sqlite"

ACCOUNT_COLUMNS = (
    ("Acknowledgement No", "acknowledgement_no"),
    ("Bank Name", "bank_name"),
    ("Account Number", "account_number"),
    ("Credited Transaction ID", "credited_transaction_id"),
    ("Total Credited Amount", "total_credited_amount"),
    ("Total Debited Amount", "total_debited_amount"),
    ("Updated Amount (Recovery)", "updated_amount"),
    ("Not Updated Amount", "not_updated_amount"),
    ("Status", "status"),
    ("Found in Other Sheets", "found_in_other_sheets"),
    ("Breakdown by Sheet", "breakdown_by_sheet"),
    ("Duplicate Entry Info", "duplicate_entry_info"),
)

BANK_COLUMNS = (
    ("Acknowledgement No", "acknowledgement_no"),
    ("Bank Name", "bank_name"),
    ("Total Credited Amount", "total_credited_amount"),
    ("Total Debited Amount", "total_debited_amount"),
    ("Updated Amount (Recovery)", "updated_amount"),
    ("Not Updated Amount", "not_updated_amount"),
    ("Status", "status"),
    ("Found in Other Sheets", "found_in_other_sheets"),
    ("Breakdown by Sheet", "breakdown_by_sheet"),
    ("Duplicate Entry Info", "duplicate_entry_info"),
)

VIEW_CONFIG = {
    "account": {
        "table": "account_summaries",
        "sheet": "Account Wise Summary",
        "columns": ACCOUNT_COLUMNS,
        "search_columns": (
            "acknowledgement_no",
            "bank_name",
            "account_number",
            "credited_transaction_id",
        ),
        "amount_columns": {
            "total_credited_amount",
            "total_debited_amount",
            "updated_amount",
            "not_updated_amount",
        },
        "column_widths": (35, 30, 25, 24, 22, 22, 22, 22, 15, 20, 60, 60),
    },
    "bank": {
        "table": "bank_summaries",
        "sheet": "Bank Wise Summary",
        "columns": BANK_COLUMNS,
        "search_columns": ("acknowledgement_no", "bank_name"),
        "amount_columns": {
            "total_credited_amount",
            "total_debited_amount",
            "updated_amount",
            "not_updated_amount",
        },
        "column_widths": (35, 30, 22, 22, 22, 22, 15, 20, 60, 60),
    },
    "partial": {
        "table": "partial_bank_summaries",
        "sheet": "Partial Bank Wise Summary",
        "columns": BANK_COLUMNS,
        "search_columns": ("acknowledgement_no", "bank_name"),
        "amount_columns": {
            "total_credited_amount",
            "total_debited_amount",
            "updated_amount",
            "not_updated_amount",
        },
        "column_widths": (35, 30, 22, 22, 22, 22, 15, 20, 60, 60),
    },
}

SOURCE_TO_DATABASE_COLUMNS = {
    "Acknowledgement No": "acknowledgement_no",
    "Bank Name": "bank_name",
    "Account Number": "account_number",
    "Credited Transaction ID": "credited_transaction_id",
    "Total Credited Amount": "total_credited_amount",
    "Total Debited Amount": "total_debited_amount",
    "Updated Amount (Recovery)": "updated_amount",
    "Not Updated Amount": "not_updated_amount",
    "Status": "status",
    "Found in Other Sheets": "found_in_other_sheets",
    "Breakdown by Sheet": "breakdown_by_sheet",
    "Duplicate Entry Info": "duplicate_entry_info",
}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def connect_database(
    database_path: str | Path = DEFAULT_DATABASE_PATH,
    *,
    readonly: bool = False,
) -> sqlite3.Connection:
    path = Path(database_path).expanduser().resolve()
    if not readonly:
        path.parent.mkdir(parents=True, exist_ok=True)
        connection = sqlite3.connect(str(path), timeout=30, check_same_thread=False)
    else:
        connection = sqlite3.connect(
            f"file:{path.as_posix()}?mode=ro",
            uri=True,
            timeout=30,
            check_same_thread=False,
        )
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    connection.execute("PRAGMA busy_timeout = 30000")
    if not readonly:
        connection.execute("PRAGMA journal_mode = WAL")
        connection.execute("PRAGMA synchronous = NORMAL")
    return connection


def _ensure_column(
    connection: sqlite3.Connection,
    table_name: str,
    column_name: str,
    declaration: str,
) -> None:
    columns = {
        row["name"]
        for row in connection.execute(f"PRAGMA table_info({table_name})")
    }
    if column_name not in columns:
        connection.execute(
            f"ALTER TABLE {table_name} ADD COLUMN {column_name} {declaration}"
        )


def _apply_schema_migrations(connection: sqlite3.Connection) -> None:
    """Add strict deduplication safeguards to new and existing databases."""
    source_columns = {
        "content_sha256": "TEXT",
        "duplicate_of_source_file_id": (
            "INTEGER REFERENCES source_files(id) ON DELETE SET NULL"
        ),
        "main_rows_read": "INTEGER NOT NULL DEFAULT 0",
        "duplicate_transaction_rows_removed": "INTEGER NOT NULL DEFAULT 0",
        "other_rows_read": "INTEGER NOT NULL DEFAULT 0",
        "duplicate_other_rows_removed": "INTEGER NOT NULL DEFAULT 0",
        "duplicate_summary_rows_removed": "INTEGER NOT NULL DEFAULT 0",
        "fast_duplicate_audit_version": "TEXT",
        "fast_duplicate_rows_found": "INTEGER NOT NULL DEFAULT 0",
        "fast_duplicate_audit_error": "TEXT",
        "fast_duplicate_reprocessed_version": "TEXT",
        "money_transfer_other_row_count": "INTEGER NOT NULL DEFAULT 0",
    }
    for column_name, declaration in source_columns.items():
        _ensure_column(
            connection,
            "source_files",
            column_name,
            declaration,
        )

    version_row = connection.execute(
        "SELECT value FROM schema_metadata WHERE key = 'schema_version'"
    ).fetchone()
    try:
        schema_version = int(version_row["value"]) if version_row else 0
    except (TypeError, ValueError):
        schema_version = 0

    if schema_version < 3:
        _remove_existing_exact_summary_duplicates(connection)
        # Version 2 briefly used an account-level key that was too broad for
        # accounts with multiple legitimate transaction groups.  Replace it
        # with the complete summary identity used by the writer below.
        connection.executescript(
            """
            DROP INDEX IF EXISTS uq_account_summary_ack_account;
            DROP INDEX IF EXISTS uq_bank_summary_ack_bank;
            DROP INDEX IF EXISTS uq_partial_summary_ack_bank;
            """
        )

    # A logical summary row may exist only once.  These keys reflect the
    # grouping performed by app_account.py and protect the database even if a
    # caller bypasses the overnight worker.
    connection.executescript(
        """
        CREATE INDEX IF NOT EXISTS idx_source_content_sha256
            ON source_files(content_sha256);
        CREATE INDEX IF NOT EXISTS idx_source_duplicate_of
            ON source_files(duplicate_of_source_file_id);

        CREATE UNIQUE INDEX IF NOT EXISTS uq_account_summary_identity
            ON account_summaries(
                TRIM(acknowledgement_no) COLLATE NOCASE,
                TRIM(COALESCE(bank_name, '')) COLLATE NOCASE,
                TRIM(COALESCE(account_number, '')) COLLATE NOCASE,
                TRIM(COALESCE(credited_transaction_id, '')) COLLATE NOCASE,
                total_credited_amount,
                total_debited_amount,
                updated_amount,
                not_updated_amount,
                TRIM(COALESCE(status, '')) COLLATE NOCASE
            );
        CREATE UNIQUE INDEX IF NOT EXISTS uq_bank_summary_identity
            ON bank_summaries(
                TRIM(acknowledgement_no) COLLATE NOCASE,
                TRIM(COALESCE(bank_name, '')) COLLATE NOCASE
            );
        CREATE UNIQUE INDEX IF NOT EXISTS uq_partial_summary_identity
            ON partial_bank_summaries(
                TRIM(acknowledgement_no) COLLATE NOCASE,
                TRIM(COALESCE(bank_name, '')) COLLATE NOCASE
            );

        CREATE TRIGGER IF NOT EXISTS trg_one_source_per_ack
        BEFORE INSERT ON account_summaries
        WHEN EXISTS (
            SELECT 1
            FROM account_summaries existing
            WHERE existing.acknowledgement_no = NEW.acknowledgement_no
                  COLLATE NOCASE
              AND existing.source_file_id <> NEW.source_file_id
        )
        BEGIN
            SELECT RAISE(
                ABORT,
                'Acknowledgement already belongs to another source file'
            );
        END;
        """
    )
    connection.execute(
        """
        INSERT INTO schema_metadata(key, value)
        VALUES ('schema_version', '4')
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """
    )


def _remove_existing_exact_summary_duplicates(
    connection: sqlite3.Connection,
) -> None:
    """Remove only provably identical legacy summary rows before constraints."""
    configurations = (
        (
            "account_summaries",
            "account_row_count",
            (
                "TRIM(COALESCE(acknowledgement_no, '')) COLLATE NOCASE",
                "TRIM(COALESCE(bank_name, '')) COLLATE NOCASE",
                "TRIM(COALESCE(account_number, '')) COLLATE NOCASE",
                "TRIM(COALESCE(credited_transaction_id, '')) COLLATE NOCASE",
                "total_credited_amount",
                "total_debited_amount",
                "updated_amount",
                "not_updated_amount",
                "TRIM(COALESCE(status, '')) COLLATE NOCASE",
                "TRIM(COALESCE(found_in_other_sheets, '')) COLLATE NOCASE",
                "TRIM(COALESCE(breakdown_by_sheet, '')) COLLATE NOCASE",
                "TRIM(COALESCE(duplicate_entry_info, '')) COLLATE NOCASE",
            ),
        ),
        (
            "bank_summaries",
            "bank_row_count",
            (
                "TRIM(COALESCE(acknowledgement_no, '')) COLLATE NOCASE",
                "TRIM(COALESCE(bank_name, '')) COLLATE NOCASE",
                "total_credited_amount",
                "total_debited_amount",
                "updated_amount",
                "not_updated_amount",
                "TRIM(COALESCE(status, '')) COLLATE NOCASE",
                "TRIM(COALESCE(found_in_other_sheets, '')) COLLATE NOCASE",
                "TRIM(COALESCE(breakdown_by_sheet, '')) COLLATE NOCASE",
                "TRIM(COALESCE(duplicate_entry_info, '')) COLLATE NOCASE",
            ),
        ),
        (
            "partial_bank_summaries",
            "partial_bank_row_count",
            (
                "TRIM(COALESCE(acknowledgement_no, '')) COLLATE NOCASE",
                "TRIM(COALESCE(bank_name, '')) COLLATE NOCASE",
                "total_credited_amount",
                "total_debited_amount",
                "updated_amount",
                "not_updated_amount",
                "TRIM(COALESCE(status, '')) COLLATE NOCASE",
                "TRIM(COALESCE(found_in_other_sheets, '')) COLLATE NOCASE",
                "TRIM(COALESCE(breakdown_by_sheet, '')) COLLATE NOCASE",
                "TRIM(COALESCE(duplicate_entry_info, '')) COLLATE NOCASE",
            ),
        ),
    )
    removed_by_source: dict[int, int] = {}
    affected_sources: set[int] = set()
    for table_name, _count_column, identity_columns in configurations:
        partition = ", ".join(identity_columns)
        duplicate_rows = connection.execute(
            f"""
            SELECT id, source_file_id
            FROM (
                SELECT id,
                       source_file_id,
                       ROW_NUMBER() OVER (
                           PARTITION BY {partition}
                           ORDER BY id
                       ) AS duplicate_number
                FROM {table_name}
            )
            WHERE duplicate_number > 1
            """
        ).fetchall()
        if not duplicate_rows:
            continue
        connection.executemany(
            f"DELETE FROM {table_name} WHERE id = ?",
            ((row["id"],) for row in duplicate_rows),
        )
        for row in duplicate_rows:
            source_file_id = int(row["source_file_id"])
            affected_sources.add(source_file_id)
            removed_by_source[source_file_id] = (
                removed_by_source.get(source_file_id, 0) + 1
            )

    for source_file_id, removed_count in removed_by_source.items():
        connection.execute(
            """
            UPDATE source_files
            SET duplicate_summary_rows_removed =
                    duplicate_summary_rows_removed + ?
            WHERE id = ?
            """,
            (removed_count, source_file_id),
        )
    for source_file_id in affected_sources:
        connection.execute(
            """
            UPDATE source_files
            SET acknowledgement_count = (
                    SELECT COUNT(DISTINCT acknowledgement_no)
                    FROM account_summaries
                    WHERE source_file_id = source_files.id
                ),
                account_row_count = (
                    SELECT COUNT(*) FROM account_summaries
                    WHERE source_file_id = source_files.id
                ),
                bank_row_count = (
                    SELECT COUNT(*) FROM bank_summaries
                    WHERE source_file_id = source_files.id
                ),
                partial_bank_row_count = (
                    SELECT COUNT(*) FROM partial_bank_summaries
                    WHERE source_file_id = source_files.id
                )
            WHERE id = ?
            """,
            (source_file_id,),
        )


def initialize_database(
    database_path: str | Path = DEFAULT_DATABASE_PATH,
) -> Path:
    path = Path(database_path).expanduser().resolve()
    connection = connect_database(path)
    try:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS schema_metadata (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS source_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_path TEXT NOT NULL COLLATE NOCASE UNIQUE,
                file_name TEXT NOT NULL,
                file_size INTEGER NOT NULL,
                mtime_ns INTEGER NOT NULL,
                fingerprint TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending'
                    CHECK (status IN ('pending', 'processing', 'completed', 'failed')),
                attempts INTEGER NOT NULL DEFAULT 0,
                discovered_at TEXT NOT NULL,
                last_seen_at TEXT NOT NULL,
                started_at TEXT,
                completed_at TEXT,
                duration_seconds REAL,
                acknowledgement_count INTEGER NOT NULL DEFAULT 0,
                account_row_count INTEGER NOT NULL DEFAULT 0,
                bank_row_count INTEGER NOT NULL DEFAULT 0,
                partial_bank_row_count INTEGER NOT NULL DEFAULT 0,
                money_transfer_other_row_count INTEGER NOT NULL DEFAULT 0,
                content_sha256 TEXT,
                duplicate_of_source_file_id INTEGER
                    REFERENCES source_files(id) ON DELETE SET NULL,
                main_rows_read INTEGER NOT NULL DEFAULT 0,
                duplicate_transaction_rows_removed INTEGER NOT NULL DEFAULT 0,
                other_rows_read INTEGER NOT NULL DEFAULT 0,
                duplicate_other_rows_removed INTEGER NOT NULL DEFAULT 0,
                duplicate_summary_rows_removed INTEGER NOT NULL DEFAULT 0,
                fast_duplicate_audit_version TEXT,
                fast_duplicate_rows_found INTEGER NOT NULL DEFAULT 0,
                fast_duplicate_audit_error TEXT,
                fast_duplicate_reprocessed_version TEXT,
                error_message TEXT
            );

            CREATE TABLE IF NOT EXISTS account_summaries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file_id INTEGER NOT NULL
                    REFERENCES source_files(id) ON DELETE CASCADE,
                acknowledgement_no TEXT NOT NULL COLLATE NOCASE,
                bank_name TEXT,
                account_number TEXT,
                credited_transaction_id TEXT,
                total_credited_amount REAL NOT NULL DEFAULT 0,
                total_debited_amount REAL NOT NULL DEFAULT 0,
                updated_amount REAL NOT NULL DEFAULT 0,
                not_updated_amount REAL NOT NULL DEFAULT 0,
                status TEXT,
                found_in_other_sheets TEXT,
                breakdown_by_sheet TEXT,
                duplicate_entry_info TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS bank_summaries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file_id INTEGER NOT NULL
                    REFERENCES source_files(id) ON DELETE CASCADE,
                acknowledgement_no TEXT NOT NULL COLLATE NOCASE,
                bank_name TEXT,
                total_credited_amount REAL NOT NULL DEFAULT 0,
                total_debited_amount REAL NOT NULL DEFAULT 0,
                updated_amount REAL NOT NULL DEFAULT 0,
                not_updated_amount REAL NOT NULL DEFAULT 0,
                status TEXT,
                found_in_other_sheets TEXT,
                breakdown_by_sheet TEXT,
                duplicate_entry_info TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS partial_bank_summaries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file_id INTEGER NOT NULL
                    REFERENCES source_files(id) ON DELETE CASCADE,
                acknowledgement_no TEXT NOT NULL COLLATE NOCASE,
                bank_name TEXT,
                total_credited_amount REAL NOT NULL DEFAULT 0,
                total_debited_amount REAL NOT NULL DEFAULT 0,
                updated_amount REAL NOT NULL DEFAULT 0,
                not_updated_amount REAL NOT NULL DEFAULT 0,
                status TEXT,
                found_in_other_sheets TEXT,
                breakdown_by_sheet TEXT,
                duplicate_entry_info TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS money_transfer_to_others_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file_id INTEGER NOT NULL
                    REFERENCES source_files(id) ON DELETE CASCADE,
                acknowledgement_no TEXT COLLATE NOCASE,
                source_row_number INTEGER NOT NULL,
                row_json TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS worker_state (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                is_running INTEGER NOT NULL DEFAULT 0,
                process_id INTEGER,
                started_at TEXT,
                heartbeat_at TEXT,
                current_file TEXT,
                message TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_source_files_status
                ON source_files(status, attempts, id);
            CREATE INDEX IF NOT EXISTS idx_account_ack
                ON account_summaries(acknowledgement_no);
            CREATE INDEX IF NOT EXISTS idx_account_source_file
                ON account_summaries(source_file_id);
            CREATE INDEX IF NOT EXISTS idx_account_status
                ON account_summaries(status);
            CREATE INDEX IF NOT EXISTS idx_account_bank
                ON account_summaries(bank_name);
            CREATE INDEX IF NOT EXISTS idx_account_number
                ON account_summaries(account_number);
            CREATE INDEX IF NOT EXISTS idx_bank_ack
                ON bank_summaries(acknowledgement_no);
            CREATE INDEX IF NOT EXISTS idx_bank_source_file
                ON bank_summaries(source_file_id);
            CREATE INDEX IF NOT EXISTS idx_bank_status
                ON bank_summaries(status);
            CREATE INDEX IF NOT EXISTS idx_partial_ack
                ON partial_bank_summaries(acknowledgement_no);
            CREATE INDEX IF NOT EXISTS idx_partial_source_file
                ON partial_bank_summaries(source_file_id);
            CREATE INDEX IF NOT EXISTS idx_money_transfer_other_source
                ON money_transfer_to_others_rows(source_file_id, source_row_number);
            CREATE INDEX IF NOT EXISTS idx_money_transfer_other_ack
                ON money_transfer_to_others_rows(acknowledgement_no);

            INSERT OR IGNORE INTO worker_state(id, is_running)
            VALUES (1, 0);
            """
        )
        _apply_schema_migrations(connection)
        connection.commit()
    finally:
        connection.close()
    return path


def _clean_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, float) and math.isnan(value):
        return default
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return default
    return text


def _raw_acknowledgement(record: Mapping[str, Any]) -> str:
    """Read the ACK from a raw Money Transfer row despite punctuation."""
    for header, value in record.items():
        if "acknowledgement" in " ".join(str(header).casefold().split()):
            return _clean_text(value)
    return ""


def _clean_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, float) and math.isnan(value):
        return 0.0
    if isinstance(value, str):
        value = (
            value.replace(",", "")
            .replace("\u20b9", "")
            .replace("â‚¹", "")
            .strip()
        )
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _record_to_values(
    record: Mapping[str, Any],
    columns: Sequence[tuple[str, str]],
) -> tuple[Any, ...]:
    values: list[Any] = []
    amount_fields = {
        "total_credited_amount",
        "total_debited_amount",
        "updated_amount",
        "not_updated_amount",
    }
    for display_name, database_name in columns:
        value = record.get(display_name)
        if value is None:
            value = record.get(database_name)
        if database_name in amount_fields:
            values.append(_clean_number(value))
        else:
            values.append(_clean_text(value))
    return tuple(values)


def _logical_key(value: Any) -> str:
    return _clean_text(value).casefold()


def _deduplicate_summary_values(
    values: Sequence[tuple[Any, ...]],
    *,
    key_positions: Sequence[int],
    label: str,
) -> tuple[list[tuple[Any, ...]], int]:
    """Drop identical logical summary rows and reject conflicting repeats."""
    unique_rows: list[tuple[Any, ...]] = []
    seen: dict[tuple[str, ...], tuple[Any, ...]] = {}
    duplicates_removed = 0
    for row in values:
        key = tuple(_logical_key(row[position]) for position in key_positions)
        comparison = tuple(
            _logical_key(value) for value in row[1:-1]
        )  # Ignore source id and generated timestamp.
        previous = seen.get(key)
        if previous is None:
            seen[key] = comparison
            unique_rows.append(row)
            continue
        if previous != comparison:
            raise ValueError(
                f"Conflicting duplicate {label} row for key: "
                f"{' / '.join(key)}"
            )
        duplicates_removed += 1
    return unique_rows, duplicates_removed


def save_file_summaries(
    database_path: str | Path,
    source_file_id: int,
    summaries: Mapping[str, Sequence[Mapping[str, Any]]],
    *,
    duration_seconds: float,
    processing_audit: Mapping[str, Any] | None = None,
) -> dict[str, int]:
    """Atomically replace one source file's summary rows and mark it completed."""
    account_records = list(summaries.get("Account Wise Summary", ()))
    bank_records = list(summaries.get("Bank Wise Summary", ()))
    partial_records = list(summaries.get("Partial Bank Wise Summary", ()))
    money_transfer_other_records = list(
        summaries.get("Money Transfer to Others", ())
    )
    created_at = utc_now()

    account_values_raw = [
        (source_file_id, *_record_to_values(record, ACCOUNT_COLUMNS), created_at)
        for record in account_records
    ]
    bank_values_raw = [
        (source_file_id, *_record_to_values(record, BANK_COLUMNS), created_at)
        for record in bank_records
    ]
    partial_values_raw = [
        (source_file_id, *_record_to_values(record, BANK_COLUMNS), created_at)
        for record in partial_records
    ]
    money_transfer_other_values = [
        (
            source_file_id,
            _raw_acknowledgement(record),
            source_row_number,
            json.dumps(
                dict(record),
                ensure_ascii=False,
                default=str,
                separators=(",", ":"),
            ),
            created_at,
        )
        for source_row_number, record in enumerate(
            money_transfer_other_records,
            start=2,
        )
    ]
    account_values, account_duplicates = _deduplicate_summary_values(
        account_values_raw,
        key_positions=(1, 2, 3, 4, 5, 6, 7, 8, 9),
        label="account",
    )
    bank_values, bank_duplicates = _deduplicate_summary_values(
        bank_values_raw,
        key_positions=(1, 2),
        label="bank",
    )
    partial_values, partial_duplicates = _deduplicate_summary_values(
        partial_values_raw,
        key_positions=(1, 2),
        label="partial-bank",
    )
    duplicate_summary_rows_removed = (
        account_duplicates + bank_duplicates + partial_duplicates
    )
    audit = dict(processing_audit or {})
    acknowledgements = {
        values[1]
        for values in account_values
        if values[1]
    }

    connection = connect_database(database_path)
    try:
        with connection:
            connection.execute(
                "DELETE FROM account_summaries WHERE source_file_id = ?",
                (source_file_id,),
            )
            connection.execute(
                "DELETE FROM bank_summaries WHERE source_file_id = ?",
                (source_file_id,),
            )
            connection.execute(
                "DELETE FROM partial_bank_summaries WHERE source_file_id = ?",
                (source_file_id,),
            )
            connection.execute(
                "DELETE FROM money_transfer_to_others_rows WHERE source_file_id = ?",
                (source_file_id,),
            )

            for acknowledgement in sorted(acknowledgements):
                owner = connection.execute(
                    """
                    SELECT source_file_id
                    FROM account_summaries
                    WHERE acknowledgement_no = ? COLLATE NOCASE
                      AND source_file_id <> ?
                    LIMIT 1
                    """,
                    (acknowledgement, source_file_id),
                ).fetchone()
                if owner is not None:
                    raise ValueError(
                        "Acknowledgement "
                        f"{acknowledgement} is already stored from source file "
                        f"ID {owner['source_file_id']}"
                    )

            if account_values:
                connection.executemany(
                    """
                    INSERT INTO account_summaries (
                        source_file_id,
                        acknowledgement_no,
                        bank_name,
                        account_number,
                        credited_transaction_id,
                        total_credited_amount,
                        total_debited_amount,
                        updated_amount,
                        not_updated_amount,
                        status,
                        found_in_other_sheets,
                        breakdown_by_sheet,
                        duplicate_entry_info,
                        created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    account_values,
                )
            if bank_values:
                connection.executemany(
                    """
                    INSERT INTO bank_summaries (
                        source_file_id,
                        acknowledgement_no,
                        bank_name,
                        total_credited_amount,
                        total_debited_amount,
                        updated_amount,
                        not_updated_amount,
                        status,
                        found_in_other_sheets,
                        breakdown_by_sheet,
                        duplicate_entry_info,
                        created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    bank_values,
                )
            if partial_values:
                connection.executemany(
                    """
                    INSERT INTO partial_bank_summaries (
                        source_file_id,
                        acknowledgement_no,
                        bank_name,
                        total_credited_amount,
                        total_debited_amount,
                        updated_amount,
                        not_updated_amount,
                        status,
                        found_in_other_sheets,
                        breakdown_by_sheet,
                        duplicate_entry_info,
                        created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    partial_values,
                )
            if money_transfer_other_values:
                connection.executemany(
                    """
                    INSERT INTO money_transfer_to_others_rows (
                        source_file_id,
                        acknowledgement_no,
                        source_row_number,
                        row_json,
                        created_at
                    ) VALUES (?, ?, ?, ?, ?)
                    """,
                    money_transfer_other_values,
                )

            connection.execute(
                """
                UPDATE source_files
                SET status = 'completed',
                    completed_at = ?,
                    duration_seconds = ?,
                    acknowledgement_count = ?,
                    account_row_count = ?,
                    bank_row_count = ?,
                    partial_bank_row_count = ?,
                    money_transfer_other_row_count = ?,
                    main_rows_read = ?,
                    duplicate_transaction_rows_removed = ?,
                    other_rows_read = ?,
                    duplicate_other_rows_removed = ?,
                    duplicate_summary_rows_removed = ?,
                    duplicate_of_source_file_id = NULL,
                    fast_duplicate_reprocessed_version = ?,
                    error_message = NULL
                WHERE id = ?
                """,
                (
                    created_at,
                    duration_seconds,
                    len(acknowledgements),
                    len(account_values),
                    len(bank_values),
                    len(partial_values),
                    len(money_transfer_other_values),
                    int(audit.get("main_rows_read", 0) or 0),
                    int(
                        audit.get("duplicate_transaction_rows_removed", 0)
                        or 0
                    ),
                    int(audit.get("other_rows_read", 0) or 0),
                    int(audit.get("duplicate_other_rows_removed", 0) or 0),
                    duplicate_summary_rows_removed,
                    audit.get("duplicate_processing_version"),
                    source_file_id,
                ),
            )
        return {
            "acknowledgements": len(acknowledgements),
            "account": len(account_values),
            "bank": len(bank_values),
            "partial": len(partial_values),
            "money_transfer_to_others": len(money_transfer_other_values),
            "duplicate_summary_rows_removed": duplicate_summary_rows_removed,
            "duplicate_transaction_rows_removed": int(
                audit.get("duplicate_transaction_rows_removed", 0) or 0
            ),
            "duplicate_other_rows_removed": int(
                audit.get("duplicate_other_rows_removed", 0) or 0
            ),
        }
    finally:
        connection.close()


def mark_file_failed(
    database_path: str | Path,
    source_file_id: int,
    error_message: str,
    *,
    duration_seconds: float,
) -> None:
    connection = connect_database(database_path)
    try:
        with connection:
            connection.execute(
                """
                UPDATE source_files
                SET status = 'failed',
                    completed_at = ?,
                    duration_seconds = ?,
                    error_message = ?
                WHERE id = ?
                """,
                (
                    utc_now(),
                    duration_seconds,
                    error_message[:4000],
                    source_file_id,
                ),
            )
    finally:
        connection.close()


def mark_file_as_duplicate(
    database_path: str | Path,
    source_file_id: int,
    canonical_source_file_id: int,
    content_sha256: str,
) -> None:
    """Record an exact workbook copy without storing its summaries again."""
    if source_file_id == canonical_source_file_id:
        raise ValueError("A source file cannot be a duplicate of itself")
    connection = connect_database(database_path)
    try:
        with connection:
            for table_name in (
                "account_summaries",
                "bank_summaries",
                "partial_bank_summaries",
                "money_transfer_to_others_rows",
            ):
                connection.execute(
                    f"DELETE FROM {table_name} WHERE source_file_id = ?",
                    (source_file_id,),
                )
            connection.execute(
                """
                UPDATE source_files
                SET status = 'completed',
                    completed_at = ?,
                    duration_seconds = 0,
                    acknowledgement_count = 0,
                    account_row_count = 0,
                    bank_row_count = 0,
                    partial_bank_row_count = 0,
                    money_transfer_other_row_count = 0,
                    content_sha256 = ?,
                    duplicate_of_source_file_id = ?,
                    main_rows_read = 0,
                    duplicate_transaction_rows_removed = 0,
                    other_rows_read = 0,
                    duplicate_other_rows_removed = 0,
                    duplicate_summary_rows_removed = 0,
                    error_message = NULL
                WHERE id = ?
                """,
                (
                    utc_now(),
                    content_sha256,
                    canonical_source_file_id,
                    source_file_id,
                ),
            )
    finally:
        connection.close()


def set_worker_state(
    database_path: str | Path,
    *,
    is_running: bool,
    process_id: int | None = None,
    current_file: str | None = None,
    message: str | None = None,
    started_at: str | None = None,
) -> None:
    now = utc_now()
    connection = connect_database(database_path)
    try:
        with connection:
            connection.execute(
                """
                UPDATE worker_state
                SET is_running = ?,
                    process_id = ?,
                    started_at = COALESCE(?, started_at),
                    heartbeat_at = ?,
                    current_file = ?,
                    message = ?
                WHERE id = 1
                """,
                (
                    int(is_running),
                    process_id,
                    started_at,
                    now,
                    current_file,
                    message,
                ),
            )
    finally:
        connection.close()


def query_progress(database_path: str | Path) -> dict[str, Any]:
    initialize_database(database_path)
    connection = connect_database(database_path, readonly=True)
    try:
        status_rows = connection.execute(
            """
            SELECT status, COUNT(*) AS file_count
            FROM source_files
            GROUP BY status
            """
        ).fetchall()
        counts = {
            "pending": 0,
            "processing": 0,
            "completed": 0,
            "failed": 0,
        }
        counts.update({row["status"]: row["file_count"] for row in status_rows})
        total = sum(counts.values())
        processed = counts["completed"] + counts["failed"]
        percent = round((processed / total * 100), 2) if total else 0.0
        totals = connection.execute(
            """
            SELECT
                COALESCE(SUM(acknowledgement_count), 0) AS acknowledgements,
                COALESCE(SUM(account_row_count), 0) AS account_rows,
                COALESCE(SUM(bank_row_count), 0) AS bank_rows,
                COALESCE(SUM(partial_bank_row_count), 0) AS partial_rows
            FROM source_files
            """
        ).fetchone()
        duplicate_audit = connection.execute(
            """
            SELECT
                COALESCE(SUM(CASE
                    WHEN duplicate_of_source_file_id IS NOT NULL
                    THEN 1 ELSE 0 END), 0) AS duplicate_files_skipped,
                COALESCE(SUM(duplicate_transaction_rows_removed), 0)
                    AS transaction_rows_removed,
                COALESCE(SUM(duplicate_other_rows_removed), 0)
                    AS other_sheet_rows_removed,
                COALESCE(SUM(duplicate_summary_rows_removed), 0)
                    AS summary_rows_removed
            FROM source_files
            WHERE status = 'completed'
            """
        ).fetchone()
        worker = connection.execute(
            "SELECT * FROM worker_state WHERE id = 1"
        ).fetchone()
        return {
            "files": {
                "total": total,
                **counts,
                "processed": processed,
                "percent": percent,
            },
            "summaries": dict(totals),
            "duplicates": dict(duplicate_audit),
            "worker": dict(worker) if worker else {},
        }
    finally:
        connection.close()


def query_recent_failures(
    database_path: str | Path,
    *,
    limit: int = 10,
) -> list[dict[str, Any]]:
    initialize_database(database_path)
    connection = connect_database(database_path, readonly=True)
    try:
        rows = connection.execute(
            """
            SELECT file_name, source_path, attempts, completed_at, error_message
            FROM source_files
            WHERE status = 'failed'
            ORDER BY completed_at DESC, id DESC
            LIMIT ?
            """,
            (max(1, min(limit, 100)),),
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        connection.close()


def _build_filter_clause(
    config: Mapping[str, Any],
    *,
    acknowledgement: str | None,
    status: str | None,
    search: str | None,
) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    parameters: list[Any] = []
    if acknowledgement and acknowledgement.upper() != "ALL":
        clauses.append("acknowledgement_no = ? COLLATE NOCASE")
        parameters.append(acknowledgement.strip())
    if status and status.upper() not in {"", "ALL"}:
        clauses.append("UPPER(status) = ?")
        parameters.append(status.upper())
    if search:
        search_term = f"%{search.strip()}%"
        search_clauses = []
        for column in config["search_columns"]:
            search_clauses.append(f"COALESCE({column}, '') LIKE ?")
            parameters.append(search_term)
        clauses.append(f"({' OR '.join(search_clauses)})")
    if not clauses:
        return "", parameters
    return " WHERE " + " AND ".join(clauses), parameters


def query_summary_page(
    database_path: str | Path,
    *,
    view: str = "account",
    acknowledgement: str | None = "ALL",
    status: str | None = "ALL",
    search: str | None = None,
    page: int = 1,
    page_size: int = 100,
) -> dict[str, Any]:
    initialize_database(database_path)
    config = VIEW_CONFIG.get(view, VIEW_CONFIG["account"])
    page = max(1, page)
    page_size = max(10, min(page_size, 500))
    where_clause, parameters = _build_filter_clause(
        config,
        acknowledgement=acknowledgement,
        status=status,
        search=search,
    )
    database_columns = [database_name for _, database_name in config["columns"]]
    select_columns = ", ".join(database_columns)

    connection = connect_database(database_path, readonly=True)
    try:
        total = connection.execute(
            f"SELECT COUNT(*) FROM {config['table']}{where_clause}",
            parameters,
        ).fetchone()[0]
        page_count = max(1, math.ceil(total / page_size))
        page = min(page, page_count)
        rows = connection.execute(
            f"""
            SELECT {select_columns}
            FROM {config['table']}
            {where_clause}
            ORDER BY not_updated_amount DESC, acknowledgement_no, id
            LIMIT ? OFFSET ?
            """,
            [*parameters, page_size, (page - 1) * page_size],
        ).fetchall()
        return {
            "view": view,
            "columns": [
                {
                    "label": display_name,
                    "key": database_name,
                    "type": (
                        "amount"
                        if database_name in config["amount_columns"]
                        else (
                            "status"
                            if database_name == "status"
                            else "text"
                        )
                    ),
                }
                for display_name, database_name in config["columns"]
            ],
            "rows": [dict(row) for row in rows],
            "pagination": {
                "page": page,
                "page_size": page_size,
                "page_count": page_count,
                "total": total,
            },
        }
    finally:
        connection.close()


def query_acknowledgements(
    database_path: str | Path,
    *,
    search: str | None = None,
    limit: int = 50,
) -> list[dict[str, Any]]:
    initialize_database(database_path)
    clauses = ""
    parameters: list[Any] = []
    if search:
        clauses = "WHERE acknowledgement_no LIKE ?"
        parameters.append(f"%{search.strip()}%")
    parameters.append(max(1, min(limit, 500)))
    connection = connect_database(database_path, readonly=True)
    try:
        rows = connection.execute(
            f"""
            SELECT acknowledgement_no, COUNT(*) AS account_rows
            FROM account_summaries
            {clauses}
            GROUP BY acknowledgement_no
            ORDER BY acknowledgement_no
            LIMIT ?
            """,
            parameters,
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        connection.close()


def query_summary_totals(
    database_path: str | Path,
    *,
    acknowledgement: str | None = "ALL",
) -> dict[str, Any]:
    initialize_database(database_path)
    where_clause = ""
    parameters: list[Any] = []
    if acknowledgement and acknowledgement.upper() != "ALL":
        where_clause = "WHERE acknowledgement_no = ? COLLATE NOCASE"
        parameters.append(acknowledgement.strip())
    connection = connect_database(database_path, readonly=True)
    try:
        row = connection.execute(
            f"""
            SELECT
                COUNT(*) AS account_rows,
                COUNT(DISTINCT acknowledgement_no) AS acknowledgements,
                COALESCE(SUM(total_credited_amount), 0) AS total_credited,
                COALESCE(SUM(total_debited_amount), 0) AS total_debited,
                COALESCE(SUM(updated_amount), 0) AS total_updated,
                COALESCE(SUM(not_updated_amount), 0) AS total_not_updated,
                SUM(CASE WHEN UPPER(status) = 'PENDING' THEN 1 ELSE 0 END)
                    AS pending_rows,
                SUM(CASE WHEN UPPER(status) = 'PARTIAL' THEN 1 ELSE 0 END)
                    AS partial_rows,
                SUM(CASE WHEN UPPER(status) IN ('COMPLETED', 'COMPLETE')
                    THEN 1 ELSE 0 END) AS completed_rows
            FROM account_summaries
            {where_clause}
            """,
            parameters,
        ).fetchone()
        result = dict(row)
        for key in ("pending_rows", "partial_rows", "completed_rows"):
            result[key] = result[key] or 0
        return result
    finally:
        connection.close()


def _iter_export_rows(
    connection: sqlite3.Connection,
    config: Mapping[str, Any],
    acknowledgement: str | None,
) -> Iterable[sqlite3.Row]:
    where_clause = ""
    parameters: list[Any] = []
    if acknowledgement and acknowledgement.upper() != "ALL":
        where_clause = "WHERE acknowledgement_no = ? COLLATE NOCASE"
        parameters.append(acknowledgement.strip())
    database_columns = ", ".join(
        database_name for _, database_name in config["columns"]
    )
    cursor = connection.execute(
        f"""
        SELECT {database_columns}
        FROM {config['table']}
        {where_clause}
        ORDER BY not_updated_amount DESC, acknowledgement_no, id
        """,
        parameters,
    )
    while True:
        rows = cursor.fetchmany(1000)
        if not rows:
            break
        yield from rows


def _iter_money_transfer_other_records(
    connection: sqlite3.Connection,
    acknowledgement: str | None,
) -> Iterable[dict[str, Any]]:
    where_clause = ""
    parameters: list[Any] = []
    if acknowledgement and acknowledgement.upper() != "ALL":
        where_clause = "WHERE acknowledgement_no = ? COLLATE NOCASE"
        parameters.append(acknowledgement.strip())
    cursor = connection.execute(
        f"""
        SELECT row_json
        FROM money_transfer_to_others_rows
        {where_clause}
        ORDER BY source_file_id, source_row_number, id
        """,
        parameters,
    )
    while True:
        rows = cursor.fetchmany(1000)
        if not rows:
            break
        for row in rows:
            record = json.loads(row["row_json"])
            if isinstance(record, dict):
                yield record


def create_excel_export(
    database_path: str | Path,
    *,
    acknowledgement: str | None = "ALL",
) -> BytesIO:
    """Build three summaries plus raw Money Transfer to Others rows."""
    initialize_database(database_path)
    connection = connect_database(database_path, readonly=True)
    workbook = Workbook(write_only=True)

    header_fill = PatternFill(
        start_color="1F2937",
        end_color="1F2937",
        fill_type="solid",
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    stripe_fill = PatternFill(
        start_color="F8FAFC",
        end_color="F8FAFC",
        fill_type="solid",
    )
    status_styles = {
        "PENDING": (
            PatternFill("solid", fgColor="FEE2E2"),
            Font(color="B91C1C", bold=True),
        ),
        "PARTIAL": (
            PatternFill("solid", fgColor="FEF3C7"),
            Font(color="B45309", bold=True),
        ),
        "COMPLETED": (
            PatternFill("solid", fgColor="DCFCE7"),
            Font(color="15803D", bold=True),
        ),
        "COMPLETE": (
            PatternFill("solid", fgColor="DCFCE7"),
            Font(color="15803D", bold=True),
        ),
    }
    thin_border = Border(
        bottom=Side(style="hair", color="CBD5E1"),
    )

    try:
        connection.execute("BEGIN")
        for view_name in ("account", "bank", "partial"):
            config = VIEW_CONFIG[view_name]
            worksheet = workbook.create_sheet(config["sheet"])
            worksheet.freeze_panes = "A2"
            worksheet.sheet_view.showGridLines = False

            header_cells = []
            for display_name, _ in config["columns"]:
                cell = WriteOnlyCell(worksheet, value=display_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                header_cells.append(cell)
            worksheet.append(header_cells)

            row_count = 0
            amount_columns = config["amount_columns"]
            for row_count, row in enumerate(
                _iter_export_rows(connection, config, acknowledgement),
                start=1,
            ):
                excel_cells = []
                striped = row_count % 2 == 0
                for _, database_name in config["columns"]:
                    value = row[database_name]
                    cell = WriteOnlyCell(worksheet, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if striped:
                        cell.fill = stripe_fill
                    if database_name in amount_columns:
                        cell.number_format = "\u20b9#,##0.00"
                        cell.alignment = Alignment(
                            horizontal="right",
                            vertical="center",
                        )
                    elif database_name == "status":
                        status_value = _clean_text(value).upper()
                        style = status_styles.get(status_value)
                        if style:
                            cell.fill, cell.font = style
                        cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                        )
                    elif database_name == "found_in_other_sheets":
                        cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                        )
                    excel_cells.append(cell)
                worksheet.append(excel_cells)

            last_column = get_column_letter(len(config["columns"]))
            worksheet.auto_filter.ref = f"A1:{last_column}{row_count + 1}"
            for index, width in enumerate(config["column_widths"], start=1):
                worksheet.column_dimensions[get_column_letter(index)].width = width

        raw_columns: list[str] = []
        seen_raw_columns: set[str] = set()
        for record in _iter_money_transfer_other_records(
            connection,
            acknowledgement,
        ):
            for column in record:
                if column not in seen_raw_columns:
                    seen_raw_columns.add(column)
                    raw_columns.append(column)

        raw_worksheet = workbook.create_sheet("Money Transfer to Others")
        raw_worksheet.freeze_panes = "A2"
        raw_worksheet.sheet_view.showGridLines = False
        if raw_columns:
            raw_header_cells = []
            for column in raw_columns:
                cell = WriteOnlyCell(raw_worksheet, value=column)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                raw_header_cells.append(cell)
            raw_worksheet.append(raw_header_cells)

            raw_row_count = 0
            for raw_row_count, record in enumerate(
                _iter_money_transfer_other_records(
                    connection,
                    acknowledgement,
                ),
                start=1,
            ):
                cells = []
                for column in raw_columns:
                    cell = WriteOnlyCell(
                        raw_worksheet,
                        value=record.get(column),
                    )
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    if raw_row_count % 2 == 0:
                        cell.fill = stripe_fill
                    cells.append(cell)
                raw_worksheet.append(cells)
            raw_last_column = get_column_letter(len(raw_columns))
            raw_worksheet.auto_filter.ref = (
                f"A1:{raw_last_column}{raw_row_count + 1}"
            )
            for index, column in enumerate(raw_columns, start=1):
                raw_worksheet.column_dimensions[
                    get_column_letter(index)
                ].width = min(50, max(12, len(column) + 2))

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        connection.rollback()
        return output
    finally:
        connection.close()
