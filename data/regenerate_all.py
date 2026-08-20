"""
Clean Bank-wise Excel Generator (All 3 folders)
================================================
Layout: Title Banner → Column Headers → Data
All columns auto-fitted so data is visible at first sight.
Status colors: COMPLETED=Green, PARTIAL=Orange, PENDING=Red
"""

import sqlite3
import os
import re
import shutil
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = r'c:\Users\admin\Desktop\LAYERED2\data\account_summaries.sqlite'
BASE_DIR = r'c:\Users\admin\Desktop\LAYERED2\data'

COLUMNS = [
    ('acknowledgement_no',      'Acknowledgement No'),
    ('bank_name',               'Bank Name'),
    ('account_number',          'Account Number'),
    ('credited_transaction_id', 'Credited Transaction Id'),
    ('total_credited_amount',   'Total Credited Amount'),
    ('total_debited_amount',    'Total Debited Amount'),
    ('updated_amount',          'Updated Amount (Recovery)'),
    ('not_updated_amount',      'Not Updated Amount'),
    ('status',                  'Status'),
    ('found_in_other_sheets',   'Found in Other Sheets'),
    ('breakdown_by_sheet',      'Breakdown by Sheet'),
]

# ── Styles ──
STATUS_FILLS = {
    'COMPLETED': PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid'),
    'COMPLETE':  PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid'),
    'PARTIAL':   PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid'),
    'PENDING':   PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid'),
}
STATUS_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

TITLE_FILL   = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
TITLE_FONT   = Font(name='Calibri', size=13, bold=True, color='FFFFFF')

HEADER_FILL  = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
HEADER_FONT  = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

ROW_EVEN = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
ROW_ODD  = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

DATA_FONT  = Font(name='Calibri', size=10)
DATA_ALIGN = Alignment(horizontal='left', vertical='center')
NUM_ALIGN  = Alignment(horizontal='right', vertical='center')
CTR_ALIGN  = Alignment(horizontal='center', vertical='center')

THIN_BORDER = Border(
    left=Side(style='thin', color='BDC3C7'),
    right=Side(style='thin', color='BDC3C7'),
    top=Side(style='thin', color='BDC3C7'),
    bottom=Side(style='thin', color='BDC3C7'),
)

MONEY_COLS = {'total_credited_amount', 'total_debited_amount', 'updated_amount', 'not_updated_amount'}


def sanitize(name):
    c = re.sub(r'[<>:"/\\|?*]', '_', name)
    c = re.sub(r'\s+', ' ', c).strip()
    return c[:80]


def auto_width(ws, col_idx, header_text, data_values):
    """Calculate best column width from header and data."""
    max_len = len(str(header_text))
    # Sample up to 100 values for speed
    for v in data_values[:100]:
        if pd.isna(v):
            continue
        vstr = str(v)
        # For numbers, approximate formatted length
        if isinstance(v, float):
            vstr = f"₹{v:,.2f}"
        max_len = max(max_len, len(vstr))
    # Add padding, cap at reasonable width
    width = min(max_len + 3, 55)
    width = max(width, 12)
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_excel(bank_name, df, output_dir, title_suffix='Account Summary Report'):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.sheet_view.showGridLines = False
    num_cols = len(COLUMNS)

    # ── Row 1: Title Banner ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    tc = ws.cell(row=1, column=1)
    tc.value = f"📊  {bank_name}  —  {title_suffix}"
    tc.font = TITLE_FONT
    tc.fill = TITLE_FILL
    tc.alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, num_cols + 1):
        ws.cell(row=1, column=c).fill = TITLE_FILL
        ws.cell(row=1, column=c).border = THIN_BORDER
    ws.row_dimensions[1].height = 28

    # ── Row 2: Column Headers ──
    for ci, (_, dn) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=ci)
        cell.value = dn
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 30

    # Freeze below header
    ws.freeze_panes = 'A3'

    # ── Row 3+: Data ──
    for i, (_, rd) in enumerate(df.iterrows()):
        dr = 3 + i
        rf = ROW_EVEN if i % 2 == 0 else ROW_ODD
        for ci, (dc, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=dr, column=ci)
            val = rd.get(dc, '')
            if pd.isna(val):
                val = ''
            cell.value = val
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            if dc == 'status':
                su = str(val).upper().strip()
                if su in STATUS_FILLS:
                    cell.fill = STATUS_FILLS[su]
                    cell.font = STATUS_FONT
                    cell.alignment = CTR_ALIGN
                else:
                    cell.fill = rf
                    cell.alignment = CTR_ALIGN
            elif dc in MONEY_COLS:
                if isinstance(val, (int, float)) and val != 0:
                    cell.number_format = '₹#,##0.00'
                cell.alignment = NUM_ALIGN
                cell.fill = rf
            elif dc in ('found_in_other_sheets',):
                cell.alignment = CTR_ALIGN
                cell.fill = rf
            else:
                cell.alignment = DATA_ALIGN
                cell.fill = rf

        ws.row_dimensions[dr].height = 20

    # ── Auto-fit column widths ──
    for ci, (dc, dn) in enumerate(COLUMNS, 1):
        col_data = df[dc].tolist() if dc in df.columns else []
        auto_width(ws, ci, dn, col_data)

    # ── Auto-filter on headers ──
    last_row = 2 + len(df)
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}{last_row}"

    # Save
    fp = os.path.join(output_dir, f"{sanitize(bank_name)}.xlsx")
    wb.save(fp)
    return fp


# ── Folder configs ──
FOLDERS = [
    {
        'name': 'bank_wise_reports',
        'label': 'ALL DATA',
        'title': 'Account Summary Report',
        'query': """
            SELECT acknowledgement_no, bank_name, account_number, credited_transaction_id,
                   total_credited_amount, total_debited_amount, updated_amount, not_updated_amount,
                   status, found_in_other_sheets, breakdown_by_sheet
            FROM account_summaries
            WHERE bank_name NOT LIKE 'Reassign Back To%'
            ORDER BY bank_name, acknowledgement_no
        """,
    },
    {
        'name': 'bank_wise_partial',
        'label': 'PARTIAL (Not Updated ≥ ₹1,000)',
        'title': 'Partial Status Report',
        'query': """
            SELECT acknowledgement_no, bank_name, account_number, credited_transaction_id,
                   total_credited_amount, total_debited_amount, updated_amount, not_updated_amount,
                   status, found_in_other_sheets, breakdown_by_sheet
            FROM account_summaries
            WHERE bank_name NOT LIKE 'Reassign Back To%'
              AND UPPER(TRIM(status)) = 'PARTIAL'
              AND not_updated_amount >= 1000
            ORDER BY bank_name, acknowledgement_no
        """,
    },
    {
        'name': 'bank_wise_pending',
        'label': 'PENDING (Not Updated ≥ ₹1,000)',
        'title': 'Pending Status Report',
        'query': """
            SELECT acknowledgement_no, bank_name, account_number, credited_transaction_id,
                   total_credited_amount, total_debited_amount, updated_amount, not_updated_amount,
                   status, found_in_other_sheets, breakdown_by_sheet
            FROM account_summaries
            WHERE bank_name NOT LIKE 'Reassign Back To%'
              AND UPPER(TRIM(status)) = 'PENDING'
              AND not_updated_amount >= 1000
            ORDER BY bank_name, acknowledgement_no
        """,
    },
]


def main():
    print("=" * 70)
    print("  REGENERATING ALL BANK-WISE EXCEL REPORTS (Clean Format)")
    print("=" * 70)

    conn = sqlite3.connect(DB_PATH)

    for folder_cfg in FOLDERS:
        folder_path = os.path.join(BASE_DIR, folder_cfg['name'])

        # Clear folder contents (handle locked files gracefully)
        if os.path.exists(folder_path):
            for f in os.listdir(folder_path):
                try:
                    os.remove(os.path.join(folder_path, f))
                except Exception:
                    pass
        os.makedirs(folder_path, exist_ok=True)

        print(f"\n{'─' * 70}")
        print(f"  📂 {folder_cfg['label']}")
        print(f"  📁 {folder_path}")
        print(f"{'─' * 70}")

        df = pd.read_sql_query(folder_cfg['query'], conn)
        print(f"  Total records: {len(df):,}")

        if len(df) == 0:
            print("  ⚠️  No records. Skipping.")
            continue

        banks = sorted(df['bank_name'].dropna().unique())
        print(f"  Banks: {len(banks)}\n")

        created = 0
        for idx, bank in enumerate(banks, 1):
            bdf = df[df['bank_name'] == bank].copy()
            if len(bdf) == 0:
                continue
            try:
                create_excel(bank, bdf, folder_path, folder_cfg['title'])
                created += 1
                print(f"  [{idx:4d}/{len(banks)}] ✅ {bank} ({len(bdf):,} records)")
            except Exception as e:
                print(f"  [{idx:4d}/{len(banks)}] ❌ {bank} - ERROR: {e}")

        print(f"\n  ✅ {created} files → {folder_cfg['name']}/")

    conn.close()
    print(f"\n{'=' * 70}")
    print("  ✅ ALL 3 FOLDERS REGENERATED!")
    print(f"{'=' * 70}")


if __name__ == '__main__':
    main()
