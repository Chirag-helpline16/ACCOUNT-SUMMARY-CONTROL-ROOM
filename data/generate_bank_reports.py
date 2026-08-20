"""
Bank-wise Excel Report Generator
=================================
Reads account_summaries from SQLite, splits by bank_name,
and creates beautifully formatted Excel files with status color coding:
  - COMPLETED = Green
  - PARTIAL   = Orange
  - PENDING   = Red
"""

import sqlite3
import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Configuration ──────────────────────────────────────────────────────────
DB_PATH = r'c:\Users\admin\Desktop\LAYERED2\data\account_summaries.sqlite'
OUTPUT_DIR = r'c:\Users\admin\Desktop\LAYERED2\data\bank_wise_reports'

# Column mapping (DB → Display)
COLUMNS = [
    ('acknowledgement_no', 'Acknowledgement No'),
    ('bank_name', 'Bank Name'),
    ('account_number', 'Account Number'),
    ('credited_transaction_id', 'Credited Transaction Id'),
    ('total_credited_amount', 'Total Credited Amount'),
    ('total_debited_amount', 'Total Debited Amount'),
    ('updated_amount', 'Updated Amount (Recovery)'),
    ('not_updated_amount', 'Not Updated Amount'),
    ('status', 'Status'),
    ('found_in_other_sheets', 'Found in Other Sheets'),
    ('breakdown_by_sheet', 'Breakdown by Sheet'),
]

# Status color scheme
STATUS_COLORS = {
    'COMPLETED': PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid'),  # Green
    'COMPLETE':  PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid'),
    'PARTIAL':   PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid'),  # Orange
    'PENDING':   PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid'),  # Red
}
STATUS_FONTS = {
    'COMPLETED': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
    'COMPLETE':  Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
    'PARTIAL':   Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
    'PENDING':   Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
}

# Style definitions
HEADER_FILL = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Alternating row colors for easy reading
ROW_FILL_EVEN = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')  # Light blue
ROW_FILL_ODD = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')    # White

DATA_FONT = Font(name='Calibri', size=11)
DATA_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
NUM_ALIGN = Alignment(horizontal='right', vertical='center')

THIN_BORDER = Border(
    left=Side(style='thin', color='BDC3C7'),
    right=Side(style='thin', color='BDC3C7'),
    top=Side(style='thin', color='BDC3C7'),
    bottom=Side(style='thin', color='BDC3C7'),
)

# Title bar style
TITLE_FILL = PatternFill(start_color='0B3D91', end_color='0B3D91', fill_type='solid')
TITLE_FONT = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
SUBTITLE_FONT = Font(name='Calibri', size=11, italic=True, color='A9CCE3')

# Summary box styles
SUMMARY_LABEL_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
SUMMARY_LABEL_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
SUMMARY_VALUE_FILL = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
SUMMARY_VALUE_FONT = Font(name='Calibri', size=11, bold=True, color='2C3E50')


def sanitize_filename(name):
    """Create a safe filename from bank name."""
    clean = re.sub(r'[<>:"/\\|?*]', '_', name)
    clean = re.sub(r'\s+', ' ', clean).strip()
    # Limit filename length
    if len(clean) > 80:
        clean = clean[:80]
    return clean


def create_excel_for_bank(bank_name, df, output_dir):
    """Create a beautifully formatted Excel file for a single bank."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Account Summary"
    
    # Freeze panes
    ws.sheet_view.showGridLines = False
    
    # ── Row 1-2: Title Banner ──
    num_cols = len(COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"📊  {bank_name}  —  Account Summary Report"
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply title fill across all columns
    for col in range(1, num_cols + 1):
        for row in [1, 2]:
            cell = ws.cell(row=row, column=col)
            cell.fill = TITLE_FILL
    
    # ── Row 3: Subtitle ──
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
    sub_cell = ws.cell(row=3, column=1)
    sub_cell.value = f"Total Records: {len(df)}  |  Generated from Bank Transaction Tracker Database"
    sub_cell.font = SUBTITLE_FONT
    sub_cell.fill = TITLE_FILL
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, num_cols + 1):
        ws.cell(row=3, column=col).fill = TITLE_FILL
    
    # ── Row 4: Summary Stats Bar ──
    completed_count = len(df[df['status'].str.upper().str.strip() == 'COMPLETED']) if 'status' in df.columns else 0
    partial_count = len(df[df['status'].str.upper().str.strip() == 'PARTIAL']) if 'status' in df.columns else 0
    pending_count = len(df[df['status'].str.upper().str.strip() == 'PENDING']) if 'status' in df.columns else 0
    
    total_credited = df['total_credited_amount'].sum() if 'total_credited_amount' in df.columns else 0
    total_debited = df['total_debited_amount'].sum() if 'total_debited_amount' in df.columns else 0
    total_recovery = df['updated_amount'].sum() if 'updated_amount' in df.columns else 0
    
    # Summary row
    summary_items = [
        ('✅ Completed', completed_count),
        ('🟠 Partial', partial_count),
        ('🔴 Pending', pending_count),
        ('💰 Total Credited', f'₹{total_credited:,.2f}'),
        ('💸 Total Debited', f'₹{total_debited:,.2f}'),
        ('🔄 Recovery', f'₹{total_recovery:,.2f}'),
    ]
    
    row_idx = 4
    # Blank separator
    for col in range(1, num_cols + 1):
        ws.cell(row=row_idx, column=col).fill = PatternFill(start_color='FDFEFE', end_color='FDFEFE', fill_type='solid')
    row_idx = 5
    
    # Write summary items in pairs across the columns
    col_idx = 1
    for label, value in summary_items:
        if col_idx > num_cols - 1:
            break
        cell_label = ws.cell(row=row_idx, column=col_idx)
        cell_label.value = label
        cell_label.font = SUMMARY_LABEL_FONT
        cell_label.fill = SUMMARY_LABEL_FILL
        cell_label.alignment = Alignment(horizontal='center', vertical='center')
        cell_label.border = THIN_BORDER
        
        cell_value = ws.cell(row=row_idx, column=col_idx + 1)
        cell_value.value = value
        cell_value.font = SUMMARY_VALUE_FONT
        cell_value.fill = SUMMARY_VALUE_FILL
        cell_value.alignment = Alignment(horizontal='center', vertical='center')
        cell_value.border = THIN_BORDER
        
        col_idx += 2
    
    # Fill remaining summary row cells
    for c in range(col_idx, num_cols + 1):
        ws.cell(row=row_idx, column=c).fill = PatternFill(start_color='FDFEFE', end_color='FDFEFE', fill_type='solid')
    
    row_idx = 6
    # Blank separator
    for col in range(1, num_cols + 1):
        ws.cell(row=row_idx, column=col).fill = PatternFill(start_color='FDFEFE', end_color='FDFEFE', fill_type='solid')
    
    # ── Row 7: Headers ──
    header_row = 7
    for col_idx, (db_col, display_name) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = display_name
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    
    ws.row_dimensions[header_row].height = 35
    
    # Freeze panes below header
    ws.freeze_panes = f'A{header_row + 1}'
    
    # ── Data Rows ──
    money_cols = {'total_credited_amount', 'total_debited_amount', 'updated_amount', 'not_updated_amount'}
    
    for i, (_, row_data) in enumerate(df.iterrows()):
        data_row = header_row + 1 + i
        is_even = (i % 2 == 0)
        row_fill = ROW_FILL_EVEN if is_even else ROW_FILL_ODD
        
        for col_idx, (db_col, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=data_row, column=col_idx)
            value = row_data.get(db_col, '')
            
            # Handle NaN
            if pd.isna(value):
                value = ''
            
            cell.value = value
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            
            # Status column special formatting
            if db_col == 'status':
                status_upper = str(value).upper().strip()
                if status_upper in STATUS_COLORS:
                    cell.fill = STATUS_COLORS[status_upper]
                    cell.font = STATUS_FONTS[status_upper]
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.fill = row_fill
                    cell.alignment = DATA_ALIGN
            elif db_col in money_cols:
                # Format as currency
                if isinstance(value, (int, float)) and value != 0:
                    cell.number_format = '₹#,##0.00'
                cell.alignment = NUM_ALIGN
                cell.fill = row_fill
            else:
                cell.fill = row_fill
                cell.alignment = DATA_ALIGN
        
        # Set row height
        ws.row_dimensions[data_row].height = 22
    
    # ── Auto-fit column widths ──
    col_widths = {
        'Acknowledgement No': 22,
        'Bank Name': 30,
        'Account Number': 25,
        'Credited Transaction Id': 28,
        'Total Credited Amount': 22,
        'Total Debited Amount': 22,
        'Updated Amount (Recovery)': 24,
        'Not Updated Amount': 22,
        'Status': 15,
        'Found in Other Sheets': 20,
        'Breakdown by Sheet': 40,
    }
    
    for col_idx, (db_col, display_name) in enumerate(COLUMNS, 1):
        width = col_widths.get(display_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # ── Add Auto-Filter ──
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(num_cols)}{header_row + len(df)}"
    
    # ── Save ──
    safe_name = sanitize_filename(bank_name)
    filepath = os.path.join(output_dir, f"{safe_name}.xlsx")
    wb.save(filepath)
    return filepath


def main():
    # Create output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    print("=" * 70)
    print("  BANK-WISE EXCEL REPORT GENERATOR")
    print("=" * 70)
    
    # Connect to database
    print(f"\n📂 Connecting to database: {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    
    # Get list of all bank names (exclude "Reassign Back To" entries)
    db_cols = [col[0] for col in COLUMNS]
    cols_str = ', '.join(db_cols)
    
    query = f"""
        SELECT {cols_str} 
        FROM account_summaries 
        WHERE bank_name NOT LIKE 'Reassign Back To%'
        ORDER BY bank_name, acknowledgement_no
    """
    
    print("📊 Loading data from account_summaries...")
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    print(f"   Total records loaded: {len(df):,}")
    
    # Get unique banks
    banks = df['bank_name'].dropna().unique()
    banks = sorted(banks)
    print(f"   Unique banks found: {len(banks)}")
    
    print(f"\n📁 Output directory: {OUTPUT_DIR}")
    print(f"\n{'─' * 70}")
    print(f"  Generating Excel files...")
    print(f"{'─' * 70}\n")
    
    created_files = []
    for idx, bank in enumerate(banks, 1):
        bank_df = df[df['bank_name'] == bank].copy()
        
        if len(bank_df) == 0:
            continue
        
        try:
            filepath = create_excel_for_bank(bank, bank_df, OUTPUT_DIR)
            created_files.append((bank, len(bank_df), filepath))
            print(f"  [{idx:4d}/{len(banks)}] ✅ {bank} ({len(bank_df):,} records)")
        except Exception as e:
            print(f"  [{idx:4d}/{len(banks)}] ❌ {bank} - ERROR: {e}")
    
    print(f"\n{'=' * 70}")
    print(f"  ✅ COMPLETE! {len(created_files)} Excel files created.")
    print(f"  📁 Location: {OUTPUT_DIR}")
    print(f"{'=' * 70}")


if __name__ == '__main__':
    main()
