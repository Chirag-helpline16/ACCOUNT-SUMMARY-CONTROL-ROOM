"""
Filtered Bank-wise Excel Report Generator
==========================================
Creates two folders:
  1. bank_wise_partial  — PARTIAL status + not_updated_amount >= 1000
  2. bank_wise_pending  — PENDING status + not_updated_amount >= 1000
Credit-only duplicate handling is applied by the overnight processor before
these stored summaries are read. Report rows are not removed here, so debit
and other-sheet/recovery values remain intact.
Outputs: individual Excel files per bank + one ZIP per folder.
"""

import sqlite3
import os
import re
import zipfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Configuration ──────────────────────────────────────────────────────────
DB_PATH = r'c:\Users\admin\Desktop\LAYERED2\data\account_summaries.sqlite'
BASE_DIR = r'c:\Users\admin\Desktop\LAYERED2\data'

FILTERS = [
    {
        'folder': 'bank_wise_partial',
        'label': 'PARTIAL (Not Updated ≥ ₹1,000)',
        'query': """
            SELECT acknowledgement_no, bank_name, account_number, credited_transaction_id,
                   total_credited_amount, total_debited_amount, updated_amount, not_updated_amount,
                   status, found_in_other_sheets, breakdown_by_sheet, duplicate_entry_info
            FROM account_summaries
            WHERE bank_name NOT LIKE 'Reassign Back To%'
              AND UPPER(TRIM(status)) = 'PARTIAL'
              AND not_updated_amount >= 1000
            ORDER BY bank_name, acknowledgement_no
        """,
        'title_tag': 'Partial Status (Not Updated ≥ ₹1,000)',
        'banner_color': 'E67E22',   # Orange theme
        'header_color': '935116',
    },
    {
        'folder': 'bank_wise_pending',
        'label': 'PENDING',
        'query': """
            SELECT acknowledgement_no, bank_name, account_number, credited_transaction_id,
                   total_credited_amount, total_debited_amount, updated_amount, not_updated_amount,
                   status, found_in_other_sheets, breakdown_by_sheet, duplicate_entry_info
            FROM account_summaries
            WHERE bank_name NOT LIKE 'Reassign Back To%'
              AND UPPER(TRIM(status)) = 'PENDING'
              AND not_updated_amount >= 1000
            ORDER BY bank_name, acknowledgement_no
        """,
        'title_tag': 'Pending Status',
        'banner_color': 'C0392B',   # Red theme
        'header_color': '7B241C',
    },
]

COLUMNS = [
    ('acknowledgement_no', 'Acknowledgement No'),
    ('bank_name', 'Bank Name'),
    ('account_number', 'Account Number'),
    ('credited_transaction_id', 'Credited Transaction Id'),
    ('total_credited_amount', 'Total Credited Amount'),
    ('total_debited_amount', 'Total Debited Amount'),
    ('updated_amount', 'Updated Amount (Recovery)'),
    ('not_updated_amount', 'Not Updated Amount'),
    ('status', 'Status'),
    ('found_in_other_sheets', 'Found in Other Sheets'),
    ('breakdown_by_sheet', 'Breakdown by Sheet'),
    ('duplicate_entry_info', 'Duplicate Entry Info'),
]

# Status color scheme
STATUS_COLORS = {
    'COMPLETED': PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid'),
    'COMPLETE':  PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid'),
    'PARTIAL':   PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid'),
    'PENDING':   PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid'),
}
STATUS_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

ROW_FILL_EVEN = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
ROW_FILL_ODD  = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
DATA_FONT     = Font(name='Calibri', size=11)
DATA_ALIGN    = Alignment(horizontal='left', vertical='center', wrap_text=True)
NUM_ALIGN     = Alignment(horizontal='right', vertical='center')

THIN_BORDER = Border(
    left=Side(style='thin', color='BDC3C7'),
    right=Side(style='thin', color='BDC3C7'),
    top=Side(style='thin', color='BDC3C7'),
    bottom=Side(style='thin', color='BDC3C7'),
)

SUMMARY_LABEL_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
SUMMARY_LABEL_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
SUMMARY_VALUE_FILL = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
SUMMARY_VALUE_FONT = Font(name='Calibri', size=11, bold=True, color='2C3E50')

COL_WIDTHS = {
    'Acknowledgement No': 22, 'Bank Name': 30, 'Account Number': 25,
    'Credited Transaction Id': 28, 'Total Credited Amount': 22,
    'Total Debited Amount': 22, 'Updated Amount (Recovery)': 24,
    'Not Updated Amount': 22, 'Status': 15, 'Found in Other Sheets': 20,
    'Breakdown by Sheet': 40, 'Duplicate Entry Info': 35,
}


def sanitize_filename(name):
    clean = re.sub(r'[<>:"/\\|?*]', '_', name)
    clean = re.sub(r'\s+', ' ', clean).strip()
    return clean[:80] if len(clean) > 80 else clean


def create_excel(bank_name, df, output_dir, banner_color, header_color, title_tag):
    wb = Workbook()
    ws = wb.active
    ws.title = "Account Summary"
    ws.sheet_view.showGridLines = False
    num_cols = len(COLUMNS)

    title_fill  = PatternFill(start_color=banner_color, end_color=banner_color, fill_type='solid')
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    title_font  = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    sub_font    = Font(name='Calibri', size=11, italic=True, color='F5CBA7')
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── Title Banner (rows 1-2) ──
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=num_cols)
    tc = ws.cell(row=1, column=1)
    tc.value = f"📊  {bank_name}  —  {title_tag}"
    tc.font = title_font; tc.fill = title_fill
    tc.alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, num_cols + 1):
        for r in (1, 2):
            ws.cell(row=r, column=c).fill = title_fill

    # ── Subtitle (row 3) ──
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
    sc = ws.cell(row=3, column=1)
    sc.value = f"Total Records: {len(df)}  |  Generated from Bank Transaction Tracker Database"
    sc.font = sub_font; sc.fill = title_fill
    sc.alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, num_cols + 1):
        ws.cell(row=3, column=c).fill = title_fill

    # ── Separator (row 4) ──
    sep_fill = PatternFill(start_color='FDFEFE', end_color='FDFEFE', fill_type='solid')
    for c in range(1, num_cols + 1):
        ws.cell(row=4, column=c).fill = sep_fill

    # ── Summary stats (row 5) ──
    total_credited = df['total_credited_amount'].sum() if 'total_credited_amount' in df.columns else 0
    total_debited  = df['total_debited_amount'].sum() if 'total_debited_amount' in df.columns else 0
    total_recovery = df['updated_amount'].sum() if 'updated_amount' in df.columns else 0
    total_not_upd  = df['not_updated_amount'].sum() if 'not_updated_amount' in df.columns else 0

    summary_items = [
        ('📋 Records', len(df)),
        ('💰 Total Credited', f'₹{total_credited:,.2f}'),
        ('💸 Total Debited', f'₹{total_debited:,.2f}'),
        ('🔄 Recovery', f'₹{total_recovery:,.2f}'),
        ('⏳ Not Updated', f'₹{total_not_upd:,.2f}'),
    ]

    ci = 1
    for label, value in summary_items:
        if ci > num_cols - 1:
            break
        cl = ws.cell(row=5, column=ci)
        cl.value = label; cl.font = SUMMARY_LABEL_FONT; cl.fill = SUMMARY_LABEL_FILL
        cl.alignment = Alignment(horizontal='center', vertical='center'); cl.border = THIN_BORDER

        cv = ws.cell(row=5, column=ci + 1)
        cv.value = value; cv.font = SUMMARY_VALUE_FONT; cv.fill = SUMMARY_VALUE_FILL
        cv.alignment = Alignment(horizontal='center', vertical='center'); cv.border = THIN_BORDER
        ci += 2

    for c in range(ci, num_cols + 1):
        ws.cell(row=5, column=c).fill = sep_fill

    # Separator row 6
    for c in range(1, num_cols + 1):
        ws.cell(row=6, column=c).fill = sep_fill

    # ── Headers (row 7) ──
    hdr = 7
    for ci2, (_, dn) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=hdr, column=ci2)
        cell.value = dn; cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = THIN_BORDER
    ws.row_dimensions[hdr].height = 35
    ws.freeze_panes = f'A{hdr + 1}'

    # ── Data rows ──
    money_cols = {'total_credited_amount', 'total_debited_amount', 'updated_amount', 'not_updated_amount'}
    for i, (_, rd) in enumerate(df.iterrows()):
        dr = hdr + 1 + i
        rf = ROW_FILL_EVEN if i % 2 == 0 else ROW_FILL_ODD
        for ci3, (dc, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=dr, column=ci3)
            val = rd.get(dc, '')
            if pd.isna(val):
                val = ''
            cell.value = val; cell.font = DATA_FONT; cell.border = THIN_BORDER

            if dc == 'status':
                su = str(val).upper().strip()
                if su in STATUS_COLORS:
                    cell.fill = STATUS_COLORS[su]; cell.font = STATUS_FONT
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.fill = rf; cell.alignment = DATA_ALIGN
            elif dc in money_cols:
                if isinstance(val, (int, float)) and val != 0:
                    cell.number_format = '₹#,##0.00'
                cell.alignment = NUM_ALIGN; cell.fill = rf
            else:
                cell.fill = rf; cell.alignment = DATA_ALIGN
        ws.row_dimensions[dr].height = 22

    # Column widths
    from openpyxl.utils import get_column_letter
    for ci4, (_, dn) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci4)].width = COL_WIDTHS.get(dn, 20)

    # Auto-filter
    ws.auto_filter.ref = f"A{hdr}:{get_column_letter(num_cols)}{hdr + len(df)}"

    safe = sanitize_filename(bank_name)
    fp = os.path.join(output_dir, f"{safe}.xlsx")
    wb.save(fp)
    return fp


def dedup_df(df):
    """Keep all summaries; credit deduplication already ran overnight."""
    return df.copy(), 0


def create_zip(folder_path, zip_path):
    """ZIP all .xlsx files in folder_path into zip_path."""
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname in sorted(os.listdir(folder_path)):
            if fname.endswith('.xlsx'):
                fpath = os.path.join(folder_path, fname)
                zf.write(fpath, fname)
    return zip_path


def main():
    print("=" * 70)
    print("  FILTERED BANK-WISE EXCEL REPORT GENERATOR")
    print("  Using Overnight Credit-Only Duplicate Totals")
    print("=" * 70)

    conn = sqlite3.connect(DB_PATH)

    for filt in FILTERS:
        folder_path = os.path.join(BASE_DIR, filt['folder'])
        os.makedirs(folder_path, exist_ok=True)

        print(f"\n{'─' * 70}")
        print(f"  📂 Filter: {filt['label']}")
        print(f"  📁 Output: {folder_path}")
        print(f"{'─' * 70}")

        df = pd.read_sql_query(filt['query'], conn)
        print(f"  Total records from DB: {len(df):,}")

        if len(df) == 0:
            print("  ⚠️  No records found for this filter. Skipping.")
            continue

        # Never drop a complete summary row here. Overnight processing already
        # excluded duplicate credits while preserving debit and recovery data.
        df, total_removed = dedup_df(df)
        print(f"  Stored rows retained: {len(df):,} (report rows removed: {total_removed:,})")

        banks = sorted(df['bank_name'].dropna().unique())
        print(f"  Unique banks: {len(banks)}\n")

        created = 0
        total_rows_written = 0
        for idx, bank in enumerate(banks, 1):
            bdf = df[df['bank_name'] == bank].copy()
            if len(bdf) == 0:
                continue
            try:
                create_excel(bank, bdf, folder_path,
                             filt['banner_color'], filt['header_color'], filt['title_tag'])
                created += 1
                total_rows_written += len(bdf)
                if idx % 50 == 0 or idx == len(banks):
                    print(f"  [{idx:4d}/{len(banks)}] ✅ Progress... {created} files, {total_rows_written:,} rows")
            except Exception as e:
                print(f"  [{idx:4d}/{len(banks)}] ❌ {bank} - ERROR: {e}")

        print(f"\n  ✅ {created} files created in {filt['folder']}/")
        print(f"  📊 Total rows written: {total_rows_written:,}")

        # Create ZIP
        zip_path = os.path.join(BASE_DIR, f"{filt['folder']}.zip")
        create_zip(folder_path, zip_path)
        zip_size = os.path.getsize(zip_path) / (1024 * 1024)
        print(f"  📦 ZIP created: {zip_path} ({zip_size:.1f} MB)")

    conn.close()
    print(f"\n{'=' * 70}")
    print("  ✅ ALL DONE!")
    print(f"  📁 PARTIAL folder: {os.path.join(BASE_DIR, 'bank_wise_partial')}")
    print(f"  📁 PENDING folder: {os.path.join(BASE_DIR, 'bank_wise_pending')}")
    print(f"  📦 ZIPs: bank_wise_partial.zip, bank_wise_pending.zip")
    print(f"{'=' * 70}")


if __name__ == '__main__':
    main()
