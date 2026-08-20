from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import Workbook

import app_account
from batch_account_summaries import (
    DUPLICATE_SUMMARY_VERSION,
    FAST_DUPLICATE_AUDIT_VERSION,
    analyse_with_existing_app,
    audit_workbook_duplicate_keys,
    claim_next_file,
    discover_files,
    fast_audit_and_queue_duplicate_files,
    release_app_memory,
)
from summary_database import (
    connect_database,
    initialize_database,
    save_file_summaries,
    utc_now,
)


def _main_row(
    *,
    serial: int,
    credited_transaction_id: str = "CREDIT-001",
    reference: str = "REF-A",
    remarks: str = "first copy",
) -> list[object]:
    return [
        serial,
        "ACK-001",
        "XXXXXX1234",
        "DEBIT-001",
        "Example Bank",
        1,
        "9876544321",
        "EXAM0001234",
        pd.Timestamp("2026-08-01 10:30:00"),
        credited_transaction_id,
        5000.0,
        2500.0,
        reference,
        remarks,
        "Officer A",
        pd.Timestamp("2026-08-02 09:00:00"),
        "unused",
    ]


def _account_record(*, acknowledgement: str = "ACK-001") -> dict[str, object]:
    return {
        "Acknowledgement No": acknowledgement,
        "Bank Name": "Example Bank",
        "Account Number": "XXXXXX4321",
        "Credited Transaction ID": "CREDIT-001",
        "Total Credited Amount": 5000,
        "Total Debited Amount": 2500,
        "Updated Amount (Recovery)": 0,
        "Not Updated Amount": 2500,
        "Status": "PENDING",
        "Found in Other Sheets": "No",
        "Breakdown by Sheet": "",
        "Duplicate Entry Info": "",
    }


def _bank_record(*, acknowledgement: str = "ACK-001") -> dict[str, object]:
    record = _account_record(acknowledgement=acknowledgement)
    record.pop("Account Number")
    record.pop("Credited Transaction ID")
    return record


def _insert_source(database_path, source_path: str) -> int:
    connection = connect_database(database_path)
    try:
        with connection:
            cursor = connection.execute(
                """
                INSERT INTO source_files (
                    source_path, file_name, file_size, mtime_ns, fingerprint,
                    status, discovered_at, last_seen_at
                ) VALUES (?, ?, 1, 1, '1:1', 'processing', ?, ?)
                """,
                (source_path, source_path, utc_now(), utc_now()),
            )
        return int(cursor.lastrowid)
    finally:
        connection.close()

def test_main_duplicate_filter_returns_credit_counting_view_only() -> None:
    duplicate_with_different_clerical_fields = _main_row(
        serial=99,
        reference="REF-B",
        remarks="copied row with edited notes",
    )
    distinct_transaction = _main_row(
        serial=3,
        credited_transaction_id="CREDIT-002",
    )
    dataframe = pd.DataFrame(
        [
            _main_row(serial=1),
            duplicate_with_different_clerical_fields,
            distinct_transaction,
        ]
    )

    result, removed = app_account.strict_deduplicate_main_transactions(dataframe)

    assert removed == 1
    assert len(dataframe) == 3
    assert len(result) == 2
    assert result.iloc[:, 9].tolist() == ["CREDIT-001", "CREDIT-002"]
    details = result.attrs["duplicate_transaction_details"]
    assert details == [
        {
            "acknowledgement_no": "ACK-001",
            "credited_transaction_id": "CREDIT-001",
            "credited_account_last_four": "4321",
            "duplicate_rows_removed": 1,
            "excluded_disputed_amount": 2500.0,
        }
    ]


def test_complete_duplicate_row_amounts_are_excluded_and_noted(tmp_path) -> None:
    workbook_path = tmp_path / "duplicate-amounts.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Money Transfer"
    worksheet.append([f"Column {index}" for index in range(1, 18)])
    original = _main_row(serial=1)
    duplicate = _main_row(
        serial=2,
        reference="edited reference",
        remarks="later duplicate",
    )
    duplicate[10] = 123456.0
    duplicate[11] = 7777.0
    duplicate[3] = "DEBIT-DUPLICATE-ROW"
    worksheet.append(original)
    worksheet.append(duplicate)

    other_sheet = workbook.create_sheet("Withdrawal Through ATM")
    other_sheet.append([f"Other {index}" for index in range(1, 7)])
    other_sheet.append(
        [1, "unused", "XXXXXX1234", "DEBIT-DUPLICATE-ROW", "unused", 400.0]
    )
    workbook.save(workbook_path)

    try:
        summaries, audit = analyse_with_existing_app(workbook_path)
    finally:
        release_app_memory()

    assert audit["duplicate_transaction_rows_removed"] == 1
    assert audit["main_rows_read"] == 2
    assert audit["main_rows_kept"] == 2
    assert audit["duplicate_other_rows_removed"] == 0
    assert audit["duplicate_processing_version"] == DUPLICATE_SUMMARY_VERSION
    account_rows = {
        str(row["Account Number"]): row
        for row in summaries["Account Wise Summary"]
    }
    credited_account = account_rows["9876544321"]
    debited_account = account_rows["XXXXXX1234"]

    # The later identity is excluded only from credited aggregation. Its debit
    # and other-sheet/recovery effects remain available.
    assert credited_account["Total Credited Amount"] == pytest.approx(2500.0)
    assert debited_account["Total Debited Amount"] == pytest.approx(10277.0)
    assert debited_account["Updated Amount (Recovery)"] == pytest.approx(400.0)
    assert debited_account["Found in Other Sheets"] == "Yes"

    credited_note = credited_account["Duplicate Entry Info"]
    assert "NOT COUNTED in Total Credited Amount" in credited_note
    assert "credited amount excluded: INR 7,777.00" in credited_note
    assert "Debit and other-sheet matching remain counted" in credited_note
    assert debited_account["Duplicate Entry Info"] in (None, "", "None")


def test_only_ack_credited_transaction_and_credited_account_last_four_match() -> None:
    original = _main_row(serial=1)
    same_identity = _main_row(serial=2)
    same_identity[2] = "DIFFERENT-DEBIT-9999"
    same_identity[3] = "DIFFERENT-DEBIT-TRANSACTION"
    same_identity[4] = "Different Bank"
    same_identity[5] = 9
    same_identity[7] = "DIFF0000001"
    same_identity[8] = pd.Timestamp("2026-08-12 23:59:00")
    same_identity[10] = 999999.99
    same_identity[11] = 1.0
    same_identity[12] = "DIFFERENT-REFERENCE"
    different_account = _main_row(serial=3)
    different_account[6] = "9876549999"
    different_ack = _main_row(serial=4)
    different_ack[1] = "ACK-002"
    blank_transaction_one = _main_row(serial=5, credited_transaction_id="")
    blank_transaction_two = _main_row(serial=6, credited_transaction_id="")
    dataframe = pd.DataFrame(
        [
            original,
            same_identity,
            different_account,
            different_ack,
            blank_transaction_one,
            blank_transaction_two,
        ]
    )

    result, removed = app_account.strict_deduplicate_main_transactions(dataframe)

    assert removed == 1
    assert len(result) == 5
    assert app_account.build_transaction_identity(result.iloc[0]) == (
        "ACK-001",
        "CREDIT-001",
        "4321",
    )


def test_other_sheet_rows_are_not_removed_by_transaction_duplicate_rule() -> None:
    dataframe = pd.DataFrame(
        [
            [1, "ACC-1", "TXN-1", 1500.0],
            [88, "ACC-1", "TXN-1", 1500.0],
            [3, "ACC-1", "TXN-2", 1500.0],
        ]
    )

    result, removed = app_account.strict_deduplicate_other_sheet(dataframe)

    assert removed == 0
    assert len(result) == 3


def test_sqlite_drops_identical_summaries_and_rejects_ack_from_other_file(
    tmp_path,
) -> None:
    database_path = tmp_path / "strict.sqlite"
    initialize_database(database_path)
    first_source_id = _insert_source(database_path, "first.xlsx")
    second_source_id = _insert_source(database_path, "second.xlsx")
    account = _account_record()
    bank = _bank_record()
    summaries = {
        "Account Wise Summary": [account, dict(account)],
        "Bank Wise Summary": [bank, dict(bank)],
        "Partial Bank Wise Summary": [],
    }

    counts = save_file_summaries(
        database_path,
        first_source_id,
        summaries,
        duration_seconds=1,
        processing_audit={"duplicate_transaction_rows_removed": 2},
    )

    assert counts["account"] == 1
    assert counts["bank"] == 1
    assert counts["duplicate_summary_rows_removed"] == 2
    connection = connect_database(database_path, readonly=True)
    try:
        assert connection.execute(
            "SELECT COUNT(*) FROM account_summaries"
        ).fetchone()[0] == 1
    finally:
        connection.close()

    conflicting = {
        "Account Wise Summary": [
            {**_account_record(), "Account Number": "XXXXXX9999"}
        ],
        "Bank Wise Summary": [_bank_record()],
        "Partial Bank Wise Summary": [],
    }
    with pytest.raises(ValueError, match="already stored from source file"):
        save_file_summaries(
            database_path,
            second_source_id,
            conflicting,
            duration_seconds=1,
        )


def test_exact_workbook_copy_is_skipped_at_claim_time(tmp_path) -> None:
    input_directory = tmp_path / "input"
    input_directory.mkdir()
    (input_directory / "a.xlsx").write_bytes(b"identical workbook bytes")
    (input_directory / "b.xlsx").write_bytes(b"identical workbook bytes")
    database_path = tmp_path / "strict.sqlite"
    initialize_database(database_path)

    discovery = discover_files(database_path, input_directory)
    assert discovery["seen"] == 2
    assert discovery["hashed"] == 2

    first = claim_next_file(database_path, maximum_attempts=2)
    assert first is not None
    connection = connect_database(database_path)
    try:
        with connection:
            connection.execute(
                "UPDATE source_files SET status = 'completed' WHERE id = ?",
                (first["id"],),
            )
    finally:
        connection.close()

    assert claim_next_file(database_path, maximum_attempts=2) is None
    connection = connect_database(database_path, readonly=True)
    try:
        duplicate = connection.execute(
            """
            SELECT duplicate_of_source_file_id
            FROM source_files
            WHERE id <> ?
            """,
            (first["id"],),
        ).fetchone()
        assert duplicate["duplicate_of_source_file_id"] == first["id"]
    finally:
        connection.close()


def test_migration_removes_only_identical_legacy_summary_rows(tmp_path) -> None:
    database_path = tmp_path / "legacy.sqlite"
    initialize_database(database_path)
    source_id = _insert_source(database_path, "legacy.xlsx")
    connection = connect_database(database_path)
    try:
        with connection:
            connection.execute("DROP INDEX uq_account_summary_identity")
            connection.execute(
                "UPDATE schema_metadata SET value = '2' WHERE key = 'schema_version'"
            )
            values = (
                source_id,
                "ACK-LEGACY",
                "Example Bank",
                "Account-1234",
                "CREDIT-LEGACY",
                1000,
                0,
                0,
                1000,
                "PENDING",
                "No",
                "",
                "",
                utc_now(),
            )
            connection.execute(
                """
                INSERT INTO account_summaries (
                    source_file_id, acknowledgement_no, bank_name,
                    account_number, credited_transaction_id,
                    total_credited_amount, total_debited_amount,
                    updated_amount, not_updated_amount, status,
                    found_in_other_sheets, breakdown_by_sheet,
                    duplicate_entry_info, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                values,
            )
            connection.execute(
                """
                INSERT INTO account_summaries (
                    source_file_id, acknowledgement_no, bank_name,
                    account_number, credited_transaction_id,
                    total_credited_amount, total_debited_amount,
                    updated_amount, not_updated_amount, status,
                    found_in_other_sheets, breakdown_by_sheet,
                    duplicate_entry_info, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    source_id,
                    "ack-legacy",
                    "example bank",
                    "account-1234",
                    "credit-legacy",
                    *values[5:],
                ),
            )
    finally:
        connection.close()

    initialize_database(database_path)
    connection = connect_database(database_path, readonly=True)
    try:
        assert connection.execute(
            "SELECT COUNT(*) FROM account_summaries"
        ).fetchone()[0] == 1
        source = connection.execute(
            """
            SELECT duplicate_summary_rows_removed
            FROM source_files WHERE id = ?
            """,
            (source_id,),
        ).fetchone()
        assert source["duplicate_summary_rows_removed"] == 1
    finally:
        connection.close()


def test_fast_audit_reads_only_requested_duplicate_key_and_queues_file(
    tmp_path,
) -> None:
    input_directory = tmp_path / "input"
    input_directory.mkdir()
    workbook_path = input_directory / "duplicate.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Money Transfer"
    worksheet.append([f"Column {index}" for index in range(1, 11)])
    first = _main_row(serial=1)[:10]
    repeated_key = _main_row(serial=2)[:10]
    repeated_key[2] = "different debit account"
    repeated_key[3] = "different debit transaction"
    repeated_key[4] = "different bank"
    repeated_key[5] = 99
    different_last_four = _main_row(serial=3)[:10]
    different_last_four[6] = "XXXX9999"
    blank_id = _main_row(serial=4, credited_transaction_id="")[:10]
    worksheet.append(first)
    worksheet.append(repeated_key)
    worksheet.append(different_last_four)
    worksheet.append(blank_id)
    worksheet.append(blank_id)
    workbook.save(workbook_path)

    duplicate_rows, data_rows = audit_workbook_duplicate_keys(workbook_path)

    assert duplicate_rows == 1
    assert data_rows == 5

    database_path = tmp_path / "strict.sqlite"
    initialize_database(database_path)
    discover_files(database_path, input_directory)
    result = fast_audit_and_queue_duplicate_files(database_path, workers=2)
    assert result["scanned"] == 1
    assert result["duplicate_files"] == 1
    assert result["duplicate_rows"] == 1
    assert result["queued"] == 1

    # A file processed by the earlier totals-only implementation must be
    # queued once more so its summary receives the excluded-amount note.
    connection = connect_database(database_path)
    try:
        with connection:
            connection.execute(
                """
                UPDATE source_files
                SET status = 'completed',
                    fast_duplicate_reprocessed_version = ?
                """,
                (FAST_DUPLICATE_AUDIT_VERSION,),
            )
    finally:
        connection.close()

    cached = fast_audit_and_queue_duplicate_files(database_path, workers=2)
    assert cached["scanned"] == 0
    assert cached["cached"] == 1
    assert cached["queued"] == 1

    connection = connect_database(database_path)
    try:
        with connection:
            connection.execute(
                """
                UPDATE source_files
                SET status = 'completed',
                    fast_duplicate_reprocessed_version = ?
                """,
                (DUPLICATE_SUMMARY_VERSION,),
            )
    finally:
        connection.close()

    fully_current = fast_audit_and_queue_duplicate_files(database_path, workers=2)
    assert fully_current["scanned"] == 0
    assert fully_current["cached"] == 1
    assert fully_current["queued"] == 0
    connection = connect_database(database_path, readonly=True)
    try:
        source = connection.execute(
            """
            SELECT status, fast_duplicate_audit_version,
                   fast_duplicate_rows_found
            FROM source_files
            """
        ).fetchone()
        assert source["status"] == "completed"
        assert source["fast_duplicate_audit_version"] == (
            FAST_DUPLICATE_AUDIT_VERSION
        )
        assert source["fast_duplicate_rows_found"] == 1
    finally:
        connection.close()
