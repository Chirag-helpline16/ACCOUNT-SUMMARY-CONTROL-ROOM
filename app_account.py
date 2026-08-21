from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import os
import json
import io
import re
from datetime import datetime
from io import BytesIO

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads_account'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Global data storage
df_main = None
df_other_sheets = {}
uploaded_files_count = 0

MONEY_TRANSFER_TO_OTHERS_SHEET = 'Money Transfer to Others'
OTHER_BANK_NAMES = {'OTHER', 'OTHERS'}

# Lookup maps for optimization
debited_acc_map = {}
credited_acc_map = {}
debited_trans_id_map = {}
credited_trans_id_map = {}
breakdown_map = {}

# Details of later transaction rows whose credited amount is excluded by the
# duplicate rule.  The rows themselves stay in df_main because debit totals,
# recovery-sheet matching, flow/status calculations, and transaction lookup
# must continue to see every source row.
last_duplicate_transaction_details = []

# Audit details from the most recently processed workbook.  The overnight
# worker persists these counts in SQLite so duplicate credit exclusion is visible and
# verifiable instead of being a silent calculation detail.
last_processing_audit = {
    'main_rows_read': 0,
    'main_rows_kept': 0,
    'duplicate_transaction_rows_removed': 0,
    'other_rows_read': 0,
    'other_rows_kept': 0,
    'duplicate_other_rows_removed': 0,
}


def _identity_text(value):
    """Normalize an Excel scalar for stable transaction comparisons."""
    if pd.isna(value):
        return ''
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = re.sub(r'\s+', ' ', str(value).strip())
    if text.lower() in ('nan', 'none', 'null'):
        return ''
    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]
    return text.upper()


def _credited_transaction_identity(value):
    """Normalize a credited transaction ID for duplicate matching only.

    Leading zeroes are not significant for the credited-duplicate business
    rule, so values such as ``000123``/``123`` and ``000ABC``/``ABC`` match.
    An all-zero value becomes blank and is not safe to deduplicate.  Callers
    continue to retain the original spreadsheet value for display and for all
    debit/recovery matching.
    """
    return _identity_text(value).lstrip('0')


def _is_money_transfer_to_others_bank(value):
    """Identify Money Transfer rows whose credited bank is Other/Others."""
    return _identity_text(value) in OTHER_BANK_NAMES


def _money_transfer_to_others_frame(dataframe=None):
    """Copy raw Other/Others rows without removing them from Money Transfer."""
    source = df_main if dataframe is None else dataframe
    if source is None:
        return pd.DataFrame()
    if source.empty or len(source.columns) <= 4:
        return source.iloc[0:0].copy()
    mask = source.iloc[:, 4].map(_is_money_transfer_to_others_bank)
    return source.loc[mask].copy()


def _account_last_four(value):
    """Match full and masked versions of the same account by last four."""
    text = _identity_text(value)
    characters = re.sub(r'[^A-Z0-9]', '', text)
    return characters[-4:] if len(characters) > 4 else characters


def build_transaction_identity(row):
    """Return the strict business identity used for duplicate transactions.

    Per the account-summary rule, a repeat requires the same acknowledgement,
    credited transaction ID, and last four characters of the credited account.
    Rows missing any one of those values cannot be identified safely and are
    therefore retained.
    """
    if len(row) < 10:
        return None
    acknowledgement = _identity_text(row.iloc[1])
    credited_account_last_four = _account_last_four(row.iloc[6])
    credited_transaction_id = _credited_transaction_identity(row.iloc[9])
    if not (
        acknowledgement
        and credited_account_last_four
        and credited_transaction_id
    ):
        return None
    return (
        acknowledgement,
        credited_transaction_id,
        credited_account_last_four,
    )


def strict_deduplicate_main_transactions(dataframe):
    """Return the rows allowed to contribute to credited totals.

    This function deliberately returns a filtered *credit-counting view*; it
    must never replace ``df_main``.  Later rows matching ACK + credited
    transaction ID + credited-account last four are excluded only from credit
    aggregation.  Their debit and other-sheet effects remain valid because the
    original dataframe is retained in full.
    """
    if dataframe is None:
        return dataframe, 0
    if dataframe.empty:
        result = dataframe.copy()
        result.attrs['duplicate_transaction_details'] = []
        return result, 0
    seen = set()
    keep_positions = []
    duplicates_removed = 0
    duplicate_groups = {}
    for position, (_, row) in enumerate(dataframe.iterrows()):
        # These raw rows are copied to their own output sheet. They remain in
        # df_main for debit aggregation, but do not participate in credited
        # totals or credited-duplicate counting.
        if len(row) > 4 and _is_money_transfer_to_others_bank(row.iloc[4]):
            continue
        identity = build_transaction_identity(row)
        if identity is None:
            keep_positions.append(position)
            continue
        if identity in seen:
            duplicates_removed += 1
            detail = duplicate_groups.setdefault(
                identity,
                {
                    'acknowledgement_no': identity[0],
                    'credited_transaction_id': identity[1],
                    'credited_account_last_four': identity[2],
                    'duplicate_rows_removed': 0,
                    'excluded_disputed_amount': 0.0,
                },
            )
            detail['duplicate_rows_removed'] += 1
            if len(row) > 11:
                detail['excluded_disputed_amount'] += clean_amount(row.iloc[11])
            continue
        seen.add(identity)
        keep_positions.append(position)

    result = dataframe.iloc[keep_positions].copy()
    result.attrs['duplicate_transaction_details'] = list(
        duplicate_groups.values()
    )
    return result, duplicates_removed


def format_duplicate_credit_note(detail):
    """Describe a duplicate without implying its debit/recovery was removed."""
    duplicate_rows = int(detail.get('duplicate_rows_removed', 0) or 0)
    row_word = 'row' if duplicate_rows == 1 else 'rows'
    return (
        f"First matching credit kept; later {duplicate_rows} duplicate credit "
        f"{row_word} NOT COUNTED in Total Credited Amount. "
        f"Credited TID: {detail.get('credited_transaction_id') or 'N/A'}; "
        "credited A/c last 4: "
        f"{detail.get('credited_account_last_four') or 'N/A'}; "
        "credited amount excluded: INR "
        f"{float(detail.get('excluded_disputed_amount', 0) or 0):,.2f}. "
        "Debit and other-sheet matching remain counted."
    )


def strict_deduplicate_other_sheet(dataframe):
    """Compatibility helper: other-sheet rows are never transaction-deduped."""
    if dataframe is None:
        return dataframe, 0
    return dataframe.copy(), 0

def rebuild_maps():
    """Rebuild lookup maps for faster processing"""
    global debited_acc_map, credited_acc_map, debited_trans_id_map, credited_trans_id_map, breakdown_map
    
    # Rebuild maps for df_main
    debited_acc_map = {}
    credited_acc_map = {}
    debited_trans_id_map = {}
    credited_trans_id_map = {}
    
    if df_main is not None:
        for idx, row in df_main.iterrows():
            # FILTER: Skip rows with null/empty credited transaction ID from all maps
            credited_trans_id_check = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
            if not credited_trans_id_check or credited_trans_id_check.lower() in ('nan', 'none', '', '-', 'null'):
                continue
            
            # Account numbers (col 2 and 6)
            deb_acc = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
            cre_acc = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
            
            if deb_acc:
                if deb_acc not in debited_acc_map:
                    debited_acc_map[deb_acc] = []
                debited_acc_map[deb_acc].append(idx)
            
            if cre_acc:
                if cre_acc not in credited_acc_map:
                    credited_acc_map[cre_acc] = []
                credited_acc_map[cre_acc].append(idx)
                
            # Transaction IDs (col 3 and 9)
            deb_tid = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
            cre_tid = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
            
            if deb_tid:
                if deb_tid not in debited_trans_id_map:
                    debited_trans_id_map[deb_tid] = []
                debited_trans_id_map[deb_tid].append(idx)
            
            if cre_tid:
                if cre_tid not in credited_trans_id_map:
                    credited_trans_id_map[cre_tid] = []
                credited_trans_id_map[cre_tid].append(idx)
    
    # Rebuild breakdown map for other sheets (by Trans ID AND Account Number)
    breakdown_map = {}
    others_less_500_trans_ids = set()  # Track which trans IDs appear in "Others Less Than 500"
    
    if df_other_sheets:
        # First pass: identify transaction IDs in "Others Less Than 500" sheet
        for sheet_name, sheet_info in df_other_sheets.items():
            if 'others less than 500' in sheet_name.lower() or 'others less then 500' in sheet_name.lower():
                df = sheet_info['data']
                if len(df.columns) > 3:
                    for _, row in df.iterrows():
                        trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                        if trans_id:
                            others_less_500_trans_ids.add(trans_id)
        
        # Second pass: build breakdown map with special handling for "Others Less Than 500"
        for sheet_name, sheet_info in df_other_sheets.items():
            df = sheet_info['data']
            amount_col = sheet_info['amount_col']
            
            is_others_less_500 = 'others less than 500' in sheet_name.lower() or 'others less then 500' in sheet_name.lower()
            
            if len(df.columns) > 3:
                processed_trans_ids_in_sheet = set()  # Track which trans IDs we've already added for this sheet
                processed_accounts_in_sheet = set()  # Track which account numbers we've already added for this sheet
                
                for row_idx, row in df.iterrows():
                    trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                    
                    # Also get account number (typically column 2 or 1 depending on sheet structure)
                    # Try column 2 first, then column 1
                    account_no = ''
                    if len(row) > 2:
                        account_no = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                    if not account_no and len(row) > 1:
                        account_no = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
                        
                    # Create shared item to avoid duplicate creation
                    shared_item = None
                    if not is_others_less_500 and len(row) > amount_col:
                        amount = clean_amount(row.iloc[amount_col])
                        if amount > 0:
                            shared_item = {
                                'sheet': sheet_name,
                                'amount': amount,
                                'row_idx': row_idx
                            }
                    
                    # Process by Transaction ID
                    if trans_id:
                        if trans_id not in breakdown_map:
                            breakdown_map[trans_id] = []
                        
                        # Special handling for "Others Less Than 500" sheet
                        if is_others_less_500:
                            # Only add ₹500 once per transaction ID, regardless of how many entries
                            if trans_id not in processed_trans_ids_in_sheet:
                                breakdown_map[trans_id].append({
                                    'sheet': sheet_name,
                                    'amount': 500.0  # Always ₹500 for this sheet
                                })
                                processed_trans_ids_in_sheet.add(trans_id)
                        elif shared_item:
                            # Normal handling for other sheets - add ALL entries (no deduplication)
                            breakdown_map[trans_id].append(shared_item)
                    
                    # ALSO Process by Account Number (NEW LOGIC)
                    if account_no and account_no not in ['nan', 'none', '']:
                        if account_no not in breakdown_map:
                            breakdown_map[account_no] = []
                        
                        # Special handling for "Others Less Than 500" sheet
                        if is_others_less_500:
                            # Only add ₹500 once per account number, regardless of how many entries
                            if account_no not in processed_accounts_in_sheet:
                                breakdown_map[account_no].append({
                                    'sheet': sheet_name,
                                    'amount': 500.0  # Always ₹500 for this sheet
                                })
                                processed_accounts_in_sheet.add(account_no)
                        elif shared_item:
                            # Normal handling for other sheets - add ALL entries (no deduplication)
                            breakdown_map[account_no].append(shared_item)


def clean_amount(value):
    """Convert amount to float"""
    if pd.isna(value) or value == '':
        return 0.0
    if isinstance(value, str):
        value = value.replace(',', '').replace('₹', '').strip()
    try:
        return float(value)
    except:
        return 0.0

def process_excel_file(filepath, is_first_file=False):
    """Process Excel file and merge with existing data"""
    global df_main, df_other_sheets, uploaded_files_count, last_processing_audit
    global last_duplicate_transaction_details

    if is_first_file or df_main is None:
        last_duplicate_transaction_details = []
    last_processing_audit = {
        'main_rows_read': 0,
        'main_rows_kept': 0,
        'duplicate_transaction_rows_removed': 0,
        'other_rows_read': 0,
        'other_rows_kept': 0,
        'duplicate_other_rows_removed': 0,
    }
    
    try:
        xl = pd.ExcelFile(filepath)
        
        print(f"DEBUG: Processing file with sheets: {xl.sheet_names}")
        
        # Find Money Transfer sheet
        money_transfer_sheet = None
        for sheet in xl.sheet_names:
            print(f"DEBUG: Checking sheet '{sheet}', lowercase: '{sheet.lower()}'")
            if 'money transfer' in sheet.lower():
                money_transfer_sheet = sheet
                print(f"DEBUG: Found Money Transfer sheet: '{money_transfer_sheet}'")
                break
        
        if not money_transfer_sheet:
            print(f"DEBUG: Money Transfer sheet NOT FOUND in sheets: {xl.sheet_names}")
            return False, f"Money Transfer sheet not found. Available sheets: {', '.join(xl.sheet_names)}"
        
        # Load main sheet
        df_new_main = pd.read_excel(xl, sheet_name=money_transfer_sheet)
        print(f"DEBUG: Loaded main sheet with {len(df_new_main)} rows")

        # Audit repeated credit identities, but keep every source row.  The
        # filtered view is used only by credited-amount aggregations below;
        # debit totals and recovery/other-sheet matching use the full data.
        last_processing_audit['main_rows_read'] = len(df_new_main)
        credit_counting_rows, duplicates_removed = strict_deduplicate_main_transactions(
            df_new_main
        )
        last_processing_audit['duplicate_transaction_rows_removed'] += duplicates_removed
        last_processing_audit['main_rows_kept'] = len(df_new_main)
        print(
            "DEBUG: Credit-only duplicate audit found "
            f"{duplicates_removed} later credited amount(s) not to count; "
            f"all {len(df_new_main)} rows remain available for debit and recovery"
        )
        
        # Merge or initialize main dataframe
        if is_first_file or df_main is None:
            df_main = df_new_main
            df_other_sheets = {}
        else:
            df_main = pd.concat([df_main, df_new_main], ignore_index=True)

        # Recompute the explanation against the complete upload set.  Do not
        # assign the filtered credit view back to df_main.
        credit_counting_rows, all_duplicate_count = (
            strict_deduplicate_main_transactions(df_main)
        )
        last_duplicate_transaction_details = credit_counting_rows.attrs.get(
            'duplicate_transaction_details', []
        )
        last_processing_audit['duplicate_transaction_rows_removed'] = (
            all_duplicate_count
        )
        
        # Reset index after merging uploads; no source rows were dropped.
        df_main = df_main.reset_index(drop=True)
        
        # Load and merge other sheets
        for sheet in xl.sheet_names:
            if sheet == money_transfer_sheet:
                continue
            
            df_sheet = pd.read_excel(xl, sheet_name=sheet)
            last_processing_audit['other_rows_read'] += len(df_sheet)
            
            # Determine amount column based on sheet name
            sheet_lower = sheet.lower()
            if 'withdrawal through atm' in sheet_lower:
                amount_col_idx = 5  # Column F (Withdrawal Amount)
            elif 'cash withdrawal through cheque' in sheet_lower:
                amount_col_idx = 9  # Column J
            elif 'withdrawal through pos' in sheet_lower:
                amount_col_idx = 6  # Column G
            elif 'cheque' in sheet_lower:
                amount_col_idx = 8  # Column I (for other cheque sheets)
            else:
                amount_col_idx = 5  # Column F (default)
            
            # Merge with existing sheet data or create new
            if sheet in df_other_sheets:
                df_other_sheets[sheet]['data'] = pd.concat(
                    [df_other_sheets[sheet]['data'], df_sheet],
                    ignore_index=True,
                )
            else:
                df_other_sheets[sheet] = {
                    'data': df_sheet,
                    'amount_col': amount_col_idx
                }

        last_processing_audit['other_rows_kept'] = sum(
            len(sheet_info['data']) for sheet_info in df_other_sheets.values()
        )
        
        uploaded_files_count += 1
        rebuild_maps()
        return True, f"File processed successfully. Total files: {uploaded_files_count}"
        
    except Exception as e:
        return False, f"Error: {str(e)}"

def get_transaction_breakdown(trans_id, account_no=None):
    """Get breakdown from other sheets - ONLY NON-ZERO amounts - Optimized with breakdown_map
    
    Searches by:
    1. Transaction ID (trans_id)
    2. Account Number (account_no) if provided
    
    Returns combined results from both lookups (deduplicated by sheet+amount)
    """
    breakdown_items = []
    seen_items = set()  # Track (sheet, amount) to avoid duplicates
    
    # Search by Transaction ID
    if trans_id:
        trans_id_str = str(trans_id).strip()
        if trans_id_str in breakdown_map:
            for item in breakdown_map[trans_id_str]:
                sheet_lower = item['sheet'].lower()
                if 'atm' in sheet_lower:
                    key = (item['sheet'], item.get('row_idx', id(item)))
                else:
                    key = (item['sheet'], item['amount'])
                if key not in seen_items:
                    breakdown_items.append(item)
                    seen_items.add(key)
    
    # Search by Account Number (if provided)
    if account_no:
        account_no_str = str(account_no).strip()
        if account_no_str and account_no_str not in ['nan', 'none', '']:
            if account_no_str in breakdown_map:
                for item in breakdown_map[account_no_str]:
                    sheet_lower = item['sheet'].lower()
                    if 'atm' in sheet_lower:
                        key = (item['sheet'], item.get('row_idx', id(item)))
                    else:
                        key = (item['sheet'], item['amount'])
                    if key not in seen_items:
                        breakdown_items.append(item)
                        seen_items.add(key)
    
    return breakdown_items

def calculate_status(child_trans_id, disputed_amount, current_layer, account_no=None):
    """Calculate transaction status - accounts for BOTH children flow AND other-sheet recovery
    
    Args:
        child_trans_id: Transaction ID to check
        disputed_amount: Amount in dispute
        current_layer: Current layer number
        account_no: Account number (optional) - used to search in other sheets
    """
    if df_main is None:
        return {
            'status': 'PENDING',
            'updated_amount': 0.0,
            'pending_amount': float(disputed_amount) if disputed_amount else 0.0,
            'has_children': False,
            'in_other_sheets': False,
            'children_total': 0.0
        }
    
    # Ensure disputed_amount is a valid number
    disputed_amount = float(disputed_amount) if disputed_amount and not pd.isna(disputed_amount) else 0.0
    
    # Optimized child lookup by transaction ID (for standard status)
    has_children = False
    children_total = 0.0
    child_id_str = str(child_trans_id).strip()
    if child_id_str in debited_trans_id_map:
        for idx in debited_trans_id_map[child_id_str]:
            if df_main.iloc[idx, 5] > current_layer:
                has_children = True
                children_total += clean_amount(df_main.iloc[idx, 11])

    # Get breakdown from breakdown_map (searches by BOTH trans_id AND account_no)
    breakdown = get_transaction_breakdown(child_trans_id, account_no)
    updated_amount = float(sum(item['amount'] for item in breakdown))
    in_other_sheets = len(breakdown) > 0
    
    # Total accounted = money that flowed to children + money recovered via other sheets
    total_accounted = children_total + updated_amount
    pending_amount = disputed_amount - total_accounted
    
    # Status logic
    if not has_children and not in_other_sheets:
        status = 'PENDING'
    elif pending_amount > 0.01:  # Allow small rounding errors
        status = 'PARTIAL'
    else:
        status = 'COMPLETE'
    
    return {
        'status': status,
        'updated_amount': float(updated_amount),
        'pending_amount': float(max(0, pending_amount)),
        'has_children': has_children,
        'in_other_sheets': in_other_sheets,
        'children_total': float(children_total)
    }

def get_layer_transactions(layer=1, parent_trans_id=None):
    """Get transactions for a layer"""
    if df_main is None:
        return []
    
    try:
        if layer == 1 and parent_trans_id is None:
            # Root level: start from minimum layer in dataset
            # Filter out NaN values first
            valid_layers = df_main[pd.notna(df_main.iloc[:, 5])]
            if len(valid_layers) == 0:
                print("ERROR: No valid layer values found")
                return []
            
            min_layer = int(valid_layers.iloc[:, 5].min())
            print(f"DEBUG: Minimum layer found: {min_layer}")
            print(f"DEBUG: Total rows in df_main: {len(df_main)}")
            filtered = df_main[df_main.iloc[:, 5] == min_layer]
            print(f"DEBUG: Filtered rows for layer {min_layer}: {len(filtered)}")
        else:
            # Get parent's layer first
            parent_rows = df_main[df_main.iloc[:, 9].astype(str).str.strip() == str(parent_trans_id).strip()]
            if len(parent_rows) == 0:
                return []
            parent_layer = int(parent_rows.iloc[0, 5])
            
            # Children: Parent Trans ID (col D) = parent_trans_id AND layer > parent_layer
            filtered = df_main[
                (df_main.iloc[:, 3].astype(str).str.strip() == str(parent_trans_id).strip()) & 
                (df_main.iloc[:, 5] > parent_layer)
            ]
    except Exception as e:
        print(f"ERROR in get_layer_transactions: {e}")
        return []
    
    transactions = []
    for _, row in filtered.iterrows():
        child_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
        
        # FILTER: Skip rows with null/empty credited transaction ID
        if not child_trans_id or child_trans_id.lower() in ('nan', 'none', '', '-', 'null'):
            continue
            
        credited_account = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
        disputed_amount = clean_amount(row.iloc[11])
        current_layer = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 1
        
        # Pass account number to calculate_status for better recovery detection
        status_info = calculate_status(child_trans_id, disputed_amount, current_layer, credited_account)
        
        # Override status to match Excel logic: if has children = "TRANSACTION CONTINUE", else use calculated status
        if status_info['has_children']:
            final_status = 'TRANSACTION CONTINUE'
        else:
            final_status = status_info['status']
        
        # Get child bank (Optimized)
        child_bank = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ''
        child_id_str = str(child_trans_id).strip()
        if child_id_str in debited_trans_id_map:
            for c_idx in debited_trans_id_map[child_id_str]:
                if df_main.iloc[c_idx, 5] > current_layer:
                    child_bank = str(df_main.iloc[c_idx, 4])
                    break
        
        transactions.append({
            's_no': int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0,
            'acknowledgement_no': str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
            'debited_account': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
            'debited_transaction_id': str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
            'bank': str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
            'layer': current_layer,
            'credited_account': str(row.iloc[6]) if pd.notna(row.iloc[6]) else '',
            'ifsc_code': str(row.iloc[7]) if pd.notna(row.iloc[7]) else '',
            'transaction_date': str(row.iloc[8]) if pd.notna(row.iloc[8]) else '',
            'credited_transaction_id': child_trans_id,
            'transaction_amount': float(clean_amount(row.iloc[10])),
            'disputed_amount': float(disputed_amount),
            'reference_no': str(row.iloc[12]) if pd.notna(row.iloc[12]) else '',
            'remarks': str(row.iloc[13]) if pd.notna(row.iloc[13]) else '',
            'action_taken_by': str(row.iloc[14]) if pd.notna(row.iloc[14]) else '',
            'date_of_action': str(row.iloc[15]) if pd.notna(row.iloc[15]) else '',
            'credited_bank': child_bank,
            'status': final_status,
            'updated_amount': float(status_info['updated_amount']),
            'pending_amount': float(status_info['pending_amount']),
            'has_children': status_info['has_children'],
            'in_other_sheets': status_info['in_other_sheets']
        })
    
    return transactions

def build_hierarchical_data(parent_id=None, layer=1, parent_path="", level=0, visited=None):
    """Build hierarchical data for Excel - INCLUDES ALL TRANSACTIONS - ACCOUNT NUMBER BASED WITH MERGING"""
    if df_main is None:
        return []
    
    # ACCOUNT NUMBER BASED WITH MERGING: Group rows by account pairs AND status, sum amounts
    if parent_id is None and layer == 1:
        all_data = []
        processed_row_indices = set()
        
        # First, group rows by (Debited Account, Credited Account) WITHOUT status to calculate base status
        temp_grouped = {}
        for idx, row in df_main.iterrows():
            # FILTER: Skip rows with null/empty credited transaction ID
            credited_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
            if not credited_trans_id or credited_trans_id.lower() in ('nan', 'none', '', '-', 'null'):
                continue
            debited_account = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
            credited_account = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
            child_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
            
            key = (debited_account, credited_account)
            if key not in temp_grouped:
                temp_grouped[key] = []
            temp_grouped[key].append({
                'idx': idx,
                'row': row,
                'child_trans_id': child_trans_id,
                'credited_account': credited_account
            })
        
        # Calculate status for each individual row
        row_status_map = {}
        for idx, row in df_main.iterrows():
            # FILTER: Skip rows with null/empty credited transaction ID
            credited_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
            if not credited_trans_id or credited_trans_id.lower() in ('nan', 'none', '', '-', 'null'):
                continue
                
            child_trans_id = credited_trans_id
            disputed_amount = clean_amount(row.iloc[11])
            current_layer = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 1
            credited_account = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
            
            # Calculate base status from other sheets (pass account number for better recovery detection)
            status_info = calculate_status(child_trans_id, disputed_amount, current_layer, credited_account)
            
            # Optimized: Check if this credited account appears as a debited account in ANY other row
            has_children = False
            credited_acc_str = str(credited_account).strip()
            if credited_acc_str in debited_acc_map:
                for check_idx in debited_acc_map[credited_acc_str]:
                    if check_idx != idx:
                        has_children = True
                        break
            
            # Determine final status
            if has_children:
                final_status = 'TRANSACTION CONTINUE'
            else:
                final_status = status_info['status']
            
            row_status_map[idx] = final_status
        
        # Now group by (Debited Account, Credited Account, Status)
        # IMPORTANT: Track unique transactions to avoid double-counting duplicates
        grouped_data = {}
        for idx, row in df_main.iterrows():
            # FILTER: Skip rows with null/empty credited transaction ID
            credited_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
            if not credited_trans_id or credited_trans_id.lower() in ('nan', 'none', '', '-', 'null'):
                continue
                
            debited_account = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
            credited_account = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
            status = row_status_map.get(idx, 'PENDING')  # Use .get() since we might have skipped this row
            
            # Key includes status for proper grouping
            key = (debited_account, credited_account, status)
            
            if key not in grouped_data:
                grouped_data[key] = {
                    'indices': [],
                    'rows': [],
                    'total_disputed': 0,
                    'total_transaction': 0,
                    'status': status,
                    'seen_transactions': set()  # Track unique transaction signatures
                }
            
            # Create a signature for this transaction to detect true duplicates
            # Match: last 4 digits of account numbers + all other fields
            debited_trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
            bank = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
            layer = str(int(row.iloc[5])) if pd.notna(row.iloc[5]) else ''
            ifsc = str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else ''
            trans_date = str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else ''
            disputed_amt = clean_amount(row.iloc[11])
            transaction_amt = clean_amount(row.iloc[10])
            
            # Get last 4 digits of account numbers
            debited_acc_last4 = debited_account[-4:] if len(debited_account) >= 4 else debited_account
            credited_acc_last4 = credited_account[-4:] if len(credited_account) >= 4 else credited_account
            
            # Create signature with last 4 digits of accounts + all other fields
            trans_signature = f"{debited_acc_last4}|{debited_trans_id}|{bank}|{layer}|{credited_acc_last4}|{ifsc}|{trans_date}|{credited_trans_id}|{transaction_amt}|{disputed_amt}"
            
            # Only add amounts if this exact transaction hasn't been seen before
            # This prevents double-counting true duplicates while allowing multiple
            # legitimate transactions between the same account pair
            if trans_signature not in grouped_data[key]['seen_transactions']:
                grouped_data[key]['seen_transactions'].add(trans_signature)
                grouped_data[key]['total_disputed'] += disputed_amt
                grouped_data[key]['total_transaction'] += transaction_amt
            else:
                print(f"DEBUG: Skipping duplicate transaction: {trans_signature}")
            
            grouped_data[key]['indices'].append(idx)
            grouped_data[key]['rows'].append(row)
        
        # Build parent-child map based on grouped data: Key = Credited Account, Value = list of account pair keys (with status)
        parent_child_map = {}
        for (debited_acc, credited_acc, status), data in grouped_data.items():
            if debited_acc:
                if debited_acc not in parent_child_map:
                    parent_child_map[debited_acc] = []
                parent_child_map[debited_acc].append((debited_acc, credited_acc, status))
        
        # Find root accounts: Credited Account NOT in any Debited Account
        all_credited_accounts = set(credited_acc for (_, credited_acc, _) in grouped_data.keys() if credited_acc)
        all_debited_accounts = set(debited_acc for (debited_acc, _, _) in grouped_data.keys() if debited_acc)
        root_credited_accounts = all_credited_accounts - all_debited_accounts
        
        # Process root account pairs
        for (debited_acc, credited_acc, status), data in grouped_data.items():
            if credited_acc in root_credited_accounts:
                for idx in data['indices']:
                    processed_row_indices.add(idx)
                
                # Create merged row data
                row_data = process_merged_account_row(data, debited_acc, credited_acc, level=0, 
                                                     parent_path=credited_acc, parent_child_map=parent_child_map,
                                                     grouped_data=grouped_data)
                all_data.append(row_data)
                
                # Add children recursively
                children_data = process_children_merged_account(credited_acc, level=1, parent_path=credited_acc, 
                                                               parent_child_map=parent_child_map, 
                                                               processed_row_indices=processed_row_indices,
                                                               grouped_data=grouped_data)
                all_data.extend(children_data)
        
        # Process remaining account pairs (orphans, circular refs)
        for (debited_acc, credited_acc, status), data in grouped_data.items():
            # Check if any index from this group is unprocessed
            has_unprocessed = any(idx not in processed_row_indices for idx in data['indices'])
            
            if has_unprocessed:
                for idx in data['indices']:
                    processed_row_indices.add(idx)
                
                # Build path
                if debited_acc and credited_acc:
                    path = f"{debited_acc} → {credited_acc}"
                elif credited_acc:
                    path = credited_acc
                else:
                    path = debited_acc if debited_acc else "No Account"
                
                row_data = process_merged_account_row(data, debited_acc, credited_acc, level=0, 
                                                     parent_path=path, parent_child_map=parent_child_map,
                                                     grouped_data=grouped_data)
                all_data.append(row_data)
                
                # Try to add children
                if credited_acc:
                    children_data = process_children_merged_account(credited_acc, level=1, parent_path=path, 
                                                                   parent_child_map=parent_child_map, 
                                                                   processed_row_indices=processed_row_indices,
                                                                   grouped_data=grouped_data)
                    all_data.extend(children_data)
        
        print(f"DEBUG: Processed {len(processed_row_indices)} rows out of {len(df_main)} total rows")
        print(f"DEBUG: Output has {len(all_data)} merged rows")
        
        return all_data
    
    return []

def process_merged_account_row(data, debited_acc, credited_acc, level, parent_path, parent_child_map, grouped_data):
    """Process merged account rows and return aggregated data"""
    # Use first row for non-numeric fields
    first_row = data['rows'][0]
    
    # Get transaction IDs (concatenate if multiple)
    trans_ids = [str(row.iloc[9]).strip() for row in data['rows'] if pd.notna(row.iloc[9])]
    child_trans_id = '; '.join(set(trans_ids)) if trans_ids else ''
    
    disputed_amount = data['total_disputed']
    transaction_amount = data['total_transaction']
    current_layer = int(first_row.iloc[5]) if pd.notna(first_row.iloc[5]) else 1
    
    # Use the pre-calculated status from grouped data
    final_status = data['status']
    
    # Check if this account has children
    has_children_rows = credited_acc in parent_child_map and len(parent_child_map[credited_acc]) > 0
    
    # Calculate children amount difference based on accounts
    if has_children_rows:
        children_total = 0
        for child_key in parent_child_map[credited_acc]:
            if child_key in grouped_data:
                children_total += grouped_data[child_key]['total_disputed']
        amount_difference = disputed_amount - children_total
    else:
        amount_difference = None
    
    # Get breakdown for all transaction IDs AND account number
    breakdown_items = []
    for trans_id in trans_ids:
        breakdown = get_transaction_breakdown(trans_id, credited_acc)
        breakdown_items.extend(breakdown)
    
    # Aggregate breakdown by sheet
    breakdown_by_sheet = {}
    for item in breakdown_items:
        sheet = item['sheet']
        if sheet not in breakdown_by_sheet:
            breakdown_by_sheet[sheet] = 0
        breakdown_by_sheet[sheet] += item['amount']
    
    breakdown_text = "; ".join([f"{sheet}: ₹{amt:,.2f}" for sheet, amt in breakdown_by_sheet.items()]) if breakdown_by_sheet else "None"
    total_updated = sum(breakdown_by_sheet.values())
    
    child_bank = str(first_row.iloc[4]) if pd.notna(first_row.iloc[4]) else ''
    if has_children_rows and len(parent_child_map[credited_acc]) > 0:
        first_child_key = parent_child_map[credited_acc][0]
        if first_child_key in grouped_data:
            first_child_row = grouped_data[first_child_key]['rows'][0]
            child_bank = str(first_child_row.iloc[4])
    
    # Concatenate other fields
    acknowledgement_nos = [str(row.iloc[1]) for row in data['rows'] if pd.notna(row.iloc[1])]
    debited_trans_ids = [str(row.iloc[3]) for row in data['rows'] if pd.notna(row.iloc[3])]
    
    row_data = {
        'Level': level,
        'Layer': current_layer,
        'Debited Account (Filter)': debited_acc,
        'Hierarchy Path': parent_path,
        'Transaction Count': len(data['rows']),
        'S.No': ', '.join([str(int(row.iloc[0])) for row in data['rows'] if pd.notna(row.iloc[0])]),
        'Debited Account': debited_acc,
        'Debited Transaction ID': '; '.join(set(debited_trans_ids)),
        'Debited Bank': str(first_row.iloc[4]) if pd.notna(first_row.iloc[4]) else '',
        'Acknowledgement No': '; '.join(set(acknowledgement_nos)),
        'Credited Account': credited_acc,
        'Credited Transaction ID': child_trans_id,
        'Credited Bank': child_bank,
        'IFSC Code': str(first_row.iloc[7]) if pd.notna(first_row.iloc[7]) else '',
        'Transaction Date': str(first_row.iloc[8]) if pd.notna(first_row.iloc[8]) else '',
        'Transaction Amount': transaction_amount,
        'Disputed Amount': disputed_amount,
        'Layerwise Amount Difference': amount_difference if amount_difference is not None else '',
        'Updated Amount': total_updated,
        'Pending Amount': max(0, disputed_amount - total_updated),
        'Status': final_status,
        'Has Children': 'Yes' if has_children_rows else 'No',
        'Found in Other Sheets': 'Yes' if total_updated > 0 else 'No',
        'Breakdown by Sheet': breakdown_text,
        'Reference No': str(first_row.iloc[12]) if pd.notna(first_row.iloc[12]) else '',
        'Remarks': str(first_row.iloc[13]) if pd.notna(first_row.iloc[13]) else '',
        'Action Taken By': str(first_row.iloc[14]) if pd.notna(first_row.iloc[14]) else '',
        'Date of Action': str(first_row.iloc[15]) if pd.notna(first_row.iloc[15]) else ''
    }
    
    return row_data

def process_children_merged_account(parent_account, level, parent_path, parent_child_map, processed_row_indices, grouped_data, max_depth=20):
    """Recursively process children with merged accounts"""
    if level > max_depth:
        return []
    
    children_data = []
    
    if parent_account in parent_child_map:
        for child_key in parent_child_map[parent_account]:
            # child_key is now a 3-tuple: (debited_acc, credited_acc, status)
            (debited_acc, credited_acc, status) = child_key
            
            if child_key not in grouped_data:
                continue
            
            data = grouped_data[child_key]
            
            # Check if any index from this group is unprocessed
            has_unprocessed = any(idx not in processed_row_indices for idx in data['indices'])
            
            if not has_unprocessed:
                continue
            
            for idx in data['indices']:
                processed_row_indices.add(idx)
            
            # Build the child's path
            child_path = f"{parent_path} → {credited_acc}" if credited_acc else parent_path
            
            # Process merged row
            row_data = process_merged_account_row(data, debited_acc, credited_acc, level=level, 
                                                 parent_path=child_path, parent_child_map=parent_child_map,
                                                 grouped_data=grouped_data)
            children_data.append(row_data)
            
            # Recursively process grandchildren
            grandchildren_data = process_children_merged_account(credited_acc, level=level+1, 
                                                                parent_path=child_path,
                                                                parent_child_map=parent_child_map, 
                                                                processed_row_indices=processed_row_indices,
                                                                grouped_data=grouped_data,
                                                                max_depth=max_depth)
            children_data.extend(grandchildren_data)
    
    return children_data

@app.route('/')
def index():
    response = render_template('index.html')
    return response, 200, {
        'Cache-Control': 'no-cache, no-store, must-revalidate',
        'Pragma': 'no-cache',
        'Expires': '0'
    }

@app.route('/static/notification.wav')
def serve_notification():
    """Serve the notification sound file"""
    from flask import send_file
    import os
    wav_path = os.path.join(os.getcwd(), 'notification.wav')
    if os.path.exists(wav_path):
        return send_file(wav_path, mimetype='audio/wav')
    else:
        return "Sound file not found", 404

@app.route('/test')
def test():
    return render_template('test.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    global uploaded_files_count
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'})
    
    files = request.files.getlist('file')
    
    if not files or files[0].filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if len(files) > 100:
        return jsonify({'success': False, 'message': 'Maximum 100 files allowed'})
    
    # Check if this is a fresh upload (clear previous data)
    clear_data = request.form.get('clear_data', 'true').lower() == 'true'
    
    if clear_data:
        uploaded_files_count = 0
    
    success_count = 0
    error_messages = []
    
    for idx, file in enumerate(files):
        if not file.filename.endswith(('.xlsx', '.xls')):
            error_messages.append(f"{file.filename}: Only Excel files allowed")
            continue
        
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], f'file_{idx}_{file.filename}')
        file.save(filepath)
        
        is_first = (idx == 0 and clear_data)
        success, message = process_excel_file(filepath, is_first_file=is_first)
        
        if success:
            success_count += 1
        else:
            error_messages.append(f"{file.filename}: {message}")
    
    if success_count > 0:
        msg = f"✓ Successfully processed {success_count} file(s). Total files loaded: {uploaded_files_count}"
        if error_messages:
            msg += f"\n⚠ Errors: {'; '.join(error_messages)}"
        return jsonify({'success': True, 'message': msg})
    else:
        return jsonify({'success': False, 'message': f"Failed to process files. Errors: {'; '.join(error_messages)}"})

@app.route('/clear-data', methods=['POST'])
def clear_data():
    """Clear all uploaded data"""
    global df_main, df_other_sheets, uploaded_files_count
    df_main = None
    df_other_sheets = {}
    uploaded_files_count = 0
    rebuild_maps()
    return jsonify({'success': True, 'message': 'All data cleared'})

@app.route('/api/min-layer')
def get_min_layer():
    """Get the minimum layer in the dataset"""
    if df_main is None:
        return jsonify({'min_layer': 1})
    
    try:
        valid_layers = df_main[pd.notna(df_main.iloc[:, 5])]
        if len(valid_layers) == 0:
            return jsonify({'min_layer': 1})
        min_layer = int(valid_layers.iloc[:, 5].min())
        return jsonify({'min_layer': min_layer})
    except:
        return jsonify({'min_layer': 1})

@app.route('/api/excluded-records')
def get_excluded_records():
    """Get records with null/empty credited transaction ID (excluded from main views)"""
    if df_main is None:
        return jsonify({'records': [], 'count': 0, 'total_amount': 0})
    
    excluded_records = []
    total_amount = 0
    
    for idx, row in df_main.iterrows():
        credited_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
        
        # Include only rows with null/empty credited transaction ID
        if not credited_trans_id or credited_trans_id.lower() in ('nan', 'none', '', '-', 'null'):
            disputed_amount = clean_amount(row.iloc[11])
            total_amount += disputed_amount
            
            excluded_records.append({
                's_no': int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0,
                'acknowledgement_no': str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
                'debited_account': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
                'debited_transaction_id': str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
                'bank': str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
                'layer': int(row.iloc[5]) if pd.notna(row.iloc[5]) else 0,
                'credited_account': str(row.iloc[6]) if pd.notna(row.iloc[6]) else '',
                'ifsc_code': str(row.iloc[7]) if pd.notna(row.iloc[7]) else '',
                'transaction_date': str(row.iloc[8]) if pd.notna(row.iloc[8]) else '',
                'credited_transaction_id': '',  # Always empty for excluded records
                'transaction_amount': float(clean_amount(row.iloc[10])),
                'disputed_amount': float(disputed_amount),
                'reference_no': str(row.iloc[12]) if pd.notna(row.iloc[12]) else '',
                'remarks': str(row.iloc[13]) if pd.notna(row.iloc[13]) else '',
                'action_taken_by': str(row.iloc[14]) if pd.notna(row.iloc[14]) else '',
                'date_of_action': str(row.iloc[15]) if pd.notna(row.iloc[15]) else ''
            })
    
    return jsonify({
        'records': excluded_records,
        'count': len(excluded_records),
        'total_amount': total_amount
    })


@app.route('/api/other-sheets-total')
def get_other_sheets_total():
    """Get total amount from all other sheets (ATM, Cheque, POS, Frozen, Others)"""
    if not df_other_sheets:
        return jsonify({
            'total': 0,
            'breakdown': {},
            'count': 0
        })
    
    total = 0
    breakdown = {}
    total_count = 0
    
    for sheet_name, sheet_info in df_other_sheets.items():
        df = sheet_info['data']
        amount_col = sheet_info['amount_col']
        sheet_lower = sheet_name.lower()
        
        # Determine which amount column to use
        if 'withdrawal through atm' in sheet_lower:
            # For ATM, use Disputed Amount (column 6)
            use_col = 6
        elif 'others less' in sheet_lower:
            # For Others, each entry is ₹500
            sheet_total = len(df) * 500.0
            breakdown[sheet_name] = {
                'total': sheet_total,
                'count': len(df)
            }
            total += sheet_total
            total_count += len(df)
            continue
        else:
            # For other sheets, try to find disputed amount column
            # First check if there's a "Disputed Amount" column
            disputed_col = None
            for idx, col in enumerate(df.columns):
                if 'disputed' in str(col).lower():
                    disputed_col = idx
                    break
            
            use_col = disputed_col if disputed_col is not None else amount_col
        
        # Calculate total for this sheet
        sheet_total = 0
        count = 0
        for idx, row in df.iterrows():
            if len(row) > use_col:
                amount = clean_amount(row.iloc[use_col])
                if amount > 0:
                    sheet_total += amount
                    count += 1
        
        if sheet_total > 0:
            breakdown[sheet_name] = {
                'total': sheet_total,
                'count': count
            }
            total += sheet_total
            total_count += count
    
    return jsonify({
        'total': total,
        'breakdown': breakdown,
        'count': total_count
    })


@app.route('/api/sheet-details/<sheet_name>')
def get_sheet_details(sheet_name):
    """Get all transactions from a specific other sheet"""
    if not df_other_sheets or sheet_name not in df_other_sheets:
        return jsonify({'error': 'Sheet not found', 'transactions': []})
    
    sheet_info = df_other_sheets[sheet_name]
    df = sheet_info['data']
    amount_col = sheet_info['amount_col']
    sheet_lower = sheet_name.lower()
    
    # Determine which amount column to use
    if 'withdrawal through atm' in sheet_lower:
        use_col = 6  # Disputed Amount
    elif 'others less' in sheet_lower:
        use_col = None  # Fixed ₹500
    else:
        # Try to find disputed amount column
        disputed_col = None
        for idx, col in enumerate(df.columns):
            if 'disputed' in str(col).lower():
                disputed_col = idx
                break
        use_col = disputed_col if disputed_col is not None else amount_col
    
    transactions = []
    for idx, row in df.iterrows():
        # Get basic info
        trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) and len(row) > 3 else ''
        account = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) and len(row) > 2 else ''
        
        # Get amount
        if use_col is None:
            amount = 500.0
        elif len(row) > use_col:
            amount = clean_amount(row.iloc[use_col])
        else:
            amount = 0
        
        if amount <= 0:
            continue
        
        # Get bank name
        bank = ''
        if 'withdrawal through atm' in sheet_lower and len(row) > 11:
            bank = str(row.iloc[11]).strip() if pd.notna(row.iloc[11]) else ''
        elif len(row) > 6:
            bank = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
        
        if not bank:
            bank = 'Unknown Bank'
        
        # Get extra info
        extra = ''
        if 'withdrawal through atm' in sheet_lower and len(row) > 8:
            location = str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else ''
            if location:
                extra = location
        elif 'pos' in sheet_lower and len(row) > 8:
            merchant = str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else ''
            if merchant:
                extra = merchant
        elif 'others' in sheet_lower:
            # For Others sheets, check column H (column 7) for remarks/reference
            if len(row) > 7:
                remarks = str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else ''
                if remarks and remarks.lower() not in ['nan', 'none', '']:
                    extra = remarks
        
        # If no extra info found yet, try to find remarks or reference columns by name
        if not extra:
            # Try to find Remarks column
            for col_idx, col_name in enumerate(df.columns):
                if 'remark' in str(col_name).lower() and len(row) > col_idx:
                    remarks = str(row.iloc[col_idx]).strip() if pd.notna(row.iloc[col_idx]) else ''
                    if remarks and remarks.lower() not in ['nan', 'none', '']:
                        extra = remarks
                        break
            
            # If still no extra info, try Reference No column
            if not extra:
                for col_idx, col_name in enumerate(df.columns):
                    if 'reference' in str(col_name).lower() and len(row) > col_idx:
                        ref = str(row.iloc[col_idx]).strip() if pd.notna(row.iloc[col_idx]) else ''
                        if ref and ref.lower() not in ['nan', 'none', '']:
                            extra = f"Ref: {ref}"
                            break
        
        # Try to find IFSC code in any column
        ifsc = ''
        for col_idx, col_name in enumerate(df.columns):
            if 'ifsc' in str(col_name).lower() and len(row) > col_idx:
                ifsc = str(row.iloc[col_idx]).strip() if pd.notna(row.iloc[col_idx]) else ''
                break
        
        # Get date if available
        date = ''
        for col_idx, col_name in enumerate(df.columns):
            if 'date' in str(col_name).lower() and 'action' not in str(col_name).lower() and len(row) > col_idx:
                date = str(row.iloc[col_idx]).strip() if pd.notna(row.iloc[col_idx]) else ''
                if date and date.lower() not in ['nan', 'none', '']:
                    break
        
        transactions.append({
            'trans_id': trans_id,
            'account': account,
            'bank': bank,
            'amount': amount,
            'extra': extra,
            'ifsc': ifsc,
            'date': date
        })
    
    return jsonify({
        'sheet_name': sheet_name,
        'transactions': transactions,
        'count': len(transactions)
    })


@app.route('/api/transactions')
def get_transactions():
    layer = int(request.args.get('layer', 1))
    parent_id = request.args.get('parent_id', None)
    
    transactions = get_layer_transactions(layer, parent_id)
    return jsonify(transactions)

@app.route('/api/all-transactions')
def get_all_transactions():
    """Get ALL transactions from the dataset for filtering - ACCOUNT-WISE SUMMARY"""
    if df_main is None:
        return jsonify([])

    # Get unique acknowledgement numbers
    all_acks = df_main.iloc[:, 1].astype(str).str.strip().unique()
    all_acks = [a for a in all_acks if a and a.lower() not in ('nan', 'none', '')]

    if not all_acks:
        all_acks = ['UNKNOWN']

    all_account_summary = []

    # Helper function to strip leading zeros from account numbers
    def strip_account(x):
        if pd.isna(x): return ''
        s = str(x).strip()
        if s.lower() in ('nan', 'none', ''): return ''
        if s.endswith('.0'):
            s = s[:-2]
        return s.lstrip('0') or '0'

    debited_series_full = df_main.iloc[:, 2].apply(strip_account)
    credited_series_full = df_main.iloc[:, 6].apply(strip_account)

    for current_ack in sorted(all_acks):
        if current_ack != 'UNKNOWN':
            ack_mask = df_main.iloc[:, 1].astype(str).str.strip() == current_ack
            df_ack_main = df_main[ack_mask]
            debited_series = debited_series_full[ack_mask]
            credited_series = credited_series_full[ack_mask]
        else:
            df_ack_main = df_main
            debited_series = debited_series_full
            credited_series = credited_series_full

        if df_ack_main.empty:
            continue

        credit_counting_df, _ = strict_deduplicate_main_transactions(
            df_ack_main
        )
        credit_counting_series = credit_counting_df.iloc[:, 6].apply(
            strip_account
        )

        credited_bank_is_countable = ~df_ack_main.iloc[:, 4].map(
            _is_money_transfer_to_others_bank
        )

        all_accounts = set(debited_series[debited_series != ''].unique()) | set(
            credited_series[
                (credited_series != '') & credited_bank_is_countable
            ].unique()
        )

        # Store the longest string variant for each account
        display_names = {}
        for col_idx in [2, 6]:
            for acc in df_ack_main.iloc[:, col_idx].astype(str).str.strip().unique():
                if acc.lower() not in ('', 'nan', 'none'):
                    stripped = acc.lstrip('0') or '0'
                    if stripped not in display_names or len(acc) > len(display_names[stripped]):
                        display_names[stripped] = acc

        for acc in sorted(all_accounts):
            display_acc = display_names.get(acc, acc)

            # Credited stats (Money coming IN)
            credited_rows = df_ack_main[
                (credited_series == acc) & credited_bank_is_countable
            ]
            credited_rows_for_total = credit_counting_df[
                credit_counting_series == acc
            ]

            total_credited = 0
            credited_transaction_ids = []

            for _, row in credited_rows.iterrows():
                # FILTER: Skip rows with null/empty credited transaction ID
                credited_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
                if not credited_trans_id or credited_trans_id.lower() in ('nan', 'none', '', '-', 'null'):
                    continue
                    
                trans_id = credited_trans_id

                # Collect credited transaction IDs
                if trans_id and trans_id.lower() not in ('', 'nan', 'none', 'unknown', '-'):
                    credited_transaction_ids.append(trans_id)

            for _, row in credited_rows_for_total.iterrows():
                credited_trans_id = (
                    str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
                )
                if not credited_trans_id or credited_trans_id.lower() in (
                    'nan', 'none', '', '-', 'null'
                ):
                    continue
                total_credited += clean_amount(row.iloc[11])

            # Join credited transaction IDs
            unique_credited_trans_ids = list(set(credited_transaction_ids))
            credited_trans_id_str = "; ".join(unique_credited_trans_ids) if unique_credited_trans_ids else "None"
            
            # SKIP accounts with no valid credited transactions (all were filtered out)
            if total_credited == 0 and len(credited_transaction_ids) == 0 and len(credited_rows) > 0:
                # This account only has rows with null credited transaction IDs, skip it
                continue

            # Debited stats (Money going OUT)
            debited_rows = df_ack_main[debited_series == acc]
            total_debited = sum(clean_amount(row.iloc[11]) for _, row in debited_rows.iterrows())

            # Bank Name Resolution
            account_bank = "N/A"
            for _, row in credited_rows.iterrows():
                b = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
                if b and b.lower() not in ('nan', 'none', ''):
                    account_bank = b
                    break

            if account_bank == "N/A":
                for _, row in debited_rows.iterrows():
                    b = str(row.iloc[14]).strip() if len(row) > 14 and pd.notna(row.iloc[14]) else ''
                    if b and b.lower() not in ('nan', 'none', ''):
                        account_bank = b
                        break

            # Updated Amount & Breakdown
            total_updated = 0
            breakdown_by_sheet = {}
            checked_trans_ids = set()

            for _, row in credited_rows.iterrows():
                # FILTER: Skip rows with null/empty credited transaction ID
                credited_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
                if not credited_trans_id or credited_trans_id.lower() in ('nan', 'none', '', '-', 'null'):
                    continue
                    
                trans_id = credited_trans_id
                if trans_id and trans_id not in checked_trans_ids:
                    checked_trans_ids.add(trans_id)
                    breakdown = get_transaction_breakdown(trans_id)
                    for item in breakdown:
                        sheet = item['sheet']
                        amount = item['amount']
                        total_updated += amount
                        breakdown_by_sheet[sheet] = breakdown_by_sheet.get(sheet, 0) + amount

            for _, row in debited_rows.iterrows():
                trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                if trans_id and trans_id not in checked_trans_ids:
                    checked_trans_ids.add(trans_id)
                    breakdown = get_transaction_breakdown(trans_id)
                    for item in breakdown:
                        sheet = item['sheet']
                        amount = item['amount']
                        total_updated += amount
                        breakdown_by_sheet[sheet] = breakdown_by_sheet.get(sheet, 0) + amount

            breakdown_text = "; ".join([f"{sheet}: ₹{amt:,.2f}" for sheet, amt in breakdown_by_sheet.items()]) if breakdown_by_sheet else "None"

            not_updated_amount = max(0, total_credited - total_debited - total_updated)

            if total_credited == 0 and len(debited_rows) > 0:
                summary_status = 'COMPLETE'
            elif total_debited == 0 and total_updated == 0:
                summary_status = 'PENDING'
            elif not_updated_amount > 0:
                summary_status = 'PARTIAL'
            else:
                summary_status = 'COMPLETE'

            all_account_summary.append({
                'acknowledgement_no': current_ack,
                'bank_name': account_bank,
                'account_number': display_acc,
                'credited_transaction_id': credited_trans_id_str,
                'total_credited': float(total_credited),
                'total_debited': float(total_debited),
                'updated_amount': float(total_updated),
                'not_updated_amount': float(not_updated_amount),
                'status': summary_status,
                'found_in_other_sheets': 'Yes' if total_updated > 0 else 'No',
                'breakdown_by_sheet': breakdown_text
            })

    return jsonify(all_account_summary)


@app.route('/api/all-transactions-by-id')
def get_all_transactions_by_id():
    """Get ALL transactions grouped by Transaction ID - TRANSACTION ID-WISE SUMMARY"""
    if df_main is None:
        return jsonify([])

    # Get unique acknowledgement numbers
    all_acks = df_main.iloc[:, 1].astype(str).str.strip().unique()
    all_acks = [a for a in all_acks if a and a.lower() not in ('nan', 'none', '')]

    if not all_acks:
        all_acks = ['UNKNOWN']

    all_transaction_summary = []

    for current_ack in sorted(all_acks):
        if current_ack != 'UNKNOWN':
            ack_mask = df_main.iloc[:, 1].astype(str).str.strip() == current_ack
            df_ack_main = df_main[ack_mask]
        else:
            df_ack_main = df_main

        if df_ack_main.empty:
            continue

        credit_counting_df, _ = strict_deduplicate_main_transactions(
            df_ack_main
        )
        credit_counting_tid_series = credit_counting_df.iloc[:, 9].apply(
            _identity_text
        )
        duplicate_details_by_transaction_id = {}
        for detail in credit_counting_df.attrs.get(
            'duplicate_transaction_details', []
        ):
            duplicate_details_by_transaction_id.setdefault(
                detail['credited_transaction_id'], []
            ).append(format_duplicate_credit_note(detail))

        # Collect all unique credited transaction IDs (column 9 only)
        all_trans_ids = set()
        credited_trans_ids = (
            credit_counting_df.iloc[:, 9].astype(str).str.strip()
        )
        # FILTER: Only include non-null credited transaction IDs
        all_trans_ids.update([tid for tid in credited_trans_ids if tid and tid.lower() not in ('nan', 'none', '', 'unknown', '-', 'null')])

        for trans_id in sorted(all_trans_ids):
            # Find all rows where this transaction ID appears as credited (column 9)
            credited_rows = credit_counting_df[
                credit_counting_df.iloc[:, 9].astype(str).str.strip()
                == trans_id
            ]

            # Collect credited account numbers (column 6) and bank names (column 4)
            credited_account_numbers = set()
            bank_names = set()
            debited_trans_ids = set()
            
            for _, row in credited_rows.iterrows():
                # Credited account (column 6)
                acc = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
                if acc and acc.lower() not in ('nan', 'none', ''):
                    credited_account_numbers.add(acc)
                
                # Bank name (column 4)
                bank = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
                if bank and bank.lower() not in ('nan', 'none', ''):
                    bank_names.add(bank)
                
                # Debited transaction ID (column 3)
                deb_tid = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                if deb_tid and deb_tid.lower() not in ('nan', 'none', '', 'unknown', '-'):
                    debited_trans_ids.add(deb_tid)

            account_numbers_str = "; ".join(sorted(credited_account_numbers)) if credited_account_numbers else "None"
            bank_name_str = "; ".join(sorted(bank_names)) if bank_names else "N/A"
            debited_trans_id_str = "; ".join(sorted(debited_trans_ids)) if debited_trans_ids else "None"

            # Duplicate identities change only the credited total.  Debit and
            # recovery calculations below continue to use df_ack_main.
            transaction_key = _identity_text(trans_id)
            credited_rows_for_total = credit_counting_df[
                credit_counting_tid_series == transaction_key
            ]
            total_credited = sum(
                clean_amount(row.iloc[11])
                for _, row in credited_rows_for_total.iterrows()
            )
            duplicate_details = duplicate_details_by_transaction_id.get(
                transaction_key, []
            )

            # Calculate outgoing/debited amount for the current credited transaction ID.
            # This should be based on child rows where the current transaction ID appears
            # as the source/debited transaction ID, not from the source ID on the same row.
            child_rows = df_ack_main[
                (df_ack_main.iloc[:, 3].astype(str).str.strip() == trans_id) &
                (df_ack_main.iloc[:, 9].astype(str).str.strip() != trans_id)
            ]
            total_debited = sum(clean_amount(row.iloc[11]) for _, row in child_rows.iterrows())

            # Updated Amount & Breakdown
            total_updated = 0
            breakdown_by_sheet = {}
            
            breakdown = get_transaction_breakdown(trans_id)
            for item in breakdown:
                sheet = item['sheet']
                amount = item['amount']
                total_updated += amount
                breakdown_by_sheet[sheet] = breakdown_by_sheet.get(sheet, 0) + amount

            breakdown_text = "; ".join([f"{sheet}: ₹{amt:,.2f}" for sheet, amt in breakdown_by_sheet.items()]) if breakdown_by_sheet else "None"

            not_updated_amount = max(0, total_credited - total_debited - total_updated)

            if total_credited == 0 and total_debited > 0:
                summary_status = 'COMPLETE'
            elif total_debited == 0 and total_updated == 0:
                summary_status = 'PENDING'
            elif not_updated_amount > 0:
                summary_status = 'PARTIAL'
            else:
                summary_status = 'COMPLETE'

            duplicate_info_str = " | ".join(duplicate_details) if duplicate_details else "None"

            all_transaction_summary.append({
                'acknowledgement_no': current_ack,
                'bank_name': bank_name_str,
                'account_number': account_numbers_str,
                'credited_transaction_id': trans_id,
                'debited_transaction_id': debited_trans_id_str,
                'total_credited': float(total_credited),
                'total_debited': float(total_debited),
                'updated_amount': float(total_updated),
                'not_updated_amount': float(not_updated_amount),
                'status': summary_status,
                'found_in_other_sheets': 'Yes' if total_updated > 0 else 'No',
                'breakdown_by_sheet': breakdown_text,
                'duplicate_entry_info': duplicate_info_str
            })

    return jsonify(all_transaction_summary)


@app.route('/api/transaction-details/<trans_id>')
def get_transaction_details(trans_id):
    breakdown = get_transaction_breakdown(trans_id)
    
    details = []
    for item in breakdown:
        details.append({
            'sheet': item['sheet'],
            'amount': item['amount'],
            'bank': '',
            'date': '',
            'remarks': ''
        })
    
    return jsonify(details)


@app.route('/api/get_transaction_details/<trans_id>')
def get_transaction_details_flow(trans_id):
    """Get detailed information for a specific transaction"""
    try:
        if df_main is None:
            return jsonify({'error': 'No data loaded'})
        
        account_no = request.args.get('account', '')
        
        # Find the transaction in df_main
        trans_row = None
        for idx, row in df_main.iterrows():
            row_trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
            if row_trans_id == trans_id:
                trans_row = row
                break
        
        if trans_row is None:
            return jsonify({'error': 'Transaction not found'})
        
        # Extract all details
        debited_account = str(trans_row.iloc[2]).strip() if pd.notna(trans_row.iloc[2]) else ''
        debited_bank = str(trans_row.iloc[4]).strip() if pd.notna(trans_row.iloc[4]) else ''
        credited_account = str(trans_row.iloc[6]).strip() if pd.notna(trans_row.iloc[6]) else ''
        credited_bank = str(trans_row.iloc[8]).strip() if pd.notna(trans_row.iloc[8]) else ''
        amount = clean_amount(trans_row.iloc[11])
        layer = int(trans_row.iloc[5]) if pd.notna(trans_row.iloc[5]) else 0
        date = str(trans_row.iloc[10]).strip() if pd.notna(trans_row.iloc[10]) else ''
        
        # Get breakdown from other sheets
        breakdown = get_transaction_breakdown(trans_id, account_no if account_no else debited_account)
        
        # Calculate status
        status_info = calculate_status(trans_id, amount, layer, debited_account)
        
        return jsonify({
            'trans_id': trans_id,
            'debited_account': debited_account,
            'debited_bank': debited_bank,
            'credited_account': credited_account,
            'credited_bank': credited_bank,
            'amount': amount,
            'layer': layer,
            'date': date,
            'status': status_info['status'],
            'updated_amount': status_info['updated_amount'],
            'pending_amount': status_info['pending_amount'],
            'children_total': status_info['children_total'],
            'breakdown': breakdown
        })
    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/download-report')
def download_report():
    try:
        if df_main is None:
            return jsonify({'success': False, 'message': 'No data loaded'})
        
        all_data = build_hierarchical_data()
        
        if not all_data:
            return jsonify({'success': False, 'message': 'No data to export'})
        
        df = pd.DataFrame(all_data)
        
        # VALIDATION NOTE: Account-based merging
        original_count = len(df_main)
        output_count = len(df)
        
        validation_warnings = []
        if output_count != original_count:
            validation_warnings.append(f"ℹ️ INFO: Original file has {original_count} transactions, output has {output_count} merged account groups")
            validation_warnings.append(f"ℹ️ This is expected: Transactions with same account pair and status are merged")
            validation_warnings.append(f"ℹ️ Merged: {abs(original_count - output_count)} transactions into groups")
        
        # Check if all accounts are present
        original_accounts = set(df_main.iloc[:, 6].astype(str).str.strip())
        output_accounts = set(df['Credited Account'].astype(str).str.strip())
        missing_accounts = original_accounts - output_accounts
        
        if missing_accounts:
            validation_warnings.append(f"⚠️ WARNING: {len(missing_accounts)} accounts missing from output!")
            validation_warnings.append(f"⚠️ Missing Accounts: {', '.join(list(missing_accounts)[:10])}...")
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Hierarchical Report', index=False)
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            workbook = writer.book
            worksheet = writer.sheets['Hierarchical Report']
            
            # Enable filtering on all columns
            worksheet.auto_filter.ref = worksheet.dimensions
            
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            column_widths = {
                'A': 8,   # Level
                'B': 8,   # Layer
                'C': 20,  # Debited Account (Filter)
                'D': 50,  # Hierarchy Path
                'E': 12,  # Transaction Count
                'F': 20,  # S.No
                'G': 18,  # Debited Account
                'H': 25,  # Debited Transaction ID
                'I': 20,  # Debited Bank
                'J': 20,  # Acknowledgement No
                'K': 18,  # Credited Account
                'L': 25,  # Credited Transaction ID
                'M': 20,  # Credited Bank
                'N': 15,  # IFSC Code
                'O': 20,  # Transaction Date
                'P': 15,  # Transaction Amount
                'Q': 15,  # Disputed Amount
                'R': 18,  # Layerwise Amount Difference
                'S': 15,  # Updated Amount
                'T': 15,  # Pending Amount
                'U': 12,  # Status
                'V': 12,  # Has Children
                'W': 18,  # Found in Other Sheets
                'X': 40,  # Breakdown by Sheet
                'Y': 15,  # Reference No
                'Z': 30,  # Remarks
                'AA': 20,  # Action Taken By
                'AB': 20  # Date of Action
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            level_colors = {
                0: 'E3F2FD', 1: 'FFF3E0', 2: 'F3E5F5', 3: 'E8F5E9',
                4: 'FFF9C4', 5: 'FCE4EC', 6: 'E0F2F1'
            }
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
                level = df.iloc[row_idx-2]['Level']
                fill_color = level_colors.get(level, 'FFFFFF')
                row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                # Get status for this row
                status = df.iloc[row_idx-2]['Status']
                
                for cell in row:
                    cell.border = thin_border
                    cell.fill = row_fill
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # Color code Status column (column U = 21)
                    if cell.column == 21:  # Status column
                        if status == 'PENDING':
                            cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')  # Light Red
                            cell.font = Font(color='C62828', bold=True)  # Dark Red text
                        elif status == 'COMPLETE':
                            cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # Light Green
                            cell.font = Font(color='2E7D32', bold=True)  # Dark Green text
                        elif status == 'PARTIAL':
                            cell.fill = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')  # Light Yellow
                            cell.font = Font(color='F57C00', bold=True)  # Dark Orange text
                        elif status == 'TRANSACTION CONTINUE':
                            cell.fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')  # Light Blue
                            cell.font = Font(color='1565C0', bold=True)  # Dark Blue text
                    
                    if cell.column in [16, 17, 18, 19, 20]:  # Amount columns (P, Q, R, S, T)
                        cell.number_format = '₹#,##0.00'
            
            # Summary - Sheet-wise disputed amounts with dynamic formulas
            summary_data = []
            
            # Money Transfer sheet summary (will use formula)
            money_transfer_total = df['Disputed Amount'].sum()
            summary_data.append({
                'Sheet Name': 'Money Transfer to',
                'Total Disputed Amount': money_transfer_total
            })
            
            # Other sheets summary
            for sheet_name, sheet_info in df_other_sheets.items():
                df_sheet = sheet_info['data']
                amount_col = sheet_info['amount_col']
                
                # Calculate total amount without removing duplicates
                if len(df_sheet.columns) > amount_col:
                    total_amount = df_sheet.iloc[:, amount_col].apply(clean_amount).sum()
                else:
                    total_amount = 0
                
                if total_amount > 0:
                    summary_data.append({
                        'Sheet Name': sheet_name,
                        'Total Disputed Amount': total_amount
                    })
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False, startrow=0)
            
            summary_ws = writer.sheets['Summary']
            
            # Money Transfer total formula - SUM of all disputed amounts
            summary_ws['B2'] = f"=SUM('Hierarchical Report'!Q:Q)"
            
            # Header formatting
            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            summary_ws.column_dimensions['A'].width = 40
            summary_ws.column_dimensions['B'].width = 25
            
            # Format amounts as currency
            for row_idx in range(2, len(summary_data) + 2):
                cell = summary_ws.cell(row=row_idx, column=2)
                cell.number_format = '₹#,##0.00'
            
            # Grand total with formula
            grand_total_row = len(summary_data) + 2
            summary_ws.cell(row=grand_total_row, column=1, value='GRAND TOTAL')
            summary_ws.cell(row=grand_total_row, column=1).font = Font(bold=True)
            summary_ws.cell(row=grand_total_row, column=2).value = f"=SUM(B2:B{grand_total_row-1})"
            summary_ws.cell(row=grand_total_row, column=2).number_format = '₹#,##0.00'
            summary_ws.cell(row=grand_total_row, column=2).font = Font(bold=True)
            
            # Add charts
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.chart.label import DataLabelList
            
            # Bar Chart
            bar_chart = BarChart()
            bar_chart.title = "Sheet-wise Disputed Amount Distribution"
            bar_chart.style = 10
            bar_chart.y_axis.title = 'Amount (₹)'
            bar_chart.x_axis.title = 'Sheet Name'
            
            data_rows = len(summary_data)
            data = Reference(summary_ws, min_col=2, min_row=1, max_row=data_rows + 1)
            cats = Reference(summary_ws, min_col=1, min_row=2, max_row=data_rows + 1)
            
            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(cats)
            bar_chart.height = 15
            bar_chart.width = 25
            
            summary_ws.add_chart(bar_chart, "D2")
            
            # Pie Chart
            pie_chart = PieChart()
            pie_chart.title = "Top Sheets by Disputed Amount"
            pie_chart.style = 10
            
            data = Reference(summary_ws, min_col=2, min_row=2, max_row=min(7, data_rows + 1))
            cats = Reference(summary_ws, min_col=1, min_row=2, max_row=min(7, data_rows + 1))
            
            pie_chart.add_data(data)
            pie_chart.set_categories(cats)
            pie_chart.height = 12
            pie_chart.width = 15
            
            pie_chart.dataLabels = DataLabelList()
            pie_chart.dataLabels.showPercent = True
            
            summary_ws.add_chart(pie_chart, "D32")
            
            # Status Summary with formulas
            status_summary_row = grand_total_row + 3
            summary_ws.cell(row=status_summary_row, column=1, value="Status Summary (Live from Hierarchical Report)")
            summary_ws.cell(row=status_summary_row, column=1).font = Font(bold=True, size=14, color='1F4E78')
            
            status_summary_row += 2
            summary_ws.cell(row=status_summary_row, column=1, value="Status")
            summary_ws.cell(row=status_summary_row, column=2, value="Count")
            summary_ws.cell(row=status_summary_row, column=3, value="Total Amount")
            
            for cell in summary_ws[status_summary_row]:
                cell.fill = header_fill
                cell.font = header_font
            
            status_summary_row += 1
            
            # Status formulas - COUNTIF and SUMIF from Hierarchical Report
            for status in ['PENDING', 'PARTIAL', 'COMPLETE', 'TRANSACTION CONTINUE']:
                summary_ws.cell(row=status_summary_row, column=1, value=status)
                # Count formula
                summary_ws.cell(row=status_summary_row, column=2).value = f"=COUNTIF('Hierarchical Report'!U:U,\"{status}\")"
                # Sum formula
                summary_ws.cell(row=status_summary_row, column=3).value = f"=SUMIF('Hierarchical Report'!U:U,\"{status}\",'Hierarchical Report'!Q:Q)"
                summary_ws.cell(row=status_summary_row, column=3).number_format = '₹#,##0.00'
                
                # Color code status
                if status == 'PENDING':
                    summary_ws.cell(row=status_summary_row, column=1).fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                elif status == 'COMPLETE':
                    summary_ws.cell(row=status_summary_row, column=1).fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                elif status == 'PARTIAL':
                    summary_ws.cell(row=status_summary_row, column=1).fill = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')
                elif status == 'TRANSACTION CONTINUE':
                    summary_ws.cell(row=status_summary_row, column=1).fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
                
                status_summary_row += 1
            
            summary_ws.column_dimensions['C'].width = 25
            
            # Bank-wise Summary by Status
            bank_summary_row = status_summary_row + 3
            summary_ws.cell(row=bank_summary_row, column=1, value="Bank-wise Summary by Status")
            summary_ws.cell(row=bank_summary_row, column=1).font = Font(bold=True, size=14, color='1F4E78')
            
            bank_summary_row += 2
            
            # PARTIAL Status - Bank-wise updated Amount
            summary_ws.cell(row=bank_summary_row, column=1, value="PARTIAL Status - updated Amount by Bank")
            summary_ws.cell(row=bank_summary_row, column=1).font = Font(bold=True, size=12)
            summary_ws.cell(row=bank_summary_row, column=1).fill = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')
            bank_summary_row += 1
            
            summary_ws.cell(row=bank_summary_row, column=1, value="Bank Name")
            summary_ws.cell(row=bank_summary_row, column=2, value="Count")
            summary_ws.cell(row=bank_summary_row, column=3, value="Total updated Amount")
            for cell in summary_ws[bank_summary_row]:
                cell.fill = header_fill
                cell.font = header_font
            bank_summary_row += 1
            
            # Get unique banks for PARTIAL status with formulas (Column I = Debited Bank, Column U = Status, Column S = updated Amount)
            partial_banks = df[df['Status'] == 'PARTIAL']['Debited Bank'].unique()
            for bank in sorted(partial_banks):
                if bank and str(bank).strip():
                    summary_ws.cell(row=bank_summary_row, column=1, value=str(bank))
                    # Count formula
                    summary_ws.cell(row=bank_summary_row, column=2).value = f"=COUNTIFS('Hierarchical Report'!U:U,\"PARTIAL\",'Hierarchical Report'!I:I,\"{bank}\")"
                    # Sum updated Amount formula
                    summary_ws.cell(row=bank_summary_row, column=3).value = f"=SUMIFS('Hierarchical Report'!S:S,'Hierarchical Report'!U:U,\"PARTIAL\",'Hierarchical Report'!I:I,\"{bank}\")"
                    summary_ws.cell(row=bank_summary_row, column=3).number_format = '₹#,##0.00'
                    bank_summary_row += 1
            
            bank_summary_row += 1
            
            # COMPLETE Status - Bank-wise updated Amount
            summary_ws.cell(row=bank_summary_row, column=1, value="COMPLETE Status - updated Amount by Bank")
            summary_ws.cell(row=bank_summary_row, column=1).font = Font(bold=True, size=12)
            summary_ws.cell(row=bank_summary_row, column=1).fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
            bank_summary_row += 1
            
            summary_ws.cell(row=bank_summary_row, column=1, value="Bank Name")
            summary_ws.cell(row=bank_summary_row, column=2, value="Count")
            summary_ws.cell(row=bank_summary_row, column=3, value="Total updated Amount")
            for cell in summary_ws[bank_summary_row]:
                cell.fill = header_fill
                cell.font = header_font
            bank_summary_row += 1
            
            # Get unique banks for COMPLETE status with formulas
            complete_banks = df[df['Status'] == 'COMPLETE']['Debited Bank'].unique()
            for bank in sorted(complete_banks):
                if bank and str(bank).strip():
                    summary_ws.cell(row=bank_summary_row, column=1, value=str(bank))
                    # Count formula
                    summary_ws.cell(row=bank_summary_row, column=2).value = f"=COUNTIFS('Hierarchical Report'!U:U,\"COMPLETE\",'Hierarchical Report'!I:I,\"{bank}\")"
                    # Sum updated Amount formula
                    summary_ws.cell(row=bank_summary_row, column=3).value = f"=SUMIFS('Hierarchical Report'!S:S,'Hierarchical Report'!U:U,\"COMPLETE\",'Hierarchical Report'!I:I,\"{bank}\")"
                    summary_ws.cell(row=bank_summary_row, column=3).number_format = '₹#,##0.00'
                    bank_summary_row += 1
            
            bank_summary_row += 1
            
            # PENDING Status - Bank-wise Pending Amount
            summary_ws.cell(row=bank_summary_row, column=1, value="PENDING Status - Pending Amount by Bank")
            summary_ws.cell(row=bank_summary_row, column=1).font = Font(bold=True, size=12)
            summary_ws.cell(row=bank_summary_row, column=1).fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
            bank_summary_row += 1
            
            summary_ws.cell(row=bank_summary_row, column=1, value="Bank Name")
            summary_ws.cell(row=bank_summary_row, column=2, value="Count")
            summary_ws.cell(row=bank_summary_row, column=3, value="Total Pending Amount")
            for cell in summary_ws[bank_summary_row]:
                cell.fill = header_fill
                cell.font = header_font
            bank_summary_row += 1
            
            # Get unique banks for PENDING status with formulas (Column T = Pending Amount)
            pending_banks = df[df['Status'] == 'PENDING']['Debited Bank'].unique()
            for bank in sorted(pending_banks):
                if bank and str(bank).strip():
                    summary_ws.cell(row=bank_summary_row, column=1, value=str(bank))
                    # Count formula
                    summary_ws.cell(row=bank_summary_row, column=2).value = f"=COUNTIFS('Hierarchical Report'!U:U,\"PENDING\",'Hierarchical Report'!I:I,\"{bank}\")"
                    # Sum Pending Amount formula
                    summary_ws.cell(row=bank_summary_row, column=3).value = f"=SUMIFS('Hierarchical Report'!T:T,'Hierarchical Report'!U:U,\"PENDING\",'Hierarchical Report'!I:I,\"{bank}\")"
                    summary_ws.cell(row=bank_summary_row, column=3).number_format = '₹#,##0.00'
                    bank_summary_row += 1
            
            summary_ws.column_dimensions['A'].width = 40
            summary_ws.column_dimensions['B'].width = 15
            summary_ws.column_dimensions['C'].width = 25
            
            # Add validation info if any
            if validation_warnings:
                warning_row = bank_summary_row + 3
                summary_ws.cell(row=warning_row, column=1, value="ℹ️ ACCOUNT-BASED MERGING INFO")
                summary_ws.cell(row=warning_row, column=1).font = Font(bold=True, size=14, color='1F4E78')
                warning_row += 2
                
                for warning in validation_warnings:
                    summary_ws.cell(row=warning_row, column=1, value=warning)
                    summary_ws.cell(row=warning_row, column=1).font = Font(color='1F4E78', bold=True)
                    summary_ws.cell(row=warning_row, column=1).fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                    warning_row += 1
        
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Account_Based_Report_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/download-leaf-nodes')
def download_leaf_nodes():
    """Download report with only leaf nodes (transactions with no children)"""
    try:
        if df_main is None:
            return jsonify({'success': False, 'message': 'No data loaded'})
        
        # Build full hierarchy
        all_data = build_hierarchical_data()
        
        if not all_data:
            return jsonify({'success': False, 'message': 'No data to export'})
        
        # Filter only leaf nodes (Has Children = 'No')
        df_full = pd.DataFrame(all_data)
        df = df_full[df_full['Has Children'] == 'No'].copy()
        
        if len(df) == 0:
            return jsonify({'success': False, 'message': 'No leaf node transactions found'})
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Leaf Nodes Report', index=False)
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            workbook = writer.book
            worksheet = writer.sheets['Leaf Nodes Report']
            
            # Enable filtering
            worksheet.auto_filter.ref = worksheet.dimensions
            
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Same column widths as main report
            column_widths = {
                'A': 8, 'B': 8, 'C': 20, 'D': 50, 'E': 8, 'F': 18, 'G': 18, 'H': 20,
                'I': 20, 'J': 18, 'K': 20, 'L': 20, 'M': 15, 'N': 20, 'O': 15,
                'P': 15, 'Q': 18, 'R': 15, 'S': 15, 'T': 12, 'U': 12, 'V': 18,
                'W': 40, 'X': 15, 'Y': 30, 'Z': 20, 'AA': 20
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
                status = df.iloc[row_idx-2]['Status']
                
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # Color code Status column
                    if cell.column == 20:  # Status column (now column T = 20)
                        if status == 'PENDING':
                            cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                            cell.font = Font(color='C62828', bold=True)
                        elif status == 'COMPLETE':
                            cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                            cell.font = Font(color='2E7D32', bold=True)
                        elif status == 'PARTIAL':
                            cell.fill = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')
                            cell.font = Font(color='F57C00', bold=True)
                        elif status == 'TRANSACTION CONTINUE':
                            cell.fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
                            cell.font = Font(color='1565C0', bold=True)
                    
                    if cell.column in [15, 16, 17, 18, 19]:  # Amount columns
                        cell.number_format = '₹#,##0.00'
            
            # Summary for leaf nodes
            summary_data = []
            
            # Count by status
            for status in ['PENDING', 'PARTIAL', 'COMPLETE']:
                status_data = df[df['Status'] == status]
                if len(status_data) > 0:
                    summary_data.append({
                        'Category': f'{status} Transactions',
                        'Count': len(status_data),
                        'Total Disputed': status_data['Disputed Amount'].sum(),
                        'Total Pending': status_data['Pending Amount'].sum()
                    })
            
            summary_data.append({})
            summary_data.append({
                'Category': 'TOTAL LEAF NODES',
                'Count': len(df),
                'Total Disputed': df['Disputed Amount'].sum(),
                'Total Pending': df['Pending Amount'].sum()
            })
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            summary_ws = writer.sheets['Summary']
            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 15
            summary_ws.column_dimensions['C'].width = 20
            summary_ws.column_dimensions['D'].width = 20
        
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Leaf_Nodes_Report_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/download-partial')
def download_partial():
    """Download report with only PARTIAL status transactions"""
    try:
        if df_main is None:
            return jsonify({'success': False, 'message': 'No data loaded'})
        
        # Build full hierarchy
        all_data = build_hierarchical_data()
        
        if not all_data:
            return jsonify({'success': False, 'message': 'No data to export'})
        
        # Filter only PARTIAL status
        df_full = pd.DataFrame(all_data)
        df = df_full[df_full['Status'] == 'PARTIAL'].copy()
        
        if len(df) == 0:
            return jsonify({'success': False, 'message': 'No PARTIAL status transactions found'})
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Partial Status Report', index=False)
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            workbook = writer.book
            worksheet = writer.sheets['Partial Status Report']
            
            # Enable filtering
            worksheet.auto_filter.ref = worksheet.dimensions
            
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Same column widths as main report
            column_widths = {
                'A': 8, 'B': 8, 'C': 20, 'D': 50, 'E': 8, 'F': 18, 'G': 18, 'H': 20,
                'I': 20, 'J': 18, 'K': 20, 'L': 20, 'M': 15, 'N': 20, 'O': 15,
                'P': 15, 'Q': 18, 'R': 15, 'S': 15, 'T': 12, 'U': 12, 'V': 18,
                'W': 40, 'X': 15, 'Y': 30, 'Z': 20, 'AA': 20
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
                status = df.iloc[row_idx-2]['Status']
                
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # Color code Status column (PARTIAL = Yellow)
                    if cell.column == 20:  # Status column
                        cell.fill = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')
                        cell.font = Font(color='F57C00', bold=True)
                    
                    if cell.column in [15, 16, 17, 18, 19]:  # Amount columns
                        cell.number_format = '₹#,##0.00'
            
            # Summary for partial transactions
            summary_data = []
            summary_data.append({
                'Category': 'PARTIAL Transactions',
                'Count': len(df),
                'Total Disputed': df['Disputed Amount'].sum(),
                'Total updated': df['Updated Amount'].sum(),
                'Total Pending': df['Pending Amount'].sum()
            })
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            summary_ws = writer.sheets['Summary']
            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 15
            summary_ws.column_dimensions['C'].width = 20
            summary_ws.column_dimensions['D'].width = 20
            summary_ws.column_dimensions['E'].width = 20
        
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Partial_Status_Report_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/download-account-summary')
def download_account_summary():
    """Download report grouped by single account number"""
    try:
        if df_main is None:
            return jsonify({'success': False, 'message': 'No data loaded'})
            
        # Ensure we have Unique Acknowledgement numbers from the first column
        # Fallback to 'UNKNOWN' if something is wrong with the file mapping
        if 'Acknowledge' in df_main.columns[1] or True: # Col 1 is Acknowledgement No.
            all_acks = df_main.iloc[:, 1].astype(str).str.strip().unique()
            all_acks = [a for a in all_acks if a and a.lower() not in ('nan', 'none', '')]
        
        if not all_acks:
            all_acks = ['UNKNOWN']

        all_account_summary = []
        all_bank_data_list = []
        all_partial_bank_data_list = []

        # Create helper series with stripped leading zeros for faster, accurate filtering
        # Improved account stripping to handle floats and varying formats
        def strip_account(x):
            if pd.isna(x): return ''
            s = str(x).strip()
            if s.lower() in ('nan', 'none', ''): return ''
            # Handle float representations (e.g., '123.0')
            if s.endswith('.0'):
                s = s[:-2]
            return s.lstrip('0') or '0'
        
        debited_series_full = df_main.iloc[:, 2].apply(strip_account)
        credited_series_full = df_main.iloc[:, 6].apply(strip_account)

        # Duplicate identities affect credited aggregation only.  Build notes
        # for the credited account; debit and other-sheet results are untouched.
        duplicate_notes_by_credited_account = {}
        for detail in last_duplicate_transaction_details:
            duplicate_rows = int(detail.get('duplicate_rows_removed', 0) or 0)
            if duplicate_rows <= 0:
                continue
            note = format_duplicate_credit_note(detail)
            acknowledgement_key = _identity_text(
                detail.get('acknowledgement_no')
            )
            credited_key = (
                acknowledgement_key,
                detail.get('credited_account_last_four', ''),
            )
            duplicate_notes_by_credited_account.setdefault(
                credited_key, []
            ).append(note)

        for current_ack in sorted(all_acks):
            if current_ack != 'UNKNOWN':
                ack_mask = df_main.iloc[:, 1].astype(str).str.strip() == current_ack
                df_ack_main = df_main[ack_mask]
                debited_series = debited_series_full[ack_mask]
                credited_series = credited_series_full[ack_mask]
            else:
                df_ack_main = df_main
                debited_series = debited_series_full
                credited_series = credited_series_full
                
            if df_ack_main.empty: continue

            credit_counting_df, _ = strict_deduplicate_main_transactions(
                df_ack_main
            )
            credit_counting_series = credit_counting_df.iloc[:, 6].apply(
                strip_account
            )

            credited_bank_is_countable = ~df_ack_main.iloc[:, 4].map(
                _is_money_transfer_to_others_bank
            )
            
            all_accounts = set(debited_series[debited_series != ''].unique()) | set(
                credited_series[
                    (credited_series != '') & credited_bank_is_countable
                ].unique()
            )
            
            # Store the longest string variant for each account to use as the display name
            display_names = {}
            for col_idx in [2, 6]:
                for acc in df_ack_main.iloc[:, col_idx].astype(str).str.strip().unique():
                    if acc.lower() not in ('', 'nan', 'none'):
                        stripped = acc.lstrip('0') or '0'
                        if stripped not in display_names or len(acc) > len(display_names[stripped]):
                            display_names[stripped] = acc
                            
            bank_data = {}         # Store info for all banks aggregation for this ACK
            partial_bank_data = {} # Store info strictly for PARTIAL accounts aggregation for this ACK
            
            # OPTIMIZATION: Pre-calculate breakdown data for all transaction IDs in this acknowledgment
            breakdown_cache = {}
            all_trans_ids_in_ack = set()
            for _, row in df_ack_main.iterrows():
                # Collect both debited and credited transaction IDs
                deb_tid = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                cre_tid = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
                if deb_tid: all_trans_ids_in_ack.add(deb_tid)
                if cre_tid: all_trans_ids_in_ack.add(cre_tid)
            
            # Pre-fetch all breakdowns at once
            for trans_id in all_trans_ids_in_ack:
                if trans_id:
                    breakdown_cache[trans_id] = get_transaction_breakdown(trans_id)
            
            for acc in sorted(all_accounts):
                display_acc = display_names.get(acc, acc)
                
                # Credited stats (Money coming IN to this account)
                credited_rows = df_ack_main[
                    (credited_series == acc) & credited_bank_is_countable
                ]
                credited_rows_for_total = credit_counting_df[
                    credit_counting_series == acc
                ]
                
                total_credited = 0
                credited_transaction_ids = []
                
                for _, row in credited_rows.iterrows():
                    trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
                    
                    # Collect credited transaction IDs
                    if trans_id and trans_id.lower() not in ('', 'nan', 'none', 'unknown', '-'):
                        credited_transaction_ids.append(trans_id)
                total_credited = sum(
                    clean_amount(row.iloc[11])
                    for _, row in credited_rows_for_total.iterrows()
                )
                
                # Remove duplicates and join credited transaction IDs
                unique_credited_trans_ids = list(set(credited_transaction_ids))
                credited_trans_id_str = "; ".join(unique_credited_trans_ids) if unique_credited_trans_ids else "None"
                
                # Debited stats (Money going OUT from this account) - always use ALL rows
                debited_rows = df_ack_main[debited_series == acc]
                total_debited = sum(clean_amount(row.iloc[11]) for _, row in debited_rows.iterrows())
                
                # Bank Name Resolution
                account_bank = "N/A"
                for _, row in credited_rows.iterrows():
                    b = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
                    if b and b.lower() not in ('nan', 'none', ''):
                        account_bank = b
                        break
                
                if account_bank == "N/A":
                    for _, row in debited_rows.iterrows():
                        b = str(row.iloc[14]).strip() if len(row) > 14 and pd.notna(row.iloc[14]) else ''
                        if b and b.lower() not in ('nan', 'none', ''):
                            account_bank = b
                            break
                
                # Updated Amount & Breakdown - USE CACHE
                total_updated = 0
                breakdown_by_sheet = {}
                checked_trans_ids = set()
                
                for _, row in credited_rows.iterrows():
                    trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
                    if trans_id and trans_id not in checked_trans_ids:
                        checked_trans_ids.add(trans_id)
                        breakdown = breakdown_cache.get(trans_id, [])
                        for item in breakdown:
                            sheet = item['sheet']
                            amount = item['amount']
                            total_updated += amount
                            breakdown_by_sheet[sheet] = breakdown_by_sheet.get(sheet, 0) + amount
                
                for _, row in debited_rows.iterrows():
                    trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                    if trans_id and trans_id not in checked_trans_ids:
                        checked_trans_ids.add(trans_id)
                        breakdown = breakdown_cache.get(trans_id, [])
                        for item in breakdown:
                            sheet = item['sheet']
                            amount = item['amount']
                            total_updated += amount
                            breakdown_by_sheet[sheet] = breakdown_by_sheet.get(sheet, 0) + amount
                
                breakdown_text = "; ".join([f"{sheet}: ₹{amt:,.2f}" for sheet, amt in breakdown_by_sheet.items()]) if breakdown_by_sheet else "None"
                
                not_updated_amount = max(0, total_credited - total_debited - total_updated)
                
                if total_credited == 0 and len(debited_rows) > 0:
                    summary_status = 'COMPLETED'
                elif total_debited == 0 and total_updated == 0:
                    summary_status = 'PENDING'
                elif not_updated_amount > 0:
                    summary_status = 'PARTIAL'
                else:
                    summary_status = 'COMPLETED'
                
                acknowledgement_key = _identity_text(current_ack)
                account_last_four = _account_last_four(display_acc)
                duplicate_notes = []
                if not credited_rows.empty:
                    duplicate_notes.extend(
                        duplicate_notes_by_credited_account.get(
                            (acknowledgement_key, account_last_four), []
                        )
                    )
                duplicate_notes = list(dict.fromkeys(duplicate_notes))
                duplicate_info_str = (
                    " | ".join(duplicate_notes) if duplicate_notes else "None"
                )

                all_account_summary.append({
                    'Acknowledgement No': current_ack,
                    'Bank Name': account_bank,
                    'Account Number': display_acc,
                    'Credited Transaction ID': credited_trans_id_str,
                    'Total Credited Amount': total_credited,
                    'Total Debited Amount': total_debited,
                    'Updated Amount (Recovery)': total_updated,
                    'Not Updated Amount': not_updated_amount,
                    'Status': summary_status,
                    'Found in Other Sheets': 'Yes' if total_updated > 0 else 'No',
                    'Breakdown by Sheet': breakdown_text,
                    'Duplicate Entry Info': duplicate_info_str
                })

                if account_bank != "N/A":
                    # 1. Total Bank Wise
                    if account_bank not in bank_data:
                        bank_data[account_bank] = {'Credited': 0, 'Debited': 0, 'Updated': 0, 'Breakdown': {}, 'Duplicate Info': []}
                    bd = bank_data[account_bank]
                    bd['Credited'] += total_credited
                    bd['Debited'] += total_debited
                    bd['Updated'] += total_updated
                    for s, a in breakdown_by_sheet.items():
                        bd['Breakdown'][s] = bd['Breakdown'].get(s, 0) + a
                    if duplicate_info_str != "None":
                        bd['Duplicate Info'].append(duplicate_info_str)

                    # 2. Partial Bank Wise
                    if summary_status == 'PARTIAL':
                        if account_bank not in partial_bank_data:
                            partial_bank_data[account_bank] = {'Credited': 0, 'Debited': 0, 'Updated': 0, 'Breakdown': {}, 'Duplicate Info': []}
                        pbd = partial_bank_data[account_bank]
                        pbd['Credited'] += total_credited
                        pbd['Debited'] += total_debited
                        pbd['Updated'] += total_updated
                        for s, a in breakdown_by_sheet.items():
                            pbd['Breakdown'][s] = pbd['Breakdown'].get(s, 0) + a
                        if duplicate_info_str != "None":
                            pbd['Duplicate Info'].append(duplicate_info_str)
                            
            def build_summary_list(data_dict, ack):
                summary_list = []
                for b_name, b_info in data_dict.items():
                    b_not_updated = max(0, b_info['Credited'] - b_info['Debited'] - b_info['Updated'])
                    if b_info['Debited'] == 0 and b_info['Updated'] == 0:
                        b_status = 'PENDING'
                    elif b_not_updated > 0:
                        b_status = 'PARTIAL'
                    else:
                        b_status = 'COMPLETED'
                    
                    b_breakdown_text = "; ".join([f"{s}: ₹{a:,.2f}" for s, a in b_info['Breakdown'].items()]) if b_info['Breakdown'] else "None"
                    
                    b_duplicate_info = " | ".join(b_info['Duplicate Info']) if b_info.get('Duplicate Info') else "None"
                    
                    summary_list.append({
                        'Acknowledgement No': ack,
                        'Bank Name': b_name,
                        'Total Credited Amount': b_info['Credited'],
                        'Total Debited Amount': b_info['Debited'],
                        'Updated Amount (Recovery)': b_info['Updated'],
                        'Not Updated Amount': b_not_updated,
                        'Status': b_status,
                        'Found in Other Sheets': 'Yes' if b_info['Updated'] > 0 else 'No',
                        'Breakdown by Sheet': b_breakdown_text,
                        'Duplicate Entry Info': b_duplicate_info
                    })
                return summary_list

            all_bank_data_list.extend(build_summary_list(bank_data, current_ack))
            all_partial_bank_data_list.extend(build_summary_list(partial_bank_data, current_ack))

        df_acc = pd.DataFrame(all_account_summary)
        df_bank = pd.DataFrame(all_bank_data_list)
        df_partial_bank = pd.DataFrame(all_partial_bank_data_list)
        df_money_transfer_to_others = _money_transfer_to_others_frame()
        
        if len(df_acc) == 0:
            return jsonify({'success': False, 'message': 'No account data found'})
            
        # Sort all by Not Updated descending
        for df in [df_acc, df_bank, df_partial_bank]:
            if len(df) > 0:
                df.sort_values(by='Not Updated Amount', ascending=False, inplace=True)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_acc.to_excel(writer, sheet_name='Account Wise Summary', index=False)
            df_bank.to_excel(writer, sheet_name='Bank Wise Summary', index=False)
            df_partial_bank.to_excel(writer, sheet_name='Partial Bank Wise Summary', index=False)
            df_money_transfer_to_others.to_excel(
                writer,
                sheet_name=MONEY_TRANSFER_TO_OTHERS_SHEET,
                index=False,
            )
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # --- STYLING LOGIC ---
            header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            # More vibrant colors for status
            pending_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')  # Bright red
            partial_fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')  # Bright orange
            completed_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # Bright green
            
            stripe_fill = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
            )

            for sheet_name in ['Account Wise Summary', 'Bank Wise Summary', 'Partial Bank Wise Summary']:
                worksheet = writer.sheets[sheet_name]
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Header Styling
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
                
                # Column Widths & Currency/Status Cols Mapping
                if sheet_name == 'Account Wise Summary':
                    col_widths = {'A': 35, 'B': 30, 'C': 25, 'D': 22, 'E': 22, 'F': 22, 'G': 22, 'H': 22, 'I': 15, 'J': 18, 'K': 60, 'L': 60}
                    amt_cols = [5, 6, 7, 8]; status_col = 9; yes_no_col = 10  # Updated for new column
                else: # Bank Wise summaries
                    col_widths = {'A': 35, 'B': 22, 'C': 22, 'D': 22, 'E': 22, 'F': 15, 'G': 18, 'H': 60, 'I': 60}
                    amt_cols = [2, 3, 4, 5]; status_col = 6; yes_no_col = 7
                    
                for col, width in col_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Data Styling
                current_df = df_acc if sheet_name == 'Account Wise Summary' else (df_bank if sheet_name == 'Bank Wise Summary' else df_partial_bank)
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(current_df)+1), start=2):
                    is_stripe = row_idx % 2 == 0
                    status_value = str(worksheet.cell(row=row_idx, column=status_col).value).upper()
                    
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                        if is_stripe: cell.fill = stripe_fill
                        
                        if cell.column == status_col:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            status_text = str(cell.value).upper() if cell.value else ""
                            if "PENDING" in status_text:
                                cell.fill = pending_fill
                                cell.font = Font(color='C62828', bold=True, size=11)  # Dark red text
                            elif "PARTIAL" in status_text:
                                cell.fill = partial_fill
                                cell.font = Font(color='E65100', bold=True, size=11)  # Dark orange text
                            elif "COMPLETED" in status_text or "COMPLETE" in status_text:
                                cell.fill = completed_fill
                                cell.font = Font(color='2E7D32', bold=True, size=11)  # Dark green text
                        
                        if cell.column in amt_cols:
                            cell.number_format = '₹#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        
                        if cell.column == yes_no_col:
                            cell.alignment = Alignment(horizontal='center', vertical='center')

            raw_worksheet = writer.sheets[MONEY_TRANSFER_TO_OTHERS_SHEET]
            raw_worksheet.freeze_panes = 'A2'
            raw_worksheet.auto_filter.ref = raw_worksheet.dimensions
            for cell in raw_worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True,
                )
                cell.border = thin_border
            for column_cells in raw_worksheet.columns:
                column_letter = column_cells[0].column_letter
                raw_worksheet.column_dimensions[column_letter].width = min(
                    50,
                    max(
                        12,
                        max(
                            len(str(cell.value or ''))
                            for cell in column_cells
                        ) + 2,
                    ),
                )
                        
        output.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Account_Summary_Report_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@app.route('/download-transaction-id-summary')
def download_transaction_id_summary():
    """Download report grouped by Transaction ID"""
    try:
        if df_main is None:
            return jsonify({'success': False, 'message': 'No data loaded'})
            
        # Get unique acknowledgement numbers
        all_acks = df_main.iloc[:, 1].astype(str).str.strip().unique()
        all_acks = [a for a in all_acks if a and a.lower() not in ('nan', 'none', '')]
        
        if not all_acks:
            all_acks = ['UNKNOWN']

        all_transaction_summary = []

        for current_ack in sorted(all_acks):
            if current_ack != 'UNKNOWN':
                ack_mask = df_main.iloc[:, 1].astype(str).str.strip() == current_ack
                df_ack_main = df_main[ack_mask]
            else:
                df_ack_main = df_main
                
            if df_ack_main.empty:
                continue

            credit_counting_df, _ = strict_deduplicate_main_transactions(
                df_ack_main
            )
            credit_counting_tid_series = credit_counting_df.iloc[:, 9].apply(
                _identity_text
            )
            duplicate_details_by_transaction_id = {}
            for detail in credit_counting_df.attrs.get(
                'duplicate_transaction_details', []
            ):
                duplicate_details_by_transaction_id.setdefault(
                    detail['credited_transaction_id'], []
                ).append(format_duplicate_credit_note(detail))
            
            # Collect all unique credited transaction IDs (column 9 only)
            all_trans_ids = set()
            credited_trans_ids = (
                credit_counting_df.iloc[:, 9].astype(str).str.strip()
            )
            all_trans_ids.update([tid for tid in credited_trans_ids if tid and tid.lower() not in ('nan', 'none', '', 'unknown', '-')])

            for trans_id in sorted(all_trans_ids):
                # Find all rows where this transaction ID appears as credited (column 9)
                credited_rows = credit_counting_df[
                    credit_counting_df.iloc[:, 9].astype(str).str.strip()
                    == trans_id
                ]

                # Collect credited account numbers (column 6) and bank names (column 4)
                credited_account_numbers = set()
                bank_names = set()
                debited_trans_ids = set()
                
                for _, row in credited_rows.iterrows():
                    # Credited account (column 6)
                    acc = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
                    if acc and acc.lower() not in ('nan', 'none', ''):
                        credited_account_numbers.add(acc)
                    
                    # Bank name (column 4)
                    bank = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
                    if bank and bank.lower() not in ('nan', 'none', ''):
                        bank_names.add(bank)
                    
                    # Debited transaction ID (column 3)
                    deb_tid = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                    if deb_tid and deb_tid.lower() not in ('nan', 'none', '', 'unknown', '-'):
                        debited_trans_ids.add(deb_tid)

                account_numbers_str = "; ".join(sorted(credited_account_numbers)) if credited_account_numbers else "None"
                bank_name_str = "; ".join(sorted(bank_names)) if bank_names else "N/A"
                debited_trans_id_str = "; ".join(sorted(debited_trans_ids)) if debited_trans_ids else "None"

                # Duplicate identities change only the credited total.  Debit
                # and other-sheet matching continue to use every source row.
                transaction_key = _identity_text(trans_id)
                credited_rows_for_total = credit_counting_df[
                    credit_counting_tid_series == transaction_key
                ]
                total_credited = sum(
                    clean_amount(row.iloc[11])
                    for _, row in credited_rows_for_total.iterrows()
                )
                duplicate_details = duplicate_details_by_transaction_id.get(
                    transaction_key, []
                )

                # Calculate outgoing/debited amount for the current credited transaction ID.
                # This must come from child rows where the current transaction ID is used as
                # the source/debited transaction ID, not from the source ID on the same row.
                child_rows = df_ack_main[
                    (df_ack_main.iloc[:, 3].astype(str).str.strip() == trans_id) &
                    (df_ack_main.iloc[:, 9].astype(str).str.strip() != trans_id)
                ]
                total_debited = sum(clean_amount(row.iloc[11]) for _, row in child_rows.iterrows())

                # Updated Amount & Breakdown
                total_updated = 0
                breakdown_by_sheet = {}
                
                # Get breakdown by transaction ID and all associated account numbers
                breakdown_items = []
                breakdown_items.extend(get_transaction_breakdown(trans_id))
                for acc_no in credited_account_numbers:
                    breakdown_items.extend(get_transaction_breakdown(None, acc_no))
                
                # Deduplicate breakdown items by (sheet, amount)
                seen_breakdown = set()
                for item in breakdown_items:
                    key = (item['sheet'], item['amount'])
                    if key not in seen_breakdown:
                        seen_breakdown.add(key)
                        sheet = item['sheet']
                        amount = item['amount']
                        total_updated += amount
                        breakdown_by_sheet[sheet] = breakdown_by_sheet.get(sheet, 0) + amount

                breakdown_text = "; ".join([f"{sheet}: ₹{amt:,.2f}" for sheet, amt in breakdown_by_sheet.items()]) if breakdown_by_sheet else "None"

                not_updated_amount = max(0, total_credited - total_debited - total_updated)

                # Status Logic: credited - (debited + updated) > 0 means PARTIAL
                if total_credited == 0 and total_debited > 0:
                    summary_status = 'COMPLETED'
                elif total_debited == 0 and total_updated == 0:
                    summary_status = 'PENDING'
                elif (total_credited - (total_debited + total_updated)) > 0.01:  # Allow small rounding errors
                    summary_status = 'PARTIAL'
                else:
                    summary_status = 'COMPLETED'

                duplicate_info_str = " | ".join(duplicate_details) if duplicate_details else "None"

                all_transaction_summary.append({
                    'Acknowledgement No': current_ack,
                    'Bank Name': bank_name_str,
                    'Account Number': account_numbers_str,
                    'Credited Transaction ID': trans_id,
                    'Debited Transaction ID': debited_trans_id_str,
                    'Total Credited Amount': total_credited,
                    'Total Debited Amount': total_debited,
                    'Updated Amount (Recovery)': total_updated,
                    'Not Updated Amount': not_updated_amount,
                    'Status': summary_status,
                    'Found in Other Sheets': 'Yes' if total_updated > 0 else 'No',
                    'Breakdown by Sheet': breakdown_text,
                    'Duplicate Entry Info': duplicate_info_str
                })

        df_trans = pd.DataFrame(all_transaction_summary)
        
        if len(df_trans) == 0:
            return jsonify({'success': False, 'message': 'No transaction data found'})
            
        # Sort by Not Updated descending
        df_trans.sort_values(by='Not Updated Amount', ascending=False, inplace=True)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_trans.to_excel(writer, sheet_name='Transaction ID Wise Summary', index=False)
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # --- STYLING LOGIC ---
            header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            # More vibrant colors for status
            pending_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')  # Bright red
            partial_fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')  # Bright orange
            completed_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # Bright green
            
            stripe_fill = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
            )

            worksheet = writer.sheets['Transaction ID Wise Summary']
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Header Styling
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Column Widths & Currency/Status Cols Mapping
            col_widths = {'A': 35, 'B': 30, 'C': 25, 'D': 22, 'E': 22, 'F': 22, 'G': 22, 'H': 22, 'I': 22, 'J': 15, 'K': 18, 'L': 60, 'M': 60}
            amt_cols = [6, 7, 8, 9]; status_col = 10; yes_no_col = 11
            
            for col, width in col_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Data Styling
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df_trans)+1), start=2):
                is_stripe = row_idx % 2 == 0
                
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    if is_stripe: cell.fill = stripe_fill
                    
                    if cell.column == status_col:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        status_text = str(cell.value).upper() if cell.value else ""
                        if "PENDING" in status_text:
                            cell.fill = pending_fill
                            cell.font = Font(color='C62828', bold=True, size=11)  # Dark red text
                        elif "PARTIAL" in status_text:
                            cell.fill = partial_fill
                            cell.font = Font(color='E65100', bold=True, size=11)  # Dark orange text
                        elif "COMPLETED" in status_text or "COMPLETE" in status_text:
                            cell.fill = completed_fill
                            cell.font = Font(color='2E7D32', bold=True, size=11)  # Dark green text
                    
                    if cell.column in amt_cols:
                        cell.number_format = '₹#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    if cell.column == yes_no_col:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
        output.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Transaction_ID_Summary_Report_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@app.route('/flow-diagram')
def flow_diagram_page():
    """Render the flow diagram page"""
    return render_template('flow_diagram.html')


@app.route('/api/get_layer1_transactions')
def get_layer1_transactions():
    """Get all Layer 1 transactions for dropdown"""
    try:
        if df_main is None:
            return jsonify({'transactions': []})
        
        layer1_txns = []
        for idx, row in df_main.iterrows():
            if row.iloc[5] == 1:  # Layer column
                # FILTER: Skip rows with null/empty credited transaction ID
                credited_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
                if not credited_trans_id or credited_trans_id.lower() in ('nan', 'none', '', '-', 'null'):
                    continue
                    
                trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                account = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                amount = clean_amount(row.iloc[11])
                
                # Get bank name from column 4
                bank = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else 'Unknown Bank'
                
                if trans_id:
                    layer1_txns.append({
                        'trans_id': trans_id,
                        'account': account,
                        'bank': bank,
                        'amount': amount
                    })
        
        return jsonify({'transactions': layer1_txns})
    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/api/get_all_layer1_flows')
def get_all_layer1_flows():
    """Get flow diagrams for ALL Layer 1 transactions in one view"""
    try:
        if df_main is None:
            return jsonify({'error': 'No data loaded'})
        
        # Get all Layer 1 transactions
        layer1_flows = []
        
        for idx, row in df_main.iterrows():
            if row.iloc[5] == 1:  # Layer column
                # FILTER: Skip rows with null/empty credited transaction ID
                credited_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
                if not credited_trans_id or credited_trans_id.lower() in ('nan', 'none', '', '-', 'null'):
                    continue
                    
                trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                if not trans_id:
                    continue
                
                # Get basic info
                account = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                bank = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else 'Unknown Bank'
                amount = clean_amount(row.iloc[11])
                ack_no = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
                date = str(row.iloc[10]).strip() if pd.notna(row.iloc[10]) else ''
                
                # Build tree structure
                tree = build_tree_from_transaction(idx, row)
                
                # Organize by layers
                layers_data = organize_tree_by_layers(tree)
                
                # Get cash-out data
                cashout_data = get_cashout_data(trans_id)
                
                layer1_flows.append({
                    'trans_id': trans_id,
                    'account': account,
                    'bank': bank,
                    'amount': amount,
                    'ack_no': ack_no,
                    'date': date,
                    'layers': layers_data,
                    'cashout': cashout_data,
                    'max_layer': max(layer['layer'] for layer in layers_data) if layers_data else 1
                })
        
        return jsonify({
            'flows': layer1_flows,
            'total_count': len(layer1_flows)
        })
    except Exception as e:
        import traceback
        print(f"Error in get_all_layer1_flows: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)})


@app.route('/api/get_all_transactions_tree')
def get_all_transactions_tree():
    """Get all transactions organized by layer for tree view"""
    try:
        if df_main is None:
            return jsonify({'error': 'No data loaded'})
        
        transactions = []
        total_amount = 0
        max_layer = 0
        cash_out = 0
        
        # Get acknowledgement and date from first row
        ack_no = ''
        date = ''
        if len(df_main) > 0:
            ack_no = str(df_main.iloc[0, 1]).strip() if pd.notna(df_main.iloc[0, 1]) else ''
            date = str(df_main.iloc[0, 10]).strip() if pd.notna(df_main.iloc[0, 10]) else ''
        
        for idx, row in df_main.iterrows():
            # FILTER: Skip rows with null/empty credited transaction ID
            credited_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
            if not credited_trans_id or credited_trans_id.lower() in ('nan', 'none', '', '-', 'null'):
                continue
                
            trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
            account = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
            bank = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else 'Unknown Bank'
            amount = clean_amount(row.iloc[11])
            layer = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 0
            
            if not trans_id:
                continue
            
            # Format account display
            if account and len(account) > 8:
                account_display = account[:4] + '...' + account[-4:]
            else:
                account_display = account
            
            transactions.append({
                'trans_id': trans_id,
                'account': account,
                'account_display': account_display,
                'bank': bank,
                'amount': amount,
                'layer': layer
            })
            
            total_amount += amount
            if layer > max_layer:
                max_layer = layer
        
        # Calculate cash-out from other sheets
        if df_other_sheets:
            for sheet_name, sheet_info in df_other_sheets.items():
                df = sheet_info['data']
                amount_col = sheet_info['amount_col']
                sheet_lower = sheet_name.lower()
                
                # Only count cash-out sheets
                if any(keyword in sheet_lower for keyword in ['atm', 'cheque', 'pos', 'hold', 'frozen']):
                    for idx, row in df.iterrows():
                        if len(row) > amount_col:
                            amt = clean_amount(row.iloc[amount_col])
                            cash_out += amt
        
        return jsonify({
            'transactions': transactions,
            'total_amount': total_amount,
            'total_layers': max_layer,
            'total_transactions': len(transactions),
            'cash_out': cash_out,
            'ack_no': ack_no,
            'date': date
        })
    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/api/get_tree_data/<trans_id>')
def get_tree_data(trans_id):
    """Generate tree data for a specific Layer 1 transaction with full hierarchy"""
    try:
        if df_main is None:
            return jsonify({'error': 'No data loaded'})
        
        # Find the Layer 1 transaction
        layer1_row = None
        for idx, row in df_main.iterrows():
            if str(row.iloc[3]).strip() == trans_id and row.iloc[5] == 1:
                layer1_row = row
                break
        
        if layer1_row is None:
            return jsonify({'error': 'Transaction not found'})
        
        # Get acknowledgement and date
        ack_no = str(layer1_row.iloc[1]).strip() if pd.notna(layer1_row.iloc[1]) else ''
        date = str(layer1_row.iloc[10]).strip() if pd.notna(layer1_row.iloc[10]) else ''
        total_disputed = clean_amount(layer1_row.iloc[11])
        
        # Build complete tree with all details
        root_node = build_complete_tree(trans_id, layer1_row)
        
        # Count total transactions and layers
        stats = count_tree_stats(root_node)
        
        # Get cash-out data
        cashout_data = get_cashout_data(trans_id)
        total_cashout = sum(group['total'] for group in cashout_data.values())
        
        return jsonify({
            'trans_id': trans_id,
            'ack_no': ack_no,
            'date': date,
            'total_disputed': total_disputed,
            'total_layers': stats['max_layer'],
            'total_transactions': stats['total_count'],
            'total_cashout': total_cashout,
            'root': root_node,
            'cashout': cashout_data
        })
    except Exception as e:
        return jsonify({'error': str(e)})


def build_complete_tree(trans_id, row):
    """Build complete tree node with all children recursively"""
    # Get basic info
    debited_account = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
    debited_bank = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else 'Unknown Bank'
    credited_account = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
    credited_bank = str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else 'Unknown Bank'
    amount = clean_amount(row.iloc[11])
    layer = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 1
    
    # Format account display
    if debited_account and len(debited_account) > 8:
        account_display = debited_account[:4] + '...' + debited_account[-4:]
    else:
        account_display = debited_account
    
    # Calculate status
    status_info = calculate_status(trans_id, amount, layer, debited_account)
    
    # Build node
    node = {
        'trans_id': trans_id,
        'account': account_display,
        'full_account': debited_account,
        'bank': debited_bank,
        'credited_account': credited_account,
        'credited_bank': credited_bank,
        'amount': amount,
        'layer': layer,
        'status': status_info['status'],
        'updated_amount': status_info['updated_amount'],
        'pending_amount': status_info['pending_amount'],
        'children_total': status_info['children_total'],
        'children': []
    }
    
    # Find all children (where this trans_id is the debited trans_id in next layer)
    if trans_id in debited_trans_id_map:
        for idx in debited_trans_id_map[trans_id]:
            child_row = df_main.iloc[idx]
            child_layer = int(child_row.iloc[5]) if pd.notna(child_row.iloc[5]) else 0
            
            # Only include if it's the next layer
            if child_layer == layer + 1:
                child_trans_id = str(child_row.iloc[3]).strip() if pd.notna(child_row.iloc[3]) else ''
                if child_trans_id:
                    # Recursively build child tree
                    child_node = build_complete_tree(child_trans_id, child_row)
                    node['children'].append(child_node)
    
    return node


def count_tree_stats(node):
    """Count total transactions and max layer in tree"""
    stats = {
        'total_count': 1,
        'max_layer': node['layer']
    }
    
    for child in node.get('children', []):
        child_stats = count_tree_stats(child)
        stats['total_count'] += child_stats['total_count']
        stats['max_layer'] = max(stats['max_layer'], child_stats['max_layer'])
    
    return stats


@app.route('/api/get_flow_diagram/<trans_id>')
def get_flow_diagram(trans_id):
    """Generate flow diagram data for a specific Layer 1 transaction"""
    try:
        if df_main is None:
            return jsonify({'error': 'No data loaded'})
        
        # Find the Layer 1 transaction
        layer1_row = None
        layer1_idx = None
        for idx, row in df_main.iterrows():
            if str(row.iloc[3]).strip() == trans_id and row.iloc[5] == 1:
                layer1_row = row
                layer1_idx = idx
                break
        
        if layer1_row is None:
            return jsonify({'error': 'Transaction not found'})
        
        # Get acknowledgement
        ack_no = str(layer1_row.iloc[1]).strip() if pd.notna(layer1_row.iloc[1]) else ''
        total_disputed = clean_amount(layer1_row.iloc[11])
        
        # Build tree structure from this transaction
        tree = build_tree_from_transaction(layer1_idx, layer1_row)
        
        # Organize by layers
        layers_data = organize_tree_by_layers(tree)
        
        # Get cash-out data
        cashout_data = get_cashout_data(trans_id)
        
        return jsonify({
            'trans_id': trans_id,
            'ack_no': ack_no,
            'total_disputed': total_disputed,
            'layers': layers_data,
            'cashout': cashout_data,
            'max_layer': max(layer['layer'] for layer in layers_data) if layers_data else 1
        })
    except Exception as e:
        import traceback
        print(f"Error in get_flow_diagram: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)})


def build_tree_from_transaction(start_idx, start_row, visited=None):
    """Build a tree structure starting from a specific transaction"""
    if visited is None:
        visited = set()
    
    if start_idx in visited:
        return None
    
    visited.add(start_idx)
    
    # Get current node data
    debited_account = str(start_row.iloc[2]).strip() if pd.notna(start_row.iloc[2]) else ''
    debited_trans_id = str(start_row.iloc[3]).strip() if pd.notna(start_row.iloc[3]) else ''
    debited_bank = str(start_row.iloc[4]).strip() if pd.notna(start_row.iloc[4]) else 'Unknown Bank'
    layer = int(start_row.iloc[5]) if pd.notna(start_row.iloc[5]) else 1
    credited_account = str(start_row.iloc[6]).strip() if pd.notna(start_row.iloc[6]) else ''
    credited_trans_id = str(start_row.iloc[9]).strip() if pd.notna(start_row.iloc[9]) else ''
    amount = clean_amount(start_row.iloc[11])
    
    node = {
        'debited_account': debited_account,
        'debited_trans_id': debited_trans_id,
        'debited_bank': debited_bank,
        'credited_account': credited_account,
        'credited_trans_id': credited_trans_id,
        'layer': layer,
        'amount': amount,
        'children': []
    }
    
    # Find children: rows where debited_trans_id matches this row's credited_trans_id
    if credited_trans_id and credited_trans_id in debited_trans_id_map:
        for child_idx in debited_trans_id_map[credited_trans_id]:
            if child_idx not in visited:
                child_row = df_main.iloc[child_idx]
                child_node = build_tree_from_transaction(child_idx, child_row, visited)
                if child_node:
                    node['children'].append(child_node)
    
    return node


def organize_tree_by_layers(tree):
    """Organize tree structure into layers"""
    if not tree:
        return []
    
    layers = {}
    
    def traverse(node):
        layer = node.get('layer', 1)
        if layer not in layers:
            layers[layer] = []
        
        # Determine node type
        node_type = 'victim' if layer == 1 else 'intermediate'
        
        # Get data
        bank = node.get('debited_bank', 'Unknown Bank')
        account = node.get('debited_account', '')
        amount = node.get('amount', 0)
        trans_id = node.get('debited_trans_id', '')
        credited_trans_id = node.get('credited_trans_id', '')
        
        # Calculate status for this node - use credited_trans_id to check for children
        status_info = calculate_status(credited_trans_id, amount, layer, account)
        status = status_info['status']
        
        layers[layer].append({
            'trans_id': trans_id,
            'account': str(account),  # Full account number
            'bank': bank,
            'amount': float(amount),
            'type': node_type,
            'status': status,
            'is_collapsed': False
        })
        
        # Traverse children
        for child in node.get('children', []):
            traverse(child)
    
    traverse(tree)
    
    # Format layers with descriptions and collapse logic
    formatted_layers = []
    for layer_num in sorted(layers.keys()):
        nodes = layers[layer_num]
        
        # Collapse if more than 5 nodes
        if len(nodes) > 5:
            visible_nodes = nodes[:3]
            collapsed_count = len(nodes) - 3
            
            # Create preview text
            preview_banks = list(set([n['bank'] for n in nodes[3:6]]))
            preview_text = ', '.join(preview_banks[:3])
            if len(preview_banks) > 3:
                preview_text += '...'
            
            visible_nodes.append({
                'is_collapsed': True,
                'collapsed_count': collapsed_count,
                'collapsed_label': 'accounts' if collapsed_count > 1 else 'account',
                'collapsed_preview': preview_text,
                'collapse_id': f'layer_{layer_num}',
                'collapsed_items': nodes[3:],
                'type': 'intermediate'
            })
            nodes = visible_nodes
        
        # Generate description
        total_amount = sum(n.get('amount', 0) for n in layers[layer_num])
        count = len(layers[layer_num])
        
        if layer_num == 1:
            description = 'seed'
        elif layer_num == 2:
            if count > 1:
                description = f'primary hub ({count} txns · ₹{total_amount/10000000:.2f} Cr disputed)'
            else:
                description = f'primary hub (1 txn · ₹{total_amount/10000000:.2f} Cr disputed)'
        else:
            description = f'{count} txns · ₹{total_amount/100000:.2f}L disputed'
        
        formatted_layers.append({
            'layer': layer_num,
            'description': description,
            'nodes': nodes
        })
    
    return formatted_layers




def get_cashout_data(root_trans_id):
    """Get cash-out data grouped by type"""
    cashout = {
        'atm': {'title': 'ATM withdrawals', 'items': [], 'count': 0, 'total': 0, 'subtitle': ''},
        'cheque': {'title': 'Cheque withdrawals', 'items': [], 'count': 0, 'total': 0, 'subtitle': ''},
        'pos': {'title': 'POS withdrawals', 'items': [], 'count': 0, 'total': 0, 'subtitle': ''},
        'frozen': {'title': 'Transactions frozen / put on hold', 'items': [], 'count': 0, 'total': 0, 'subtitle': ''},
        'others': {'title': 'Others Less Than 500', 'items': [], 'count': 0, 'total': 0, 'subtitle': ''}
    }
    
    if not df_other_sheets:
        return cashout
    
    # Collect all transaction IDs and accounts in the tree
    all_trans_ids = set()
    all_accounts = set()
    
    def collect_ids(node):
        trans_id = node.get('debited_trans_id', '')
        account = node.get('debited_account', '')
        credited_trans_id = node.get('credited_trans_id', '')
        credited_account = node.get('credited_account', '')
        
        if trans_id:
            all_trans_ids.add(trans_id)
        if credited_trans_id:
            all_trans_ids.add(credited_trans_id)
        if account:
            all_accounts.add(account)
        if credited_account:
            all_accounts.add(credited_account)
            
        for child in node.get('children', []):
            collect_ids(child)
    
    # Find the Layer 1 transaction and build tree
    layer1_row = None
    layer1_idx = None
    for idx, row in df_main.iterrows():
        # FILTER: Skip rows with null/empty credited transaction ID
        credited_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
        if not credited_trans_id or credited_trans_id.lower() in ('nan', 'none', '', '-', 'null'):
            continue
            
        if str(row.iloc[3]).strip() == root_trans_id and row.iloc[5] == 1:
            layer1_row = row
            layer1_idx = idx
            break
    
    if layer1_row is not None:
        tree = build_tree_from_transaction(layer1_idx, layer1_row)
        if tree:
            collect_ids(tree)
    
    # Process each sheet
    for sheet_name, sheet_info in df_other_sheets.items():
        df = sheet_info['data']
        amount_col = sheet_info['amount_col']
        sheet_lower = sheet_name.lower()
        
        # Determine category
        category = None
        if 'withdrawal through atm' in sheet_lower:
            category = 'atm'
        elif 'cheque' in sheet_lower:
            category = 'cheque'
        elif 'withdrawal through pos' in sheet_lower:
            category = 'pos'
        elif 'put on hold' in sheet_lower or 'frozen' in sheet_lower:
            category = 'frozen'
        elif 'others less' in sheet_lower:
            category = 'others'
        
        if not category:
            continue
        
        # Collect items from this sheet - group by account number
        account_data = {}  # {account: {'bank': '', 'amounts': [], 'trans_ids': [], 'extras': []}}
        
        for idx, row in df.iterrows():
            trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) and len(row) > 3 else ''
            account = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) and len(row) > 2 else ''
            
            # Check if this transaction/account is in our tree
            if trans_id not in all_trans_ids and account not in all_accounts:
                continue
            
            if not account:
                continue
            
            # Get amount - use disputed amount for ATM and cheque
            if category == 'others':
                amount = 500.0
            elif category == 'atm':
                # For ATM, use Disputed Amount (column 6)
                amount = clean_amount(row.iloc[6]) if len(row) > 6 else 0
            else:
                amount = clean_amount(row.iloc[amount_col]) if len(row) > amount_col else 0
            
            if amount <= 0:
                continue
            
            # Get bank name based on sheet type
            bank = ''
            if category == 'atm':
                # For ATM, bank name is in column 11 (Action Taken By bank)
                if len(row) > 11:
                    bank = str(row.iloc[11]).strip() if pd.notna(row.iloc[11]) else ''
            elif category == 'frozen':
                # For frozen/hold, bank name is in column 6 (Action Taken By bank)
                if len(row) > 6:
                    bank = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
            elif category == 'others':
                # For others, bank name is in column 6 (Action Taken By bank)
                if len(row) > 6:
                    bank = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
            elif category == 'cheque':
                # For Cheque, try column 6 first (Action Taken By bank), then fallback
                if len(row) > 6:
                    bank = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
                if not bank and len(row) > 4:
                    bank = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
            elif category == 'pos':
                # For POS, try column 6 first (Action Taken By bank), then fallback
                if len(row) > 6:
                    bank = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
                if not bank and len(row) > 4:
                    bank = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
            
            if not bank:
                bank = 'Unknown Bank'
            
            # Get location/extra info based on category
            extra = ''
            if category == 'atm':
                # ATM: Show location from column 8
                if len(row) > 8:
                    location = str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else ''
                    if location:
                        extra = f"📍 {location}"
            elif category == 'pos':
                # POS: Show merchant name from column 8
                if len(row) > 8:
                    merchant = str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else ''
                    if merchant:
                        extra = f"🏪 {merchant}"
            
            # Group by account number
            if account not in account_data:
                account_data[account] = {
                    'bank': bank,
                    'amounts': [],
                    'trans_ids': [],
                    'extras': set(),
                    'count': 0,
                    'details': []  # Store individual transaction details
                }
            
            account_data[account]['amounts'].append(amount)
            if trans_id:
                account_data[account]['trans_ids'].append(trans_id)
            if extra:
                account_data[account]['extras'].add(extra)
            account_data[account]['count'] += 1
            
            # Store individual transaction details
            account_data[account]['details'].append({
                'trans_id': trans_id,
                'amount': amount,
                'extra': extra
            })
        
        # Convert grouped data to items
        for account, data in account_data.items():
            total_amount = sum(data['amounts'])
            
            # Determine node type
            node_type = f'cashout-{category}'
            if category == 'frozen':
                node_type = 'frozen'
            elif category == 'others':
                node_type = 'others'
            
            # Build extra info
            extra_info = ''
            if data['extras']:
                # If multiple locations/merchants, show count
                if len(data['extras']) > 1:
                    if category == 'atm':
                        extra_info = f"📍 {len(data['extras'])} locations"
                    elif category == 'pos':
                        extra_info = f"🏪 {len(data['extras'])} merchants"
                else:
                    extra_info = list(data['extras'])[0]
            
            # Add count if multiple entries
            if data['count'] > 1:
                if extra_info:
                    extra_info += f" · {data['count']} txns"
                else:
                    extra_info = f"{data['count']} transactions"
            
            cashout[category]['items'].append({
                'trans_id': data['trans_ids'][0] if data['trans_ids'] else '',
                'account': account,
                'bank': data['bank'],
                'amount': total_amount,
                'extra': extra_info,
                'type': node_type,
                'is_collapsed': False,
                'is_grouped': data['count'] > 1,
                'grouped_details': data['details'] if data['count'] > 1 else []
            })
            
            cashout[category]['count'] += data['count']
            cashout[category]['total'] += total_amount
    
    # Add subtitles and collapse logic
    for cat_key in cashout:
        cat = cashout[cat_key]
        items = cat['items']
        
        if len(items) == 0:
            continue
        
        # Generate subtitle
        if cat_key == 'atm':
            unique_atms = len(set(item.get('extra', '') for item in items if item.get('extra')))
            unique_accounts = len(set(item['account'] for item in items))
            cat['subtitle'] = f'{unique_atms} ATMs · {unique_accounts} accounts'
        elif cat_key == 'cheque':
            disputed = sum(item['amount'] for item in items)
            cat['subtitle'] = f'₹{disputed/100000:.2f}L disputed'
        elif cat_key == 'pos':
            unique_merchants = len(set(item.get('extra', '') for item in items if item.get('extra')))
            cat['subtitle'] = f'{unique_merchants} merchants'
        elif cat_key == 'frozen':
            cat['subtitle'] = f'₹{cat["total"]/100000:.2f}L secured'
        else:
            cat['subtitle'] = f'{cat["count"]} records'
        
        # Collapse if more than 3 items
        if len(items) > 3:
            visible = items[:2]
            collapsed_count = len(items) - 2
            
            preview_banks = list(set([item['bank'] for item in items[2:5]]))
            preview_text = ', '.join(preview_banks[:3])
            
            visible.append({
                'is_collapsed': True,
                'collapsed_count': collapsed_count,
                'collapsed_label': f'{cat_key.upper()} accounts',
                'collapsed_preview': preview_text,
                'collapse_id': f'cashout_{cat_key}',
                'collapsed_items': items[2:],
                'type': f'cashout-{cat_key}'
            })
            
            cat['items'] = visible
    
    return cashout


@app.route('/download_flow_diagram/<trans_id>')
def download_flow_diagram(trans_id):
    """Download flow diagram as standalone HTML file"""
    try:
        if df_main is None:
            return jsonify({'error': 'No data loaded'})
        
        # Get flow data
        response = get_flow_diagram(trans_id)
        data = response.get_json()
        
        if 'error' in data:
            return jsonify({'error': data['error']})
        
        # Generate standalone HTML
        html_content = generate_flow_diagram_html(data)
        
        # Create filename
        filename = f'Flow_Diagram_{trans_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.html'
        
        # Return as downloadable file
        output = io.BytesIO()
        output.write(html_content.encode('utf-8'))
        output.seek(0)
        
        return send_file(
            output,
            mimetype='text/html',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)})


def generate_flow_diagram_html(data):
    """Generate standalone HTML for flow diagram"""
    # Read the template
    with open('templates/flow_diagram.html', 'r', encoding='utf-8') as f:
        template = f.read()
    
    # Inject data as JavaScript
    data_json = json.dumps(data, indent=2)
    
    # Modify template to be standalone
    standalone_html = template.replace(
        'window.onload = function() {',
        f'''
        let flowData = {data_json};
        
        window.onload = function() {{
            // Hide controls for standalone version
            document.querySelector('.controls').style.display = 'none';
            
            // Render diagram immediately
            renderFlowDiagram(flowData);
            
            // Original onload code (disabled)
            return;
        '''
    )
    
    return standalone_html


if __name__ == '__main__':
    app.run(debug=False, port=5001, host='127.0.0.1')
